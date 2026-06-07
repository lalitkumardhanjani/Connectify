import os
import sys
import unittest
import openpyxl
from unittest.mock import MagicMock, patch

# Add workspace to path
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

from config.constants import SCRAPER_HEADERS
from config.user_profiles import get_selected_user_name, load_all_configs, save_all_configs

# We will run E2E against a temporary excel file E2E
TEST_TRACKER_FILE = os.path.abspath(os.path.join(os.path.dirname(__file__), "test_job_tracker.xlsx"))

# Mock path helper
def get_mock_job_tracker_file():
    return TEST_TRACKER_FILE

class MockWebElement:
    def __init__(self, urn, text):
        self.urn = urn
        self.text = text
    def get_attribute(self, attr):
        if attr == "data-urn":
            return self.urn
        return None

class TestE2EEmailScraperOutreach(unittest.TestCase):
    
    def setUp(self):
        # Clean up any leftover test file
        if os.path.exists(TEST_TRACKER_FILE):
            os.remove(TEST_TRACKER_FILE)
            
        # Back up active user config
        self.username = get_selected_user_name()
        self.orig_config = load_all_configs()
        
    def tearDown(self):
        # Restore active user config
        if self.orig_config:
            save_all_configs(self.orig_config)
        # Clean up E2E test file
        if os.path.exists(TEST_TRACKER_FILE):
            os.remove(TEST_TRACKER_FILE)

    @patch('core.storage.database.get_job_tracker_file', get_mock_job_tracker_file)
    @patch('config.settings.get_job_tracker_file', get_mock_job_tracker_file)
    @patch('config.settings.JOB_TRACKER_FILE', TEST_TRACKER_FILE)
    @patch('pipelines.email_outreach.pipeline.JOB_TRACKER_FILE', TEST_TRACKER_FILE)
    def test_end_to_end_scraper_and_outreach_workflow(self):
        # 1. Update config parameters under email_scraper for E2E validation
        config = load_all_configs()
        user_conf = config.get("users", {}).get(self.username, {})
        
        # Setup E2E configurations
        user_conf["email_scraper"] = {
            "interval": "15",
            "review_mode": True,
            "max_emails_per_run": "2",
            "keywords": ["DBA"],
            "excluded_keywords": ["Intern"],
            "email_subject": "DBA Job Opportunity",
            "email_template": "Hello {FIRST_NAME}, let's connect."
        }
        user_conf["profile"] = {
            "preferred_locations": "Bangalore"
        }
        save_all_configs(config)
        
        # 2. Mock Selenium Scraper driver interactions
        mock_driver = MagicMock()
        mock_driver.current_url = "https://www.linkedin.com/search/results/content/"
        
        # Setup LinkedIn feed post simulation
        # Post 1: Matches keyword "DBA", has email "recruiter@startup.com"
        # Post 2: Matches keyword "DBA", has email "hiring.manager@enterprise.com"
        # Post 3: Matches keyword "DBA" but ALSO matches exclusion "Intern", has email "intern@startup.com"
        post1 = MockWebElement("urn:li:activity:11111", "Hiring DBA! Apply at recruiter@startup.com")
        post2 = MockWebElement("urn:li:activity:22222", "Hiring DBA! Apply at hiring.manager@enterprise.com")
        post3 = MockWebElement("urn:li:activity:33333", "Hiring Intern DBA! Apply at intern@startup.com")
        
        # Initialize Scraper
        from pipelines.email_outreach.services.scraper import LinkedInScraper
        scraper = LinkedInScraper(mock_driver)
        
        # Mock scraper methods to bypass browser scrolling
        scraper.login = MagicMock(return_value=True)
        scraper.search_for_keyword = MagicMock(return_value=True)
        
        # Mock _find_post_containers to return E2E test posts
        scraper._find_post_containers = MagicMock(side_effect=[
            [post1, post2, post3],
            [] # Empty list to end E2E scroll loop on second iteration
        ])
        
        # Mock extract_post_data to extract text content
        def mock_extract(post_element):
            return {"content": post_element.text}
        scraper.extract_post_data = MagicMock(side_effect=mock_extract)
        
        # 3. Run Scraper Pipeline (Phase 1)
        from pipelines.email_outreach.pipeline import run_phase_one
        print("\n[E2E] Running Phase 1: Post scraping & Database updates...")
        run_phase_one(scraper)
        
        # 4. Database Validation (Inserts and Exclusion Match Checks)
        self.assertTrue(os.path.exists(TEST_TRACKER_FILE), "Scraper spreadsheet should have been created")
        
        wb = openpyxl.load_workbook(TEST_TRACKER_FILE)
        ws = wb.active
        
        # Verify Headers
        headers = [cell.value for cell in ws[1]]
        self.assertEqual(headers, SCRAPER_HEADERS, "Workbook must contain standard E2E scraper headers")
        
        # Retrieve collected data
        rows_data = []
        for row in range(2, ws.max_row + 1):
            row_vals = [ws.cell(row=row, column=c).value for c in range(1, len(SCRAPER_HEADERS) + 1)]
            rows_data.append(row_vals)
            
        print("[E2E] Scraped leads in Excel:", rows_data)
        
        # Assertions
        collected_emails = [r[1] for r in rows_data]
        self.assertIn("recruiter@startup.com", collected_emails, "recruiter@startup.com should be collected")
        self.assertIn("hiring.manager@enterprise.com", collected_emails, "hiring.manager@enterprise.com should be collected")
        self.assertNotIn("intern@startup.com", collected_emails, "intern@startup.com must be skipped due to exclusion keyword Intern")
        
        # 5. Run Quality Gate & Outreach pipeline (Phase 2)
        # Mock E2E Quality Gate triggers:
        # First email recruiter@startup.com -> skipped
        # Second email hiring.manager@enterprise.com -> sent
        def mock_send_email(driver, email, review_mode=None):
            if email == "recruiter@startup.com":
                return "skipped"
            elif email == "hiring.manager@enterprise.com":
                return True
            return False
            
        print("\n[E2E] Running Phase 2: Quality Gate email outreach sends...")
        from pipelines.email_outreach.pipeline import run_phase_two
        
        with patch('pipelines.email_outreach.pipeline.send_email_via_gmail', side_effect=mock_send_email):
            run_phase_two(scraper, review_mode=True)
            
        # 6. Database Validation (Status transitions)
        wb = openpyxl.load_workbook(TEST_TRACKER_FILE)
        ws = wb.active
        
        col_map = {cell.value: idx+1 for idx, cell in enumerate(ws[1])}
        email_col = col_map.get('Email')
        status_col = col_map.get('Status')
        
        final_statuses = {}
        for row in range(2, ws.max_row + 1):
            em = ws.cell(row=row, column=email_col).value
            st = ws.cell(row=row, column=status_col).value
            final_statuses[em] = st
            
        print("[E2E] Final email statuses in Excel:", final_statuses)
        self.assertEqual(final_statuses["recruiter@startup.com"], "skipped", "First email should transition to skipped")
        self.assertEqual(final_statuses["hiring.manager@enterprise.com"], "sent", "Second email should transition to sent")
        
        print("\n[E2E] END-TO-END WORKFLOW VALIDATION PASSED SUCCESSFULLY!")

if __name__ == "__main__":
    unittest.main()
