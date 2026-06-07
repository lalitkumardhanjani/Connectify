import os
import sys
import unittest
import openpyxl
from unittest.mock import MagicMock, patch

# Ensure workspace is in PYTHONPATH
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

from config.user_profiles import get_selected_user_name, load_all_configs, save_all_configs

TEST_LEADS_FILE = os.path.abspath(os.path.join(os.path.dirname(__file__), "test_job_leads.xlsx"))

def get_mock_job_leads_file():
    return TEST_LEADS_FILE

class TestE2EJobFinder(unittest.TestCase):
    
    def setUp(self):
        # Clean up any leftover test files
        if os.path.exists(TEST_LEADS_FILE):
            os.remove(TEST_LEADS_FILE)
            
        # Back up active user config
        self.username = get_selected_user_name()
        self.orig_config = load_all_configs()
        
    def tearDown(self):
        # Restore active user config
        if self.orig_config:
            save_all_configs(self.orig_config)
        # Clean up test files
        if os.path.exists(TEST_LEADS_FILE):
            os.remove(TEST_LEADS_FILE)

    @patch('core.storage.database.get_job_leads_file', get_mock_job_leads_file)
    @patch('config.settings.get_job_leads_file', get_mock_job_leads_file)
    @patch('config.settings.JOB_LEADS_FILE', TEST_LEADS_FILE)
    @patch('pipelines.linkedin_outreach.services.job_finder.get_driver')
    @patch('pipelines.linkedin_outreach.services.job_finder.wait_until_logged_in', return_value=True)
    def test_job_finder_sequential_iteration_and_apply(self, mock_wait_login, mock_get_driver):
        # 1. Initialize E2E excel worksheets
        from core.storage.database import init_job_leads_store
        init_job_leads_store()
        
        # Setup E2E settings configuration
        config = load_all_configs()
        user_conf = config.get("users", {}).get(self.username, {})
        user_conf["linkedin_connect"] = {
            "interval": "15",
            "review_mode": True,
            "keywords": ["DBA"],
            "excluded_keywords": ["Oracle"]
        }
        user_conf["profile"] = {
            "preferred_locations": "Bangalore"
        }
        save_all_configs(config)
        
        # Mock Selenium Driver
        mock_driver = MagicMock()
        mock_get_driver.return_value = mock_driver
        
        # Setup Window Handles: main tab and new tab
        mock_driver.window_handles = ["main_tab"]
        mock_driver.current_window_handle = "main_tab"
        
        # When switching to the new tab, return the external URL
        mock_driver.current_url = "https://apple.com/jobs/apply/dba-1234"
        
        # Helper to generate mock card elements
        def make_mock_card(title_text, company_text, job_url):
            card = MagicMock()
            
            title_el = MagicMock()
            title_el.text = title_text
            
            company_el = MagicMock()
            company_el.text = company_text
            
            link_el = MagicMock()
            link_el.get_attribute.return_value = job_url
            
            def find_element_mock(by, value):
                if "span" in value or "view" in value or "jobs/view" in value:
                    return title_el
                return company_el
                
            def find_elements_mock(by, value):
                if "view" in value or "jobs/view" in value:
                    return [link_el]
                return []
                
            card.find_element.side_effect = find_element_mock
            card.find_elements.side_effect = find_elements_mock
            return card
            
        # Mock cards
        mock_card_apple = make_mock_card("DBA", "Apple", "https://linkedin.com/jobs/view/apple-dba-1")
        mock_card_ms = make_mock_card("DBA", "Microsoft", "https://linkedin.com/jobs/view/ms-dba-2")
        mock_card_google = make_mock_card("Oracle DBA", "Google", "https://linkedin.com/jobs/view/google-dba-3")
        mock_card_netflix = make_mock_card("Frontend Developer", "Netflix", "https://linkedin.com/jobs/view/netflix-fe-4")

        # Mock left pane container
        mock_left_pane = MagicMock()
        mock_left_pane.is_displayed.return_value = True

        # Mock right details pane container
        mock_right_pane = MagicMock()
        mock_right_pane.is_displayed.return_value = True

        # Setup find_element on driver
        def mock_driver_find_element(by, value):
            if any(details_css in value for details_css in ['.jobs-search__job-details', '.jobs-details', '#job-details', '.scaffold-layout__detail']):
                return mock_right_pane
            return mock_left_pane
            
        mock_driver.find_element.side_effect = mock_driver_find_element

        # Mock Apple and MS apply buttons
        mock_apply_btn_apple = MagicMock()
        mock_apply_btn_apple.text = "Apply on company website"
        
        mock_apply_btn_ms = MagicMock()
        mock_apply_btn_ms.text = "Easy Apply"

        # Toggle active card State
        current_card = [None]
        page_counter = [1]

        # Mock find_elements on left pane and right details pane
        def mock_left_pane_find_elements(by, value):
            if page_counter[0] == 1:
                return [mock_card_apple, mock_card_ms, mock_card_google, mock_card_netflix]
            return []
            
        mock_left_pane.find_elements.side_effect = mock_left_pane_find_elements

        def mock_right_pane_find_elements(by, value):
            if any(sel in value for sel in ['Apply', 'Easy', 'safety', 'apply-button']):
                if current_card[0] == 'apple':
                    return [mock_apply_btn_apple]
                elif current_card[0] == 'ms':
                    return [mock_apply_btn_ms]
            return []
            
        mock_right_pane.find_elements.side_effect = mock_right_pane_find_elements

        # Mock driver execute script clicks
        def mock_execute_script(script, *args):
            if args:
                el = args[0]
                if el == mock_card_apple:
                    current_card[0] = 'apple'
                elif el == mock_card_ms:
                    current_card[0] = 'ms'
                elif el == mock_card_google:
                    current_card[0] = 'google'
                elif el == mock_card_netflix:
                    current_card[0] = 'netflix'
                elif el == mock_apply_btn_apple:
                    mock_driver.window_handles = ["main_tab", "new_apply_tab"]
            return None

        mock_driver.execute_script.side_effect = mock_execute_script
        
        # Mock pagination page clicks
        def mock_next_page_func(driver_arg):
            if page_counter[0] == 1:
                page_counter[0] += 1
                return True
            return False

        # Apply patching
        with patch('pipelines.linkedin_outreach.services.job_finder.go_to_next_jobs_page', side_effect=mock_next_page_func):
             from pipelines.linkedin_outreach.services.job_finder import run_job_finder
             print("\n[E2E] Running Job Finder sequential iteration workflow...")
             run_job_finder()

        # 3. Verify Job Leads stored in Excel sheet
        wb = openpyxl.load_workbook(TEST_LEADS_FILE)
        ws_jobs = wb["Jobs"]
        
        job_records = []
        for r in range(2, ws_jobs.max_row + 1):
            job_records.append([ws_jobs.cell(row=r, column=c).value for c in (2, 3, 4, 7)])
            
        print("[E2E] Saved jobs in Excel database:", job_records)
        
        # Verify Apple job is stored as "Interested"
        self.assertEqual(len(job_records), 1, "Only Apple regular apply job should be saved in Excel database")
        self.assertEqual(job_records[0][0], "DBA")
        self.assertEqual(job_records[0][1], "Apple")
        self.assertEqual(job_records[0][2], "https://apple.com/jobs/apply/dba-1234")
        self.assertEqual(job_records[0][3], "NEW")
        
        print("\n[E2E] JOB FINDER SEQUENTIAL ITERATION TEST PASSED SUCCESSFULLY!")

if __name__ == "__main__":
    unittest.main()
