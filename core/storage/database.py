import os
import shutil
import openpyxl
from datetime import datetime
from openpyxl.worksheet.table import Table, TableStyleInfo
from config.settings import (
    BASE_DIR, get_data_dir, get_job_tracker_file, get_job_leads_file, get_referrals_file
)
from config.constants import SCRAPER_HEADERS, JOB_LEADS_HEADERS, REFERRAL_HEADERS
from core.logging.config import logger
from core.utils.url_utils import is_valid_external_url, normalize_external_url

# Cache set for fast duplicate checks of external apply URLs
seen_external_urls = set()

def migrate_old_data_files():
    """Migrates existing tracking databases from the root directory to the active user's data/ directory."""
    from config.settings import get_active_user
    active_user = get_active_user()
    if not active_user:
        return
        
    old_files = {
        "job_tracker.xlsx": get_job_tracker_file(),
        "LinkedIn_Job_Tracker.xlsx": get_job_leads_file()
    }
    for old_name, new_path in old_files.items():
        old_path = os.path.join(BASE_DIR, old_name)
        if os.path.exists(old_path) and not os.path.exists(new_path):
            try:
                shutil.move(old_path, new_path)
                logger.info(f"Migrated legacy file to new storage directory: {old_name} -> {new_path}")
            except Exception as e:
                logger.error(f"Failed to migrate legacy file {old_name}: {e}")

def migrate_referrals_data():
    """Migrates existing referrals worksheet data from LinkedIn_Job_Tracker.xlsx to referrals.xlsx for all users."""
    users_dir = os.path.join(BASE_DIR, "users")
    if not os.path.exists(users_dir):
        return
        
    for user_name in os.listdir(users_dir):
        user_path = os.path.join(users_dir, user_name)
        if not os.path.isdir(user_path) or user_name == "default":
            continue
            
        data_dir = os.path.join(user_path, "data")
        tracker_path = os.path.join(data_dir, "LinkedIn_Job_Tracker.xlsx")
        referrals_path = os.path.join(data_dir, "referrals.xlsx")
        
        if os.path.exists(tracker_path):
            try:
                wb_tracker = openpyxl.load_workbook(tracker_path)
                if "Referrals" in wb_tracker.sheetnames:
                    ws_src = wb_tracker["Referrals"]
                    
                    # Read rows from source
                    src_headers = [cell.value for cell in ws_src[1]]
                    if not src_headers or len(src_headers) == 0:
                        continue
                        
                    rows_data = []
                    for r in range(2, ws_src.max_row + 1):
                        row_dict = {}
                        for col_idx, h in enumerate(src_headers, start=1):
                            if h:
                                row_dict[h] = ws_src.cell(row=r, column=col_idx).value
                        if any(val is not None for val in row_dict.values()):
                            rows_data.append(row_dict)
                            
                    if rows_data:
                        logger.info(f"Found {len(rows_data)} referral records to migrate for user {user_name}.")
                        if os.path.exists(referrals_path):
                            wb_dest = openpyxl.load_workbook(referrals_path)
                        else:
                            wb_dest = openpyxl.Workbook()
                            
                        if "Referrals" not in wb_dest.sheetnames:
                            ws_dest = wb_dest.create_sheet(title="Referrals")
                        else:
                            ws_dest = wb_dest["Referrals"]
                            
                        if ws_dest.max_row <= 1:
                            ws_dest.append(REFERRAL_HEADERS)
                            
                        for row_dict in rows_data:
                            row_data = []
                            for header in REFERRAL_HEADERS:
                                row_data.append(row_dict.get(header))
                            ws_dest.append(row_data)
                            
                        ws_dest._tables.clear()
                        ref_range = f"A1:{chr(64 + len(REFERRAL_HEADERS))}{ws_dest.max_row}"
                        from openpyxl.worksheet.table import Table, TableStyleInfo
                        tab = Table(displayName="ReferralsTable", ref=ref_range)
                        style = TableStyleInfo(
                            name="TableStyleLight9",
                            showFirstColumn=False,
                            showLastColumn=False,
                            showRowStripes=True,
                            showColumnStripes=False,
                        )
                        tab.tableStyleInfo = style
                        ws_dest.add_table(tab)
                        
                        if "Sheet" in wb_dest.sheetnames:
                            del wb_dest["Sheet"]
                            
                        wb_dest.save(referrals_path)
                        logger.info(f"Successfully migrated referrals to {referrals_path}")
                        
                    # Delete worksheet
                    del wb_tracker["Referrals"]
                    wb_tracker.save(tracker_path)
                    logger.info(f"Removed 'Referrals' worksheet from {tracker_path}")
            except Exception as e:
                logger.error(f"Failed to migrate referrals for user {user_name}: {e}")

# Run data files migration automatically upon module import
migrate_old_data_files()
migrate_referrals_data()


def _trigger_mac_excel_reload(path):
    """Refreshes Microsoft Excel / Numbers document on macOS if open to show instantaneous updates."""
    try:
        import subprocess
        abs_path = os.path.abspath(path)
        filename = os.path.basename(path)
        
        reload_script = f'''\
        tell application "System Events"
            set excelRunning to (name of processes) contains "Microsoft Excel"
            set numbersRunning to (name of processes) contains "Numbers"
        end tell
        
        if excelRunning then
            tell application "Microsoft Excel"
                try
                    if exists workbook "{filename}" then
                        close workbook "{filename}" saving no
                        open POSIX file "{abs_path}"
                    end if
                end try
            end tell
        end if
        
        if numbersRunning then
            tell application "Numbers"
                try
                    set docName to "{filename}"
                    set docExists to false
                    repeat with doc in documents
                        if name of doc is docName then
                            set docExists to true
                            exit repeat
                        end if
                    end repeat
                    if docExists then
                        close document docName saving no
                        open POSIX file "{abs_path}"
                    end if
                end try
            end tell
        end if
        '''
        subprocess.run(["osascript", "-e", reload_script], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
    except Exception:
        pass


# =========================================================================
# 1. Email Scraper Database Operations (job_tracker.xlsx)
# =========================================================================

def init_scraper_store(path=None):
    """Initializes the scraper database Excel sheet if it does not exist."""
    if path is None:
        path = get_job_tracker_file()
    if not os.path.exists(path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Jobs"
        ws.append(SCRAPER_HEADERS)
        
        col_count = len(SCRAPER_HEADERS)
        tab = Table(displayName="JobTrackerTable", ref=f"A1:{chr(64+col_count)}1")
        style = TableStyleInfo(
            name="TableStyleLight9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        tab.tableStyleInfo = style
        ws.add_table(tab)
        wb.save(path)
        _trigger_mac_excel_reload(path)
        
    trim_scraper_excel_to_schema(path)
    migrate_pending_to_new(path)

def trim_scraper_excel_to_schema(path=None):
    """Enforces scraper Excel headers conform to standard schema."""
    if path is None:
        path = get_job_tracker_file()
    if not os.path.exists(path):
        return
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    
    current_headers = [cell.value for cell in ws[1]]
    needed = SCRAPER_HEADERS
    if current_headers == needed and len(ws._tables) > 0:
        return
    header_map = {h: i+1 for i, h in enumerate(current_headers) if h in needed}
    
    data_rows = []
    for row in range(1, ws.max_row + 1):
        row_data = {}
        for h in needed:
            idx = header_map.get(h)
            if idx is not None and idx <= ws.max_column:
                row_data[h] = ws.cell(row=row, column=idx).value
            else:
                row_data[h] = h if row == 1 else None
        data_rows.append(row_data)
        
    ws.delete_rows(1, ws.max_row)
    
    for r_idx, row_dict in enumerate(data_rows, start=1):
        for c_idx, h in enumerate(needed, start=1):
            ws.cell(row=r_idx, column=c_idx, value=row_dict[h])
            
    max_col = ws.max_column
    if max_col > len(needed):
        ws.delete_cols(len(needed)+1, max_col - len(needed))
        
    ws._tables.clear()
    ref = f"A1:{chr(64+len(needed))}{ws.max_row}"
    tab = Table(displayName="JobTrackerTable", ref=ref)
    style = TableStyleInfo(
        name="TableStyleLight9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    tab.tableStyleInfo = style
    ws.add_table(tab)
    wb.save(path)
    _trigger_mac_excel_reload(path)

def append_email(email, keyword='', post_url='', company_name='', experience='', location='', path=None):
    """Appends extracted email to the job tracker if it's unique."""
    if path is None:
        path = get_job_tracker_file()
    init_scraper_store(path)
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    col_indices = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}
    id_col = col_indices.get('ID', 1)
    email_col = col_indices.get('Email', 2)
    
    normalized_email = str(email or '').strip().lower()
    for row in range(2, ws.max_row + 1):
        if str(ws.cell(row=row, column=email_col).value or '').strip().lower() == normalized_email:
            return False
            
    max_id = max([ws.cell(row=row, column=id_col).value or 0 for row in range(2, ws.max_row + 1)], default=0)
    new_id = int(max_id) + 1
    timestamp = datetime.utcnow().isoformat()
    # SCRAPER_HEADERS = ['ID', 'Email', 'Status', 'Timestamp', 'Keyword', 'PostURL', 'CompanyName', 'Experience', 'Location']
    new_row = [new_id, email, 'New', timestamp, keyword, post_url or '', company_name or '', experience or '', location or '']
    ws.append(new_row)
    
    ws._tables.clear()
    col_count = len(SCRAPER_HEADERS)
    ref = f"A1:{chr(64+col_count)}{ws.max_row}"
    tab = Table(displayName="JobTrackerTable", ref=ref)
    style = TableStyleInfo(
        name="TableStyleLight9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    tab.tableStyleInfo = style
    ws.add_table(tab)
    wb.save(path)
    _trigger_mac_excel_reload(path)
    return True

def update_status(email, status, path=None):
    """Updates status for a specific scraped email row."""
    if path is None:
        path = get_job_tracker_file()
    init_scraper_store(path)
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    col_indices = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}
    email_col = col_indices.get('Email', 2)
    status_col = col_indices.get('Status', 3)
    timestamp_col = col_indices.get('Timestamp', 4)
    normalized_email = str(email or '').strip().lower()
    updated = False
    
    for row in range(2, ws.max_row + 1):
        if str(ws.cell(row=row, column=email_col).value or '').strip().lower() == normalized_email:
            ws.cell(row=row, column=status_col, value=status)
            ws.cell(row=row, column=timestamp_col, value=datetime.utcnow().isoformat())
            updated = True
            break
            
    if updated:
        wb.save(path)
        _trigger_mac_excel_reload(path)
    return updated

def count_unique_emails(path=None):
    """Counts unique emails processed in the scraper sheet."""
    if path is None:
        path = get_job_tracker_file()
    if not os.path.exists(path):
        return 0
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    col_indices = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}
    email_col = col_indices.get('Email', 2)
    unique_emails = set()
    for row in range(2, ws.max_row + 1):
        val = ws.cell(row=row, column=email_col).value
        if val:
            unique_emails.add(str(val).strip().lower())
    return len(unique_emails)

def edit_row(row_id, email, status, keyword, post_url=None, company_name=None, experience=None, location=None, path=None):
    """Modifies details of an existing scraper log by row ID."""
    if path is None:
        path = get_job_tracker_file()
    init_scraper_store(path)
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    col_indices = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}
    id_col = col_indices.get('ID', 1)
    email_col = col_indices.get('Email', 2)
    status_col = col_indices.get('Status', 3)
    keyword_col = col_indices.get('Keyword', 5)
    post_url_col = col_indices.get('PostURL')
    company_col = col_indices.get('CompanyName')
    experience_col = col_indices.get('Experience')
    location_col = col_indices.get('Location')
    
    updated = False
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=id_col).value == row_id:
            if email_col:
                ws.cell(row=row, column=email_col, value=email)
            if status_col:
                ws.cell(row=row, column=status_col, value=status)
            if keyword_col:
                ws.cell(row=row, column=keyword_col, value=keyword)
            if post_url_col and post_url is not None:
                ws.cell(row=row, column=post_url_col, value=post_url)
            if company_col and company_name is not None:
                ws.cell(row=row, column=company_col, value=company_name)
            if experience_col and experience is not None:
                ws.cell(row=row, column=experience_col, value=experience)
            if location_col and location is not None:
                ws.cell(row=row, column=location_col, value=location)
            updated = True
            break
            
    if updated:
        wb.save(path)
        _trigger_mac_excel_reload(path)
    return updated

def migrate_pending_to_new(path=None):
    """Legacy helper converting old 'Pending' labels to standard 'New'."""
    if path is None:
        path = get_job_tracker_file()
    if not os.path.exists(path):
        return
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    col_indices = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}
    status_col = col_indices.get('Status')
    if not status_col:
        return
    updated = False
    for row in range(2, ws.max_row + 1):
        val = ws.cell(row=row, column=status_col).value
        if val and str(val).strip().lower() == 'pending':
            ws.cell(row=row, column=status_col, value='New')
            updated = True
    if updated:
        wb.save(path)
        _trigger_mac_excel_reload(path)


# =========================================================================
# 2. LinkedIn Outreach Database Operations (LinkedIn_Job_Tracker.xlsx)
# =========================================================================

def init_job_leads_store(path=None):
    """Ensures the job search leads database exists and contains standard headers."""
    if path is None:
        path = get_job_leads_file()
    if not os.path.exists(path):
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Jobs"
            ws.append(JOB_LEADS_HEADERS)
            update_job_leads_table(ws)
            wb.save(path)
            logger.info(f"Initialized new Excel tracker: {path}")
        except Exception as e:
            logger.error(f"Unable to initialize Excel tracker: {str(e)}")
    else:
        try:
            trim_job_leads_excel_to_schema(path)
        except Exception as e:
            logger.error(f"Error trimming job leads Excel: {e}")

def trim_job_leads_excel_to_schema(path=None):
    """Enforces job leads Excel headers conform to standard schema."""
    if path is None:
        path = get_job_leads_file()
    if not os.path.exists(path):
        return
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    
    current_headers = [cell.value for cell in ws[1]]
    needed = JOB_LEADS_HEADERS
    if current_headers == needed and len(ws._tables) > 0:
        return
    header_map = {h: i+1 for i, h in enumerate(current_headers) if h in needed}
    
    data_rows = []
    for row in range(1, ws.max_row + 1):
        row_data = {}
        for h in needed:
            idx = header_map.get(h)
            if idx is not None and idx <= ws.max_column:
                row_data[h] = ws.cell(row=row, column=idx).value
            else:
                row_data[h] = h if row == 1 else None
        data_rows.append(row_data)
        
    ws.delete_rows(1, ws.max_row)
    
    for r_idx, row_dict in enumerate(data_rows, start=1):
        for c_idx, h in enumerate(needed, start=1):
            ws.cell(row=r_idx, column=c_idx, value=row_dict[h])
            
    max_col = ws.max_column
    if max_col > len(needed):
        ws.delete_cols(len(needed)+1, max_col - len(needed))
        
    ws._tables.clear()
    ref = f"A1:{chr(64+len(needed))}{ws.max_row}"
    tab = Table(displayName="JobTrackerTable", ref=ref)
    style = TableStyleInfo(
        name="TableStyleLight9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    tab.tableStyleInfo = style
    ws.add_table(tab)
    wb.save(path)
    _trigger_mac_excel_reload(path)

def update_job_leads_table(ws):
    """Updates formatting on the openpyxl job leads worksheet."""
    ws._tables.clear()
    max_row = max(ws.max_row, 1)
    col_letter = chr(64 + len(JOB_LEADS_HEADERS))
    ref_range = f"A1:{col_letter}{max_row}"
    tab = Table(displayName="JobTrackerTable", ref=ref_range)
    style = TableStyleInfo(
        name="TableStyleLight9", 
        showFirstColumn=False,
        showLastColumn=False, 
        showRowStripes=True, 
        showColumnStripes=False
    )
    tab.tableStyleInfo = style
    ws.add_table(tab)

def load_saved_jobs(path=None):
    """Loads all saved job applications into memory."""
    if path is None:
        path = get_job_leads_file()
    jobs = []
    normalized_urls = set()

    if os.path.exists(path):
        try:
            wb = openpyxl.load_workbook(path)
            ws = wb.active
            col_indices = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}
            
            company_url_col = col_indices.get('CompanyURL')
            company_col = col_indices.get('CompanyName')
            job_id_col = col_indices.get('JobID')
            keyword_col = col_indices.get('SearchKeyword')
            job_title_col = col_indices.get('JobTitle')

            if company_url_col:
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not any(cell is not None for cell in row):
                        continue
                    company_url = row[company_url_col - 1]
                    if company_url:
                        company_url_str = str(company_url).strip()
                        normalized_urls.add(normalize_external_url(company_url_str))
                        
                        jobs.append({
                            "url": company_url_str,
                            "company": str(row[company_col - 1]).strip() if company_col and row[company_col - 1] else "",
                            "job_id": int(row[job_id_col - 1]) if job_id_col and row[job_id_col - 1] else None,
                            "keyword": str(row[keyword_col - 1]).strip() if keyword_col and row[keyword_col - 1] else "",
                            "position": str(row[job_title_col - 1]).strip() if job_title_col and row[job_title_col - 1] else ""
                        })
        except Exception as e:
            logger.error(f"Error loading tracker Excel file: {str(e)}")

    return jobs, normalized_urls

def save_job(data, path=None):
    """Saves a discovered external apply job to the Excel leads sheet."""
    if path is None:
        path = get_job_leads_file()
    url = (data.get("url") or "").strip()
    company = (data.get("company") or "").strip()
    search_keyword = (data.get("search_keyword") or "").strip()
    job_title = (data.get("position") or data.get("job_title") or "").strip()

    if not url or not is_valid_external_url(url):
        return False

    normalized_url = normalize_external_url(url)
    jobs, existing_urls = load_saved_jobs(path)

    if normalized_url in existing_urls:
        logger.info(f"Duplicate job URL skipped (normalized match): {url}")
        return False

    for existing_job in jobs:
        if existing_job.get("url") == url:
            logger.info(f"Duplicate job URL skipped (exact string match): {url}")
            return False

    try:
        init_job_leads_store(path)
        wb = openpyxl.load_workbook(path)
        ws = wb.active

        max_id = 0
        col_indices = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}
        job_id_col = col_indices.get('JobID', 1)

        for row in range(2, ws.max_row + 1):
            val = ws.cell(row=row, column=job_id_col).value
            if val is not None:
                try:
                    val_int = int(val)
                    if val_int > max_id:
                        max_id = val_int
                except ValueError:
                    pass

        new_job_id = max_id + 1
        created_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        row_data = []
        for header in JOB_LEADS_HEADERS:
            if header == 'JobID':
                row_data.append(new_job_id)
            elif header == 'JobTitle':
                row_data.append(job_title)
            elif header == 'CompanyName':
                row_data.append(company)
            elif header == 'LinkedIn_Company_URL':
                slug = company.lower().replace(" ", "-").replace(".", "").replace(",", "")
                row_data.append(f"https://www.linkedin.com/company/{slug}/")
            elif header == 'CompanyURL':
                row_data.append(url)
            elif header == 'ShortenURL':
                row_data.append('')
            elif header == 'SearchKeyword':
                row_data.append(search_keyword)
            elif header == 'Status':
                row_data.append('NEW')
            elif header == 'ShortUrlCreated':
                row_data.append('No')
            elif header == 'CreatedDateTime':
                row_data.append(created_time)

        ws.append(row_data)

        # Sort jobs by ID
        rows_to_sort = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(cell is not None for cell in row):
                rows_to_sort.append(list(row))

        job_id_idx = JOB_LEADS_HEADERS.index('JobID')
        rows_to_sort.sort(key=lambda r: int(r[job_id_idx]) if r[job_id_idx] is not None else 0)

        while ws.max_row > 1:
            ws.delete_rows(2)

        for r in rows_to_sort:
            ws.append(r)

        update_job_leads_table(ws)
        wb.save(path)
        _trigger_mac_excel_reload(path)

        seen_external_urls.add(normalized_url)
        logger.info(f"Saved & Sorted in Excel Tracker (JobID {new_job_id}): {url}")
        return True

    except Exception as e:
        logger.error(f"Error saving job to Excel tracker: {str(e)}")
        return False

def load_jobs_for_referral(path=None, status_filter='Asked for Referral'):
    """Loads all lead row dictionaries filtered by status.
    
    If status_filter is None, returns ALL rows regardless of status.
    """
    if path is None:
        path = get_job_leads_file()
    if not os.path.exists(path):
        return []
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    col_indices = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}
    rows = []
    
    for row in range(2, ws.max_row + 1):
        raw_status = ws.cell(row=row, column=col_indices.get('Status')).value
        if status_filter is None:
            # Return all rows
            row_dict = {header: ws.cell(row=row, column=col_idx).value for header, col_idx in col_indices.items()}
            if any(val is not None for val in row_dict.values()):
                rows.append(row_dict)
        else:
            status = str(raw_status).strip().lower() if raw_status is not None else ''
            if status == status_filter.strip().lower():
                row_dict = {header: ws.cell(row=row, column=col_idx).value for header, col_idx in col_indices.items()}
                rows.append(row_dict)
    return rows




def update_status_by_id(job_id, status, path=None):
    """Updates status for a specific lead row identified by JobID."""
    if path is None:
        path = get_job_leads_file()
    if not os.path.exists(path):
        return False
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    col_indices = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}
    id_col = col_indices.get('JobID')
    status_col = col_indices.get('Status')
    timestamp_col = col_indices.get('CreatedDateTime')
    
    if not id_col or not status_col:
        return False
    updated = False
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=id_col).value == job_id:
            ws.cell(row=row, column=status_col, value=status)
            if timestamp_col:
                ws.cell(row=row, column=timestamp_col, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
            updated = True
            break
            
    if updated:
        wb.save(path)
        _trigger_mac_excel_reload(path)
    return updated


def edit_lead_row(job_id, company, url, shorten, keyword, position, status, path=None):
    """Modifies details of an existing job lead by JobID."""
    if path is None:
        path = get_job_leads_file()
    init_job_leads_store(path)
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    col_indices = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}
    id_col = col_indices.get('JobID', 1)
    company_col = col_indices.get('CompanyName')
    url_col = col_indices.get('CompanyURL')
    shorten_col = col_indices.get('ShortenURL')
    keyword_col = col_indices.get('SearchKeyword')
    position_col = col_indices.get('JobTitle')
    status_col = col_indices.get('Status')
    short_url_created_col = col_indices.get('ShortUrlCreated')
    
    updated = False
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=id_col).value == job_id:
            if company_col and company is not None:
                ws.cell(row=row, column=company_col, value=company)
            if url_col and url is not None:
                ws.cell(row=row, column=url_col, value=url)
            if shorten_col and shorten is not None:
                ws.cell(row=row, column=shorten_col, value=shorten)
                if short_url_created_col:
                    if shorten and str(shorten).strip().startswith("http"):
                        ws.cell(row=row, column=short_url_created_col, value="Yes")
                    else:
                        ws.cell(row=row, column=short_url_created_col, value="No")
            if keyword_col and keyword is not None:
                ws.cell(row=row, column=keyword_col, value=keyword)
            if position_col and position is not None:
                ws.cell(row=row, column=position_col, value=position)
            if status_col and status is not None:
                ws.cell(row=row, column=status_col, value=status)
            updated = True
            break
            
    if updated:
        wb.save(path)
        _trigger_mac_excel_reload(path)
    return updated


# =========================================================================
# 3. Referral Outreach Database Operations (Referrals Sheet)
# =========================================================================

def init_referrals_store(path=None):
    """Initializes the referrals worksheet inside referrals.xlsx if it doesn't exist."""
    if path is None:
        path = get_referrals_file()
    
    os.makedirs(os.path.dirname(os.path.abspath(path)), exist_ok=True)
    
    if not os.path.exists(path):
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Referrals"
            ws.append(REFERRAL_HEADERS)
            
            ref_range = f"A1:{chr(64 + len(REFERRAL_HEADERS))}1"
            tab = Table(displayName="ReferralsTable", ref=ref_range)
            style = TableStyleInfo(
                name="TableStyleLight9",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            tab.tableStyleInfo = style
            ws.add_table(tab)
            wb.save(path)
            logger.info(f"Initialized new Referrals tracker: {path}")
        except Exception as e:
            logger.error(f"Unable to initialize Referrals tracker: {str(e)}")
    else:
        try:
            trim_referrals_excel_to_schema(path)
        except Exception as e:
            logger.error(f"Error trimming referrals Excel: {e}")

def trim_referrals_excel_to_schema(path=None):
    """Enforces Referrals Excel sheet headers conform to standard schema."""
    if path is None:
        path = get_referrals_file()
    if not os.path.exists(path):
        return
    wb = openpyxl.load_workbook(path)
    if "Referrals" not in wb.sheetnames:
        return
    ws = wb["Referrals"]
    
    current_headers = [cell.value for cell in ws[1]]
    needed = REFERRAL_HEADERS
    if current_headers == needed and len(ws._tables) > 0:
        return
    header_map = {h: i+1 for i, h in enumerate(current_headers) if h in needed}
    
    data_rows = []
    for row in range(1, ws.max_row + 1):
        row_data = {}
        for h in needed:
            idx = header_map.get(h)
            if idx is not None and idx <= ws.max_column:
                row_data[h] = ws.cell(row=row, column=idx).value
            else:
                row_data[h] = h if row == 1 else None
        data_rows.append(row_data)
        
    ws.delete_rows(1, ws.max_row)
    
    for r_idx, row_dict in enumerate(data_rows, start=1):
        for c_idx, h in enumerate(needed, start=1):
            ws.cell(row=r_idx, column=c_idx, value=row_dict[h])
            
    max_col = ws.max_column
    if max_col > len(needed):
        ws.delete_cols(len(needed)+1, max_col - len(needed))
        
    ws._tables.clear()
    ref = f"A1:{chr(64+len(needed))}{ws.max_row}"
    tab = Table(displayName="ReferralsTable", ref=ref)
    style = TableStyleInfo(
        name="TableStyleLight9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    tab.tableStyleInfo = style
    ws.add_table(tab)
    wb.save(path)
    _trigger_mac_excel_reload(path)

def load_all_referrals(path=None):
    """Loads all referral contact records from the Referrals sheet."""
    if path is None:
        path = get_referrals_file()
    init_referrals_store(path)
    wb = openpyxl.load_workbook(path)
    if "Referrals" not in wb.sheetnames:
        return []
    ws = wb["Referrals"]
    col_indices = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}
    
    rows = []
    for row in range(2, ws.max_row + 1):
        row_dict = {header: ws.cell(row=row, column=col_idx).value for header, col_idx in col_indices.items()}
        rows.append(row_dict)
    return rows

def add_or_update_referral(referral_data, path=None):
    """Adds a new referral contact or updates status/fields of an existing one by profile URL and job URL combination."""
    if path is None:
        path = get_referrals_file()
    init_referrals_store(path)
    wb = openpyxl.load_workbook(path)
    ws = wb["Referrals"]
    col_indices = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}
    
    # Backwards compatibility / translation
    if 'Job_URL' not in referral_data and 'Company_URL' in referral_data:
        referral_data['Job_URL'] = referral_data['Company_URL']
        
    profile_url = referral_data.get('Referral_Person_Profile_URL')
    job_url = referral_data.get('Job_URL')
    if not profile_url:
        return False
        
    url_col = col_indices.get('Referral_Person_Profile_URL')
    job_url_col = col_indices.get('Job_URL')
    
    existing_row = None
    if url_col and job_url_col:
        for r in range(2, ws.max_row + 1):
            cell_val = ws.cell(row=r, column=url_col).value
            cell_job = ws.cell(row=r, column=job_url_col).value
            prof_match = cell_val and str(cell_val).strip() == str(profile_url).strip()
            job_match = str(cell_job or '').strip() == str(job_url or '').strip()
            if prof_match and job_match:
                existing_row = r
                break
                
    if existing_row:
        for key, val in referral_data.items():
            col_idx = col_indices.get(key)
            if col_idx and key != 'ReferralID':
                ws.cell(row=existing_row, column=col_idx, value=val)
    else:
        id_col = col_indices.get('ReferralID', 1)
        max_id = 0
        for r in range(2, ws.max_row + 1):
            val = ws.cell(row=r, column=id_col).value
            if val is not None:
                try:
                    max_id = max(max_id, int(val))
                except ValueError:
                    pass
        new_id = max_id + 1
        
        row_data = []
        for header in REFERRAL_HEADERS:
            if header == 'ReferralID':
                row_data.append(new_id)
            else:
                row_data.append(referral_data.get(header))
        ws.append(row_data)
        
    wb.save(path)
    
    wb = openpyxl.load_workbook(path)
    ws = wb["Referrals"]
    ws._tables.clear()
    ref_range = f"A1:{chr(64 + len(REFERRAL_HEADERS))}{ws.max_row}"
    tab = Table(displayName="ReferralsTable", ref=ref_range)
    style = TableStyleInfo(
        name="TableStyleLight9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    tab.tableStyleInfo = style
    ws.add_table(tab)
    wb.save(path)
    _trigger_mac_excel_reload(path)
    try:
        sync_job_lead_referral_statuses()
    except Exception as e:
        logger.error(f"Error calling sync_job_lead_referral_statuses: {e}")
    return True

def is_profile_already_contacted(profile_url, job_url=None, path=None):
    """Checks if connection profile URL has been contacted or skipped before (optionally scoped to a specific Job_URL)."""
    if path is None:
        path = get_referrals_file()
    init_referrals_store(path)
    wb = openpyxl.load_workbook(path)
    ws = wb["Referrals"]
    col_indices = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}
    url_col = col_indices.get('Referral_Person_Profile_URL')
    status_col = col_indices.get('Referral_Status')
    job_url_col = col_indices.get('Job_URL')
    
    if not url_col:
        return False
        
    for row in range(2, ws.max_row + 1):
        cell_url = ws.cell(row=row, column=url_col).value
        if cell_url and str(cell_url).strip().lower() == str(profile_url).strip().lower():
            if job_url is not None and job_url_col:
                cell_job = ws.cell(row=row, column=job_url_col).value
                if str(cell_job or '').strip().lower() != str(job_url).strip().lower():
                    continue
            if status_col:
                status = str(ws.cell(row=row, column=status_col).value or "").strip().lower()
                if status in ('sent', 'replied', 'referral received', 'skipped'):
                    return True
    return False

def edit_referral_contact_row(referral_id, referral_data, path=None):
    """Modifies an existing referral contact record directly by ReferralID."""
    if path is None:
        path = get_referrals_file()
    init_referrals_store(path)
    wb = openpyxl.load_workbook(path)
    ws = wb["Referrals"]
    col_indices = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}
    id_col = col_indices.get('ReferralID')
    if not id_col:
        return False
        
    updated = False
    for row in range(2, ws.max_row + 1):
        val = ws.cell(row=row, column=id_col).value
        if val is not None and str(val).strip() == str(referral_id).strip():
            for key, value in referral_data.items():
                col_idx = col_indices.get(key)
                if col_idx and key != 'ReferralID':
                    ws.cell(row=row, column=col_idx, value=value)
            updated = True
            break
            
    if updated:
        wb.save(path)
        _trigger_mac_excel_reload(path)
    return updated

def get_company_sent_count(company_name, path=None):
    """Counts successfully sent referral/outreach messages or connection requests for a company."""
    if path is None:
        path = get_referrals_file()
    if not os.path.exists(path):
        return 0
    init_referrals_store(path)
    wb = openpyxl.load_workbook(path)
    if "Referrals" not in wb.sheetnames:
        return 0
    ws = wb["Referrals"]
    col_indices = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}
    company_col = col_indices.get('CompanyName')
    status_col = col_indices.get('Referral_Status')
    if not company_col or not status_col:
        return 0
    
    count = 0
    normalized_company = str(company_name or '').strip().lower()
    for row in range(2, ws.max_row + 1):
        row_company = str(ws.cell(row=row, column=company_col).value or '').strip().lower()
        if row_company == normalized_company:
            row_status = str(ws.cell(row=row, column=status_col).value or '').strip().lower()
            if row_status == 'sent':
                count += 1
    return count


def get_company_referrals_count(company_name, path=None):
    """Counts total existing referral contacts (regardless of status) in database for a company."""
    if path is None:
        path = get_referrals_file()
    if not os.path.exists(path):
        return 0
    init_referrals_store(path)
    wb = openpyxl.load_workbook(path)
    if "Referrals" not in wb.sheetnames:
        return 0
    ws = wb["Referrals"]
    col_indices = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}
    company_col = col_indices.get('CompanyName')
    if not company_col:
        return 0
    
    count = 0
    normalized_company = str(company_name or '').strip().lower()
    for row in range(2, ws.max_row + 1):
        row_company = str(ws.cell(row=row, column=company_col).value or '').strip().lower()
        if row_company == normalized_company:
            count += 1
    return count


def get_employee_outreach_progress(company_name, path=None):
    """Counts employee outreach progress towards target connection count.
    
    Includes existing employee connections and connection requests with status
    in ('Pending', 'Sent', 'Replied', 'Referral Received').
    """
    if path is None:
        path = get_referrals_file()
    if not os.path.exists(path):
        return 0
    init_referrals_store(path)
    wb = openpyxl.load_workbook(path)
    if "Referrals" not in wb.sheetnames:
        return 0
    ws = wb["Referrals"]
    col_indices = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}
    company_col = col_indices.get('CompanyName')
    status_col = col_indices.get('Referral_Status')
    source_col = col_indices.get('Referral_Source')
    if not company_col or not status_col or not source_col:
        return 0
        
    count = 0
    normalized_company = str(company_name or '').strip().lower()
    for row in range(2, ws.max_row + 1):
        row_company = str(ws.cell(row=row, column=company_col).value or '').strip().lower()
        if row_company == normalized_company:
            row_source = str(ws.cell(row=row, column=source_col).value or '').strip().lower()
            # Exclude recruiter source
            if 'recruiter' in row_source:
                continue
            row_status = str(ws.cell(row=row, column=status_col).value or '').strip().lower()
            if row_status in ('pending', 'sent', 'replied', 'referral received'):
                count += 1
    return count


def get_recruiter_outreach_progress(company_name, path=None):
    """Counts recruiter outreach progress towards recruiter target count.
    
    Includes existing recruiter connections and recruiter connection requests
    with status in ('Pending', 'Sent', 'Replied', 'Referral Received').
    """
    if path is None:
        path = get_referrals_file()
    if not os.path.exists(path):
        return 0
    init_referrals_store(path)
    wb = openpyxl.load_workbook(path)
    if "Referrals" not in wb.sheetnames:
        return 0
    ws = wb["Referrals"]
    col_indices = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}
    company_col = col_indices.get('CompanyName')
    status_col = col_indices.get('Referral_Status')
    source_col = col_indices.get('Referral_Source')
    if not company_col or not status_col or not source_col:
        return 0
        
    count = 0
    normalized_company = str(company_name or '').strip().lower()
    for row in range(2, ws.max_row + 1):
        row_company = str(ws.cell(row=row, column=company_col).value or '').strip().lower()
        if row_company == normalized_company:
            row_source = str(ws.cell(row=row, column=source_col).value or '').strip().lower()
            # Include recruiter source only
            if 'recruiter' not in row_source:
                continue
            row_status = str(ws.cell(row=row, column=status_col).value or '').strip().lower()
            if row_status in ('pending', 'sent', 'replied', 'referral received'):
                count += 1
    return count



def clean_company_url(url):
    """Normalizes any LinkedIn company URL to standard format."""
    if not url:
        return ""
    url = str(url).strip()
    if "/company/" in url:
        parts = url.split("/company/")
        slug = parts[1].split("/")[0].split("?")[0].rstrip("/")
        return f"https://www.linkedin.com/company/{slug}/"
    return url


def get_completed_referral_count(company_name, job_url=None, job_id=None, path=None):
    """Counts completed referral outreach actions (sent, replied, referral received) for a company and optional JobID/JobURL."""
    if path is None:
        path = get_referrals_file()
    if not os.path.exists(path):
        return 0
    init_referrals_store(path)
    wb = openpyxl.load_workbook(path)
    if "Referrals" not in wb.sheetnames:
        return 0
    ws = wb["Referrals"]
    col_indices = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}
    company_col = col_indices.get('CompanyName')
    job_url_col = col_indices.get('Job_URL')
    status_col = col_indices.get('Referral_Status')
    source_col = col_indices.get('Referral_Source')
    job_id_col = col_indices.get('JobID')
    if not company_col or not status_col or not source_col:
        return 0
        
    count = 0
    normalized_company = str(company_name or '').strip().lower()
    normalized_url = str(job_url).strip().lower() if job_url else ""
    
    for row in range(2, ws.max_row + 1):
        row_company = str(ws.cell(row=row, column=company_col).value or '').strip().lower()
        if row_company == normalized_company:
            if normalized_url and job_url_col:
                row_url = str(ws.cell(row=row, column=job_url_col).value or "").strip().lower()
                if row_url and row_url != normalized_url:
                    continue
            if job_id is not None and job_id_col:
                row_job_id = ws.cell(row=row, column=job_id_col).value
                if row_job_id is not None and str(row_job_id).strip() != str(job_id).strip():
                    continue
            row_source = str(ws.cell(row=row, column=source_col).value or '').strip().lower()
            if row_source in ('existing employee', 'sent employee connection'):
                row_status = str(ws.cell(row=row, column=status_col).value or '').strip().lower()
                if row_status in ('sent', 'replied', 'referral received'):
                    count += 1
    return count


def sync_job_lead_referral_statuses(path=None):
    """Automatically updates job tracker status to 'Referral Outreach Completed' if targets are met."""
    if path is None:
        path = get_job_leads_file()
    if not os.path.exists(path):
        return
        
    try:
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        col_indices = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}
        
        id_col = col_indices.get('JobID')
        company_col = col_indices.get('CompanyName')
        status_col = col_indices.get('Status')
        linkedin_url_col = col_indices.get('LinkedIn_Company_URL')
        
        if not id_col or not company_col or not status_col:
            return
            
        from config.user_profiles import get_selected_user_config
        user_conf = get_selected_user_config()
        connect_conf = user_conf.get("linkedin_connect", {})
        target = int(connect_conf.get("max_connections_per_run") or 5)
        
        referrals = load_all_referrals()
        company_data = {}
        for ref in referrals:
            c_name = str(ref.get("CompanyName") or "").strip().lower()
            c_url = clean_company_url(ref.get("Company_URL") or "")
            job_id_val = str(ref.get("JobID") or "").strip()
            ref_source = str(ref.get("Referral_Source") or "").strip().lower()
            ref_status = str(ref.get("Referral_Status") or "").strip().lower()
            is_employee = ref_source in ("existing employee", "sent employee connection")
            is_valid_status = ref_status in ("sent", "replied", "referral received")
            
            key = (c_name, c_url, job_id_val)
            if key not in company_data:
                company_data[key] = 0
            if is_employee and is_valid_status:
                company_data[key] += 1
                
        updated = False
        for row in range(2, ws.max_row + 1):
            company_name = str(ws.cell(row=row, column=company_col).value or "").strip()
            c_name_lower = company_name.lower()
            row_job_id = str(ws.cell(row=row, column=id_col).value or "").strip()
            
            company_url = ""
            if linkedin_url_col:
                company_url = clean_company_url(ws.cell(row=row, column=linkedin_url_col).value or "")
                
            if not company_url:
                for (cn, cu, jid) in company_data.keys():
                    if cn == c_name_lower and cu:
                        company_url = cu
                        break
            if not company_url:
                slug = company_name.lower().replace(" ", "-").replace(".", "").replace(",", "")
                company_url = f"https://www.linkedin.com/company/{slug}/"
                
            completed_count = 0
            for (cn, cu, jid), count in company_data.items():
                if cn == c_name_lower and jid == row_job_id:
                    if not company_url or not cu or company_url == cu:
                        completed_count += count
                        
            if completed_count >= target:
                current_status = str(ws.cell(row=row, column=status_col).value or "").strip()
                if current_status != 'Referral Outreach Completed' and current_status.lower() != 'done':
                    ws.cell(row=row, column=status_col, value='Referral Outreach Completed')
                    updated = True
                    
            if linkedin_url_col:
                current_linkedin_url = ws.cell(row=row, column=linkedin_url_col).value
                if not current_linkedin_url:
                    ws.cell(row=row, column=linkedin_url_col, value=company_url)
                    updated = True
                    
        if updated:
            wb.save(path)
            _trigger_mac_excel_reload(path)
    except Exception as e:
        logger.error(f"Error syncing job tracker referral statuses: {e}")


def load_job_leads_with_referral_counts():
    """Loads all job leads enriched with dynamic referral progress metrics."""
    from config.user_profiles import get_selected_user_config
    try:
        user_conf = get_selected_user_config()
    except Exception:
        user_conf = {}
    connect_conf = user_conf.get("linkedin_connect", {})
    target = int(connect_conf.get("max_connections_per_run") or 5)

    leads_file = get_job_leads_file()
    init_job_leads_store(leads_file)
    
    # Run sync to ensure latest states are written back
    sync_job_lead_referral_statuses()
    
    wb = openpyxl.load_workbook(leads_file)
    ws = wb.active
    col_indices = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}
    
    leads = []
    for row in range(2, ws.max_row + 1):
        row_dict = {}
        for h, idx in col_indices.items():
            if h:
                row_dict[h] = ws.cell(row=row, column=idx).value
        if any(val is not None for val in row_dict.values()):
            leads.append(row_dict)
            
    referrals = load_all_referrals()
    company_data = {}
    for ref in referrals:
        c_name = str(ref.get("CompanyName") or "").strip().lower()
        c_url = clean_company_url(ref.get("Company_URL") or "")
        job_id_val = str(ref.get("JobID") or "").strip()
        ref_source = str(ref.get("Referral_Source") or "").strip().lower()
        ref_status = str(ref.get("Referral_Status") or "").strip().lower()

        is_employee = ref_source in ("existing employee", "sent employee connection")
        is_valid_status = ref_status in ("sent", "replied", "referral received")

        key = (c_name, c_url, job_id_val)
        if key not in company_data:
            company_data[key] = {
                "completed": 0,
                "urls": set()
            }
        
        if c_url:
            company_data[key]["urls"].add(c_url)
        if is_employee and is_valid_status:
            company_data[key]["completed"] += 1

    for lead in leads:
        company_name = str(lead.get("CompanyName") or "").strip()
        c_name_lower = company_name.lower()
        lead_job_id = str(lead.get("JobID") or "").strip()
        lead_status = str(lead.get("Status") or "").strip().lower()
        
        # If the job is marked as Not Interested, show zero for all referral metrics
        if lead_status == "not interested":
            lead["LinkedIn_Company_URL"] = clean_company_url(lead.get("LinkedIn_Company_URL") or "")
            lead["Referral_Target"] = 0
            lead["Referral_Completed"] = 0
            lead["Referral_Remaining"] = 0
            lead["Referral_Target_Achieved"] = "N/A"
            continue
        
        company_url = clean_company_url(lead.get("LinkedIn_Company_URL") or "")
        if not company_url:
            for (cn, cu, jid) in company_data.keys():
                if cn == c_name_lower and cu:
                    company_url = cu
                    break
        if not company_url:
            slug = company_name.lower().replace(" ", "-").replace(".", "").replace(",", "")
            company_url = f"https://www.linkedin.com/company/{slug}/"

        completed_count = 0
        for (cn, cu, jid), data in company_data.items():
            if cn == c_name_lower and jid == lead_job_id:
                if not company_url or not cu or company_url == cu:
                    completed_count += data["completed"]

        remaining = max(0, target - completed_count)
        achieved = "Yes" if remaining == 0 else "No"
        
        lead["LinkedIn_Company_URL"] = company_url
        lead["Referral_Target"] = target
        lead["Referral_Completed"] = completed_count
        lead["Referral_Remaining"] = remaining
        lead["Referral_Target_Achieved"] = achieved

    return leads



def migrate_company_url_and_tracking_fields():
    """Migrates existing data files to add Company_URL / LinkedIn_Company_URL columns if missing, and backfills values."""
    users_dir = os.path.join(BASE_DIR, "users")
    if not os.path.exists(users_dir):
        return
        
    for user_name in os.listdir(users_dir):
        user_path = os.path.join(users_dir, user_name)
        if not os.path.isdir(user_path) or user_name == "default":
            continue
            
        data_dir = os.path.join(user_path, "data")
        referrals_path = os.path.join(data_dir, "referrals.xlsx")
        tracker_path = os.path.join(data_dir, "LinkedIn_Job_Tracker.xlsx")
        
        if os.path.exists(referrals_path):
            try:
                # 1. Load job leads url map for backfilling
                job_leads_map = {}
                if os.path.exists(tracker_path):
                    try:
                        tracker_wb = openpyxl.load_workbook(tracker_path)
                        tracker_ws = tracker_wb.active
                        tracker_cols = {cell.value: idx for idx, cell in enumerate(tracker_ws[1], start=1)}
                        jid_idx = tracker_cols.get("JobID")
                        surl_idx = tracker_cols.get("ShortenURL")
                        curl_idx = tracker_cols.get("CompanyURL")
                        if jid_idx:
                            for r in range(2, tracker_ws.max_row + 1):
                                jid = tracker_ws.cell(row=r, column=jid_idx).value
                                if jid is not None:
                                    val = ""
                                    if surl_idx:
                                        val = tracker_ws.cell(row=r, column=surl_idx).value or ""
                                    if not val and curl_idx:
                                        val = tracker_ws.cell(row=r, column=curl_idx).value or ""
                                    job_leads_map[str(jid).strip()] = str(val).strip()
                    except Exception as e:
                        logger.warning(f"Could not load tracker for backfilling Job_URL: {e}")

                wb = openpyxl.load_workbook(referrals_path)
                if "Referrals" in wb.sheetnames:
                    ws = wb["Referrals"]
                    headers = [cell.value for cell in ws[1]]
                    
                    changed = False
                    if "Company_URL" in headers and "Job_URL" not in headers:
                        company_url_idx = headers.index("Company_URL") + 1
                        ws.cell(row=1, column=company_url_idx, value="Job_URL")
                        headers = [cell.value for cell in ws[1]]
                        changed = True
                    elif "Job_URL" not in headers:
                        logger.info(f"Adding Job_URL column to referrals.xlsx for user {user_name}")
                        ws.insert_cols(4)
                        ws.cell(row=1, column=4, value="Job_URL")
                        headers = [cell.value for cell in ws[1]]
                        changed = True

                    col_indices = {h: idx for idx, h in enumerate(headers, start=1)}
                    job_id_col = col_indices.get("JobID")
                    url_col = col_indices.get("Job_URL")
                    
                    if url_col:
                        for r in range(2, ws.max_row + 1):
                            current_val = ws.cell(row=r, column=url_col).value
                            is_old_company_url = current_val and str(current_val).strip().startswith("https://www.linkedin.com/company/")
                            if not current_val or str(current_val).strip() == "" or is_old_company_url:
                                jid_val = str(ws.cell(row=r, column=job_id_col).value or "").strip() if job_id_col else ""
                                if jid_val in job_leads_map:
                                    ws.cell(row=r, column=url_col, value=job_leads_map[jid_val])
                                    changed = True
                                    
                    if changed:
                        ws._tables.clear()
                        ref_range = f"A1:{chr(64 + len(REFERRAL_HEADERS))}{ws.max_row}"
                        tab = Table(displayName="ReferralsTable", ref=ref_range)
                        style = TableStyleInfo(
                            name="TableStyleLight9",
                            showFirstColumn=False,
                            showLastColumn=False,
                            showRowStripes=True,
                            showColumnStripes=False,
                        )
                        tab.tableStyleInfo = style
                        ws.add_table(tab)
                        wb.save(referrals_path)
                        logger.info(f"Successfully migrated referrals schema for user {user_name}")
            except Exception as e:
                logger.error(f"Failed to migrate referrals.xlsx schema for user {user_name}: {e}")
                
        if os.path.exists(tracker_path):
            try:
                wb = openpyxl.load_workbook(tracker_path)
                ws = wb.active
                headers = [cell.value for cell in ws[1]]
                
                changed = False
                if "LinkedIn_Company_URL" not in headers:
                    logger.info(f"Adding LinkedIn_Company_URL column to LinkedIn_Job_Tracker.xlsx for user {user_name}")
                    ws.insert_cols(4)
                    ws.cell(row=1, column=4, value="LinkedIn_Company_URL")
                    headers = [cell.value for cell in ws[1]]
                    changed = True

                col_indices = {h: idx for idx, h in enumerate(headers, start=1)}
                company_col = col_indices.get("CompanyName")
                linkedin_url_col = col_indices.get("LinkedIn_Company_URL")
                
                if company_col and linkedin_url_col:
                    for r in range(2, ws.max_row + 1):
                        current_val = ws.cell(row=r, column=linkedin_url_col).value
                        if not current_val or str(current_val).strip() == "":
                            comp_name = str(ws.cell(row=r, column=company_col).value or "").strip()
                            if comp_name:
                                slug = comp_name.lower().replace(" ", "-").replace(".", "").replace(",", "")
                                default_url = f"https://www.linkedin.com/company/{slug}/"
                                ws.cell(row=r, column=linkedin_url_col, value=default_url)
                                changed = True
                                
                if changed:
                    ws._tables.clear()
                    ref_range = f"A1:{chr(64 + len(JOB_LEADS_HEADERS))}{ws.max_row}"
                    tab = Table(displayName="JobTrackerTable", ref=ref_range)
                    style = TableStyleInfo(
                        name="TableStyleLight9",
                        showFirstColumn=False,
                        showLastColumn=False,
                        showRowStripes=True,
                        showColumnStripes=False,
                    )
                    tab.tableStyleInfo = style
                    ws.add_table(tab)
                    wb.save(tracker_path)
                    logger.info(f"Successfully migrated job tracker schema for user {user_name}")
            except Exception as e:
                logger.error(f"Failed to migrate LinkedIn_Job_Tracker.xlsx schema for user {user_name}: {e}")

# Run schema migrations for company URLs and tracking fields
migrate_company_url_and_tracking_fields()
