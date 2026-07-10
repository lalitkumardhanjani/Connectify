import os
import openpyxl
from config.settings import JOB_LEADS_FILE
from core.integrations.selenium_driver import get_driver
from core.storage.database import update_status_by_id, load_saved_jobs
from core.logging.config import logger

def load_job_data_excel(filename):
    """Load job data from the Excel tracker and return rows, workbook, sheet, headers."""
    from core.storage.database import get_sheets_config
    sheets_conf = get_sheets_config()
    if sheets_conf:
        url, creds = sheets_conf
        from core.storage.sheets import read_rows
        try:
            rows = read_rows(url, creds, "Job Leads")
            from config.constants import GOOGLE_SHEET_WORKSHEETS
            headers = GOOGLE_SHEET_WORKSHEETS["jobs"]["headers"]
            return rows, None, None, headers
        except Exception as e:
            logger.error(f"Error loading job leads from Google Sheets: {e}")
            return [], None, None, []

    if not os.path.exists(filename):
        logger.error(f"Excel file '{filename}' not found!")
        return [], None, None, []
    try:
        wb = openpyxl.load_workbook(filename)
        ws = wb.active
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            rows.append(dict(zip(headers, row)))
        return rows, wb, ws, headers
    except Exception as e:
        logger.exception(f"Error loading Excel file: {e}")
        return [], None, None, []

def save_job_data_excel(wb, ws, headers, rows, filename):
    """Write rows back to the Excel worksheet and save the workbook."""
    from core.storage.database import get_sheets_config
    sheets_conf = get_sheets_config()
    if sheets_conf:
        url, creds = sheets_conf
        from core.storage.sheets import write_rows
        try:
            write_rows(url, creds, "Job Leads", rows)
            logger.info("Saved progress to Google Sheets.")
        except Exception as e:
            logger.error(f"Error saving job leads to Google Sheets: {e}")

    if ws is None or wb is None:
        if os.path.exists(filename):
            try:
                import openpyxl
                wb = openpyxl.load_workbook(filename)
                ws = wb.active
            except Exception as e:
                logger.error(f"Error loading local workbook for backup: {e}")
                return

    if ws is not None and wb is not None:
        try:
            ws.delete_rows(2, ws.max_row)
            for row_dict in rows:
                ws.append([row_dict.get(h, "") for h in headers])
            wb.save(filename)
            logger.info(f"Excel file '{filename}' saved.")
        except Exception as e:
            logger.error(f"Error saving local excel copy: {e}")

def run_reviewer():
    """Runs the terminal CLI reviewer interface."""
    rows, wb, ws, headers = load_job_data_excel(JOB_LEADS_FILE)
    if not rows:
        logger.warning("No jobs found in Excel tracker database!")
        return

    def col(name):
        for h in headers:
            if isinstance(h, str) and h.lower() == name.lower():
                return h
        return None

    url_col = col('CompanyURL')
    company_col = col('CompanyName')
    status_col = col('Status')
    
    processed_rows = []
    jobs_to_process = []
    
    for row in rows:
        status = row.get(status_col)
        if status and str(status).strip().upper() == "NEW":
            jobs_to_process.append(row)
        else:
            processed_rows.append(row)
            
    total = len(jobs_to_process)
    if total == 0:
        logger.info("No jobs left with status 'NEW' to review.")
        return

    try:
        driver = get_driver()
    except Exception as e:
        logger.error(f"Error starting Chrome browser: {e}")
        import sys
        sys.exit(1)

    updated_rows = []
    try:
        for index, row in enumerate(jobs_to_process, start=1):
            url = row.get(url_col, "")
            company = row.get(company_col, "Unknown")

            if not url:
                logger.info(f"No URL found for {company}. Skipping...")
                updated_rows.append(row)
                continue

            print("\n" + "=" * 60)
            print(f"Job {index}/{total}")
            print(f"Company : {company}")
            print("=" * 60)
            driver.get(url)

            while True:
                print("\nOptions:")
                print("1 -> Interested (set status)")
                print("2 -> Not Interested (set status)")
                print("q -> Quit")
                choice = input("\nEnter choice: ").strip().lower()
                if choice in ['1', '2', 'q']:
                    break
                print("Invalid choice. Please enter 1, 2, or q.")

            if choice == 'q':
                logger.info("Stopping reviewer script...")
                remaining = jobs_to_process[index-1:]
                updated_rows.extend(remaining)
                all_rows = processed_rows + updated_rows
                save_job_data_excel(wb, ws, headers, all_rows, JOB_LEADS_FILE)
                import sys
                sys.exit(2)
            elif choice == '1':
                logger.info("Interested. Updating status.")
                if status_col:
                    row[status_col] = "Interested"
                updated_rows.append(row)
            elif choice == '2':
                logger.info("Not interested. Updating status.")
                if status_col:
                    row[status_col] = "Not Interested"
                updated_rows.append(row)


            # After handling the choice, save current state
            remaining = jobs_to_process[index:]
            all_rows = processed_rows + updated_rows + remaining
            save_job_data_excel(wb, ws, headers, all_rows, JOB_LEADS_FILE)
            logger.info("Saved progress to Excel.")
    except Exception as e:
        logger.exception(f"\nFatal error during review session: {e}")
        import sys
        sys.exit(1)
    finally:
        logger.info("\nClosing browser...")
        try:
            driver.quit()
        except NameError:
            pass
        logger.info("Reviewer script finished.")

if __name__ == "__main__":
    run_reviewer()
