import os
import sys
import openpyxl
from datetime import datetime

# Add current directory to path
sys.path.append(os.getcwd())

from core.storage.engine import read_database_rows, write_database_rows, get_active_username
from pipelines.linkedin_outreach.services.job_details_scraper import scrape_job_details
from config.settings import get_job_leads_file

def run_extraction():
    # Set the active user profile
    username = get_active_username()
    print(f"Starting Job Details Extraction for user profile: {username}")
    
    # 1. Load the active Excel job tracker file path
    file_path = get_job_leads_file()
    if not os.path.exists(file_path):
        print(f"Error: Excel job leads file not found at {file_path}")
        return

    # 2. Open workbook using openpyxl to ensure columns exist and get sheet reference
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    
    # Get current headers
    headers = [cell.value for cell in ws[1]]
    print(f"Current headers: {headers}")
    
    # Ensure Experience and Location columns are added to headers if they are missing
    modified_headers = False
    if "Experience" not in headers:
        headers.append("Experience")
        ws.cell(row=1, column=len(headers), value="Experience")
        modified_headers = True
    if "Location" not in headers:
        headers.append("Location")
        ws.cell(row=1, column=len(headers), value="Location")
        modified_headers = True
        
    if modified_headers:
        wb.save(file_path)
        print(f"Added missing headers 'Experience' and 'Location' to local Excel sheet.")

    # 3. Read database rows via the storage engine (to ensure we work with the clean dictionary data)
    rows = read_database_rows("jobs", username=username, bypass_cache=True)
    total_jobs = len(rows)
    print(f"Found {total_jobs} total job records in the database.")

    # Cache for already scraped domains/URLs to speed up processing
    scraped_cache = {}
    updated_count = 0

    for idx, row in enumerate(rows, start=1):
        url = row.get("CompanyURL") or row.get("LinkedIn_Company_URL")
        company = row.get("CompanyName") or "Unknown"
        
        # Check if the details are already populated (so we don't scrape unnecessarily)
        current_exp = str(row.get("Experience") or "").strip()
        current_loc = str(row.get("Location") or "").strip()
        current_id = str(row.get("JobID") or "").strip()
        
        # If both Location and Experience are already set and not default placeholders, skip
        if current_exp and current_loc and current_exp != "Not Specified" and current_loc != "Remote / India":
            print(f"[{idx}/{total_jobs}] Skipping {company} (already has Location: {current_loc} | Experience: {current_exp})")
            continue

        if not url or "forms/d" in url or "hsforms.com" in url:
            print(f"[{idx}/{total_jobs}] Skipping {company} (invalid/form URL: {url})")
            # Set default placeholders for forms/missing URLs
            row["Experience"] = "Not Specified"
            row["Location"] = "Remote / India"
            continue

        print(f"\n[{idx}/{total_jobs}] Scraping details for {company}...")
        print(f"  URL: {url}")
        
        # Check cache first
        if url in scraped_cache:
            details = scraped_cache[url]
            print(f"  (Cache Hit) Reusing details for {url}")
        else:
            details = scrape_job_details(url)
            scraped_cache[url] = details
            
        print(f"  Extracted -> JobID: {details['job_id']} | Location: {details['location']} | Experience: {details['experience']}")
        
        # Update row values
        # If scraper extracted a specific Job ID and the existing Job ID is empty or numeric/LinkedIn ID, enrich it
        if details["job_id"] != "N/A":
            row["JobID"] = details["job_id"]
            
        row["Location"] = details["location"]
        row["Experience"] = details["experience"]
        updated_count += 1

        # Periodic save every 5 jobs so progress is not lost if the script is stopped
        if updated_count % 5 == 0:
            write_database_rows("jobs", rows, username=username)
            print(f"--- Saved intermediate progress of {updated_count} updates to database. ---")

    # 4. Final write to save all updates and trigger sheets backup sync
    if updated_count > 0:
        write_database_rows("jobs", rows, username=username)
        print(f"\nExtraction complete! Updated details for {updated_count} jobs.")
    else:
        print("\nAll jobs already had location/experience details populated. No updates needed.")

if __name__ == "__main__":
    run_extraction()
