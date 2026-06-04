from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException, StaleElementReferenceException
from selenium.webdriver.chrome.options import Options
import os
import time
import random
import re
import csv
from urllib.parse import quote

from config import CHROME_PROFILE_PATH, LINKEDIN_EMAIL, LINKEDIN_PASSWORD, DEFAULT_SEARCH_KEYWORDS, REVIEW_MODE
from email_extractor import extract_emails
from email_sender_gmail_web import send_email_via_gmail
from data_store import init_store, append_email, update_status
from logger import logger
import openpyxl

class LinkedInScraper:
    def __init__(self):
        self.driver = self._setup_driver()

    def _setup_driver(self):
        options = Options()
        options.add_argument("--remote-debugging-port=9222")
        options.add_argument(f"--user-data-dir={CHROME_PROFILE_PATH}")
        options.add_argument("--disable-blink-features=AutomationControlled")
        options.add_experimental_option("excludeSwitches", ["enable-automation"])
        options.add_experimental_option('useAutomationExtension', False)

        driver = webdriver.Chrome(options=options)
        driver.maximize_window()
        return driver

    # ---------------------------------------------------------------------
    # Post discovery helpers
    # ---------------------------------------------------------------------
    def _find_post_containers(self):
        selectors = [
            ("css", ".feed-shared-update-v2"),
            ("xpath", "//div[@role='listitem' and .//button[@data-testid='expandable-text-button'] ]"),
            ("xpath", "//div[@role='listitem']"),
            ("xpath", "//article")
        ]
        for strategy, selector in selectors:
            try:
                if strategy == "css":
                    elements = self.driver.find_elements(By.CSS_SELECTOR, selector)
                else:
                    elements = self.driver.find_elements(By.XPATH, selector)
                if elements:
                    logger.info(f"Found {len(elements)} potential posts using selector: {selector}")
                    return elements
            except Exception as e:
                logger.warning(f"Post container selector failed: {selector} -> {e}")
        return []

    def _expand_post(self, post_element):
        expand_selectors = [
            ".//button[@data-testid='expandable-text-button']",
            ".//button[contains(normalize-space(.), 'more')]",
            ".//button[contains(translate(., 'MORE', 'more'), 'more')]"
        ]
        for selector in expand_selectors:
            try:
                button = post_element.find_element(By.XPATH, selector)
                self.driver.execute_script("arguments[0].scrollIntoView({block:'center'});", button)
                time.sleep(0.5)
                self.driver.execute_script("arguments[0].click();", button)
                time.sleep(0.7)
                return True
            except (NoSuchElementException, StaleElementReferenceException):
                continue
            except Exception as e:
                logger.warning(f"Unable to expand post content: {e}")
                continue
        return False

    # ---------------------------------------------------------------------
    # Authentication
    # ---------------------------------------------------------------------
    def login(self):
        """Handle LinkedIn login"""
        self.driver.get("https://www.linkedin.com/feed/")
        logger.info("Checking if already logged in...")
        try:
            search_bar_selectors = [
                "input[placeholder*='Search']",
                ".search-global-typeahead__input",
                "input[role='combobox']"
            ]
            for selector in search_bar_selectors:
                try:
                    search_bar = WebDriverWait(self.driver, 2).until(
                        EC.presence_of_element_located((By.CSS_SELECTOR, selector))
                    )
                    if search_bar:
                        logger.info("Already logged in! (Search bar found)")
                        return True
                except TimeoutException:
                    continue
            logger.info("Not logged in. Attempting login...")
            if not LINKEDIN_EMAIL or not LINKEDIN_PASSWORD:
                logger.error("LinkedIn credentials missing in config.")
                return False
            email_selectors = [
                "input[id*='username']",
                "input[name='session_key']",
                "input[type='email']",
                "#username"
            ]
            email_input = None
            for selector in email_selectors:
                try:
                    email_input = WebDriverWait(self.driver, 3).until(
                        EC.visibility_of_element_located((By.CSS_SELECTOR, selector))
                    )
                    break
                except TimeoutException:
                    continue
            if email_input:
                email_input.send_keys(LINKEDIN_EMAIL)
                password_input = self.driver.find_element(By.CSS_SELECTOR, "input[type='password']")
                password_input.send_keys(LINKEDIN_PASSWORD)
                button_selectors = [
                    "button[type='submit']",
                    "button[data-litms-control-urn*='sign-in']",
                    ".btn__primary--large"
                ]
                for selector in button_selectors:
                    try:
                        self.driver.find_element(By.CSS_SELECTOR, selector).click()
                        break
                    except NoSuchElementException:
                        continue
                time.sleep(5)
                logger.info("Login successful!")
                return True
        except Exception as e:
            logger.error(f"Login error: {str(e)}")
            return False

    # ---------------------------------------------------------------------
    # Search handling per keyword
    # ---------------------------------------------------------------------
    def search_for_keyword(self, keyword):
        """Navigate to the LinkedIn search page for a given keyword and ensure the Posts/LATEST view is active."""
        search_query = quote(keyword)
        url = f"https://www.linkedin.com/search/results/content/?keywords={search_query}&origin=GLOBAL_SEARCH_HEADER&sortBy=%22date_posted%22"
        logger.info(f"Navigating to: {url}")
        self.driver.get(url)
        time.sleep(5)
        if "search/results/content" not in self.driver.current_url:
            logger.warning("Direct navigation failed – falling back to manual UI steps.")
            self.driver.get("https://www.linkedin.com/feed/")
            time.sleep(3)
            try:
                search_input = self.driver.find_element(By.CLASS_NAME, "search-global-typeahead__input")
                search_input.send_keys(keyword)
                search_input.send_keys(u'\ue007')  # Enter
                time.sleep(5)
                posts_btn = self.driver.find_element(By.XPATH, "//button[text()='Posts']")
                posts_btn.click()
                time.sleep(3)
                sort_btn = self.driver.find_element(By.XPATH, "//button[contains(@aria-label, 'Sort by')]")
                sort_btn.click()
                time.sleep(1)
                latest_option = self.driver.find_element(By.XPATH, "//span[text()='Latest']")
                latest_option.click()
                time.sleep(3)
            except Exception as e:
                logger.error(f"Failed to navigate to Posts view: {e}")
                return False
        return True

    # ---------------------------------------------------------------------
    # Post data extraction (content only – we ignore name/date/url)
    # ---------------------------------------------------------------------
    def extract_post_data(self, post_element):
        """Extract raw text content from a post element."""
        try:
            self._expand_post(post_element)
            content_selectors = [
                ".//*[contains(@class, 'feed-shared-text')]",
                ".//*[contains(@class, 'update-components-text')]",
                ".//div[contains(@data-display-contents, 'true')]",
                ".//div[contains(@class, 'show-more-less-html')]",
                ".//div[@data-test-id='feed-item-text']",
                ".//p",
                ".//span",
                "."  # fallback to the whole element text
            ]
            post_content = ""
            for selector in content_selectors:
                try:
                    elements = post_element.find_elements(By.XPATH, selector)
                    if not elements:
                        continue
                    texts = []
                    for el in elements:
                        try:
                            txt = el.text.strip()
                            if txt and len(txt) > 5:
                                texts.append(txt)
                        except Exception:
                            continue
                    post_content = " ".join(texts).strip()
                    if len(post_content) > 50:
                        break
                except Exception:
                    continue
            post_content = re.sub(r"\s+", " ", post_content).replace("… more", "").strip()
            return {"content": post_content}
        except StaleElementReferenceException:
            logger.warning("Post element went stale while processing.")
            return None
        except Exception as e:
            logger.warning(f"Error extracting post data: {e}")
            return None

    # ---------------------------------------------------------------------
    # Scrolling loop with per‑keyword timeout
    # ---------------------------------------------------------------------
    def process_keyword(self, keyword, timeout_seconds=60):
        start_time = time.time()
        processed_ids = set()
        last_height = self.driver.execute_script("return document.body.scrollHeight")
        no_new_posts = 0
        while True:
            # Check timeout first
            if time.time() - start_time > timeout_seconds:
                logger.info(f"Keyword '{keyword}' timed out after {timeout_seconds}s.")
                break
            try:
                posts = self._find_post_containers()
                if not posts:
                    logger.warning("No posts found on the current page.")
                    no_new_posts += 1
                else:
                    no_new_posts = 0
                for post in posts:
                    post_id = post.get_attribute("data-urn") or str(hash(post.text))
                    if post_id in processed_ids:
                        continue
                    data = self.extract_post_data(post)
                    if data and data.get('content'):
                        content = data['content']
                        # Verify at least one keyword appears
                        if any(kw.lower() in content.lower() for kw in DEFAULT_SEARCH_KEYWORDS):
                            emails = extract_emails(content)
                            for email in emails:
                                # append_email handles duplicate checking and immediate save
                                appended = append_email(email, keyword)
                                if appended:
                                    logger.info(f"Collected new email: {email}")
                        else:
                            logger.debug("Post did not contain any target keyword – skipped.")
                    processed_ids.add(post_id)
            except Exception as e:
                logger.error(f"Error during post processing: {e}")

            # Scroll down
            self.driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            time.sleep(random.uniform(2, 3))
            new_height = self.driver.execute_script("return document.body.scrollHeight")
            if new_height == last_height:
                no_new_posts += 1
                logger.info(f"No new posts loaded (attempt {no_new_posts}/3)")
                if no_new_posts >= 3:
                    logger.info("Reached bottom of page – exiting scroll loop.")
                    break
            else:
                no_new_posts = 0
            last_height = new_height

    def close(self):
        self.driver.quit()

# -------------------------------------------------------------------------
# Phase execution helpers
# -------------------------------------------------------------------------
def run_phase_one(scraper):
    init_store()
    for kw in DEFAULT_SEARCH_KEYWORDS:
        logger.info(f"=== Processing keyword: '{kw}' ===")
        if scraper.search_for_keyword(kw):
            scraper.process_keyword(kw, timeout_seconds=60)
        else:
            logger.warning(f"Skipping keyword '{kw}' due to navigation failure.")

def run_phase_two(scraper, review_mode=REVIEW_MODE):
    excel_path = 'job_tracker.xlsx'
    if not os.path.exists(excel_path):
        logger.warning("Excel file not found – nothing to send.")
        return
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    col_map = {cell.value: idx+1 for idx, cell in enumerate(ws[1])}
    email_col = col_map.get('Email')
    status_col = col_map.get('Status')
    if not email_col or not status_col:
        logger.error("Excel schema missing required columns.")
        return
    for row in range(2, ws.max_row + 1):
        status = (ws.cell(row=row, column=status_col).value or '').strip().lower()
        if status == 'sent':
            continue
        email = ws.cell(row=row, column=email_col).value
        if not email:
            continue
        logger.info(f"Sending email to {email} (row {row})")
        sent = send_email_via_gmail(scraper.driver, email, review_mode=review_mode)
        if sent:
            update_status(email, 'sent')
        else:
            logger.info(f"Email to {email} not sent – leaving status unchanged.")

if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description="LinkedIn email scraper with optional phases.")
    parser.add_argument(
        "--phase",
        choices=["full", "phase1", "phase2"],
        default="full",
        help="Which part of the workflow to run: full (both phases), phase1 only, or phase2 only.",
    )
    parser.add_argument(
        "--review",
        action="store_true",
        help="Ask for confirmation before sending each email (default false).",
    )
    args = parser.parse_args()

    scraper = LinkedInScraper()
    try:
        if not scraper.login():
            logger.error("Login failed – aborting.")
        else:
            if args.phase in ("full", "phase1"):
                run_phase_one(scraper)
            if args.phase in ("full", "phase2"):
                run_phase_two(scraper, review_mode=args.review)
    finally:
        scraper.close()
