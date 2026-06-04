import os
from datetime import datetime
import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo

EXCEL_FILE = 'job_tracker.xlsx'
HEADERS = ['ID', 'Email', 'Status', 'Timestamp', 'Keyword']

def init_store(path=EXCEL_FILE):
    if not os.path.exists(path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Jobs"
        ws.append(HEADERS)
        
        # Format as Table with all columns including Referral Person
        col_count = len(HEADERS)
        tab = Table(displayName="JobTrackerTable", ref=f"A1:{chr(64+col_count)}1")
        style = TableStyleInfo(
            name="TableStyleLight9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        tab.tableStyleInfo = style
        ws.add_table(tab)
        wb.save(path)
        _trigger_mac_excel_reload(path)
    # Ensure any existing file conforms to the minimal schema
    trim_excel_to_schema(path)
    migrate_pending_to_new(path)

def _trigger_mac_excel_reload(path):
    # On macOS, if the user has Excel/Numbers open, reload the file so they see updates instantly
    try:
        import subprocess
        abs_path = os.path.abspath(path)
        filename = os.path.basename(path)
        
        reload_script = f'''\
        tell application "System Events"
            set excelRunning to (name of processes) contains "Microsoft Excel"
            set numbersRunning to (name of processes) contains "Numbers"
        end tell
        
        if excelRunning then
            tell application "Microsoft Excel"
                try
                    if exists workbook "{filename}" then
                        close workbook "{filename}" saving no
                        open POSIX file "{abs_path}"
                    end if
                end try
            end tell
        end if
        
        if numbersRunning then
            tell application "Numbers"
                try
                    set docName to "{filename}"
                    set docExists to false
                    repeat with doc in documents
                        if name of doc is docName then
                            set docExists to true
                            exit repeat
                        end if
                    end repeat
                    if docExists then
                        close document docName saving no
                        open POSIX file "{abs_path}"
                    end if
                end try
            end tell
        end if
        '''
        subprocess.run(["osascript", "-e", reload_script], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
    except Exception:
        pass

def trim_excel_to_schema(path=EXCEL_FILE):
    """Ensure the Excel file contains at least the required columns and the optional 'Referral Person' column.
    Missing columns are added; extra columns beyond the defined headers are removed.
    """
    if not os.path.exists(path):
        return
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    # Current headers
    current_headers = [cell.value for cell in ws[1]]
    needed = HEADERS  # includes Referral Person
    header_map = {h: i+1 for i, h in enumerate(current_headers) if h in needed}
    # Add missing headers at the end
    for h in needed:
        if h not in header_map:
            ws.cell(row=1, column=len(current_headers)+1, value=h)
            header_map[h] = len(current_headers)+1
            current_headers.append(h)
    # Rearrange columns in place to match needed order
    for target_idx, h in enumerate(needed, start=1):
        src_idx = header_map[h]
        if src_idx != target_idx:
            for row in range(1, ws.max_row + 1):
                ws.cell(row=row, column=target_idx, value=ws.cell(row=row, column=src_idx).value)
    # Remove any extra columns beyond needed
    max_col = ws.max_column
    if max_col > len(needed):
        ws.delete_cols(len(needed)+1, max_col - len(needed))
    # Refresh table reference
    ws._tables.clear()
    ref = f"A1:{chr(64+len(needed))}{ws.max_row}"
    tab = Table(displayName="JobTrackerTable", ref=ref)
    style = TableStyleInfo(
        name="TableStyleLight9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    tab.tableStyleInfo = style
    ws.add_table(tab)
    wb.save(path)
    _trigger_mac_excel_reload(path)

def append_email(email, keyword='', path=EXCEL_FILE):
    init_store(path)
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    col_indices = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}
    id_col = col_indices.get('ID', 1)
    email_col = col_indices.get('Email', 2)
    # Check duplicate Email
    normalized_email = str(email or '').strip().lower()
    for row in range(2, ws.max_row + 1):
        if str(ws.cell(row=row, column=email_col).value or '').strip().lower() == normalized_email:
            return False
    # Auto‑increment ID
    max_id = max([ws.cell(row=row, column=id_col).value or 0 for row in range(2, ws.max_row + 1)], default=0)
    new_id = int(max_id) + 1
    timestamp = datetime.utcnow().isoformat()
    new_row = [new_id, email, 'New', timestamp, keyword]
    ws.append(new_row)
    # Refresh table reference
    ws._tables.clear()
    col_count = len(HEADERS)
    ref = f"A1:{chr(64+col_count)}{ws.max_row}"
    tab = Table(displayName="JobTrackerTable", ref=ref)
    style = TableStyleInfo(
        name="TableStyleLight9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    tab.tableStyleInfo = style
    ws.add_table(tab)
    wb.save(path)
    _trigger_mac_excel_reload(path)
    return True

def update_status(email, status, path=EXCEL_FILE):
    """Update status and timestamp for a row identified by email."""
    init_store(path)
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    col_indices = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}
    email_col = col_indices.get('Email', 2)
    status_col = col_indices.get('Status', 3)
    timestamp_col = col_indices.get('Timestamp', 4)
    normalized_email = str(email or '').strip().lower()
    updated = False
    for row in range(2, ws.max_row + 1):
        if str(ws.cell(row=row, column=email_col).value or '').strip().lower() == normalized_email:
            ws.cell(row=row, column=status_col, value=status)
            ws.cell(row=row, column=timestamp_col, value=datetime.utcnow().isoformat())
            updated = True
            break
    if updated:
        wb.save(path)
        _trigger_mac_excel_reload(path)
    return updated

def count_unique_emails(path=EXCEL_FILE):
    if not os.path.exists(path):
        return 0
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    col_indices = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}
    email_col = col_indices.get('Email', 2)
    unique_emails = set()
    for row in range(2, ws.max_row + 1):
        val = ws.cell(row=row, column=email_col).value
        if val:
            unique_emails.add(str(val).strip().lower())
    return len(unique_emails)


def load_jobs_for_referral(path='LinkedIn_Job_Tracker.xlsx', status_filter='Ask for referral'):
    """Return a list of dict rows where Status matches the filter, including all columns."""
    if not os.path.exists(path):
        return []
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    # Build a mapping from header names to column indices
    col_indices = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}
    rows = []
    for row in range(2, ws.max_row + 1):
        raw_status = ws.cell(row=row, column=col_indices.get('Status')).value
        status = str(raw_status).strip().lower() if raw_status is not None else ''
        if status == status_filter.strip().lower():
            # Include all columns in the dict
            row_dict = {header: ws.cell(row=row, column=col_idx).value for header, col_idx in col_indices.items()}
            rows.append(row_dict)
    return rows

def append_referral_person(job_id, person_name, path='LinkedIn_Job_Tracker.xlsx'):
    """Append a person's name to the ReferralPerson column for the given JobID."""
    if not os.path.exists(path):
        return False
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    col_indices = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}
    id_col = col_indices.get('JobID')
    referral_col = col_indices.get('ReferralPerson')
    if not id_col or not referral_col:
        return False
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=id_col).value == job_id:
            existing = ws.cell(row=row, column=referral_col).value
            if existing:
                names = [n.strip() for n in existing.split(',')]
                if person_name not in names:
                    new_val = f"{existing}, {person_name}"
                else:
                    new_val = existing
            else:
                new_val = person_name
            ws.cell(row=row, column=referral_col, value=new_val)
            wb.save(path)
            _trigger_mac_excel_reload(path)
            return True
    return False

def update_status_by_id(job_id, status, path='LinkedIn_Job_Tracker.xlsx'):
    """Update the Status column for a row identified by JobID."""
    if not os.path.exists(path):
        return False
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    col_indices = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}
    id_col = col_indices.get('JobID')
    status_col = col_indices.get('Status')
    timestamp_col = col_indices.get('CreatedDateTime')
    if not id_col or not status_col:
        return False
    updated = False
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=id_col).value == job_id:
            ws.cell(row=row, column=status_col, value=status)
            if timestamp_col:
                ws.cell(row=row, column=timestamp_col, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
            updated = True
            break
    if updated:
        wb.save(path)
        _trigger_mac_excel_reload(path)
    return updated

def edit_row(row_id, email, status, keyword, path=EXCEL_FILE):
    """Modify details of an existing record in the Scraper Database by ID."""
    init_store(path)
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    col_indices = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}
    id_col = col_indices.get('ID', 1)
    email_col = col_indices.get('Email', 2)
    status_col = col_indices.get('Status', 3)
    keyword_col = col_indices.get('Keyword', 5)
    
    updated = False
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=id_col).value == row_id:
            if email_col:
                ws.cell(row=row, column=email_col, value=email)
            if status_col:
                ws.cell(row=row, column=status_col, value=status)
            if keyword_col:
                ws.cell(row=row, column=keyword_col, value=keyword)
            updated = True
            break
            
    if updated:
        wb.save(path)
        _trigger_mac_excel_reload(path)
    return updated

def migrate_pending_to_new(path=EXCEL_FILE):
    """Convert any existing records with status 'Pending' to 'New'."""
    if not os.path.exists(path):
        return
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    col_indices = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}
    status_col = col_indices.get('Status')
    if not status_col:
        return
    updated = False
    for row in range(2, ws.max_row + 1):
        val = ws.cell(row=row, column=status_col).value
        if val and str(val).strip().lower() == 'pending':
            ws.cell(row=row, column=status_col, value='New')
            updated = True
    if updated:
        wb.save(path)
        _trigger_mac_excel_reload(path)

