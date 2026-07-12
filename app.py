import warnings
warnings.filterwarnings('ignore', message='.*urllib3 v2 only supports OpenSSL.*')
import logging

class NoDevServerWarningFilter(logging.Filter):
    def filter(self, record):
        msg = record.getMessage()
        return "development server" not in msg and "production deployment" not in msg

logging.getLogger("werkzeug").addFilter(NoDevServerWarningFilter())
import os
import sys
import json
import time
import subprocess
import threading
import signal
from datetime import datetime
from flask import Flask, jsonify, request, render_template, send_from_directory, send_file, Response
from werkzeug.utils import secure_filename
import pandas as pd
import openpyxl

# Connectify Consolidated Configurations and Core Imports
from config.settings import (
    BASE_DIR, get_job_tracker_file, get_job_leads_file, get_referrals_file, get_resumes_dir, get_active_user, get_user_dir
)
from config.user_profiles import (
    load_all_configs, save_all_configs, get_selected_user_name,
    get_selected_user_config, get_global_settings, get_resume_file_path
)
from config.email_templates import DEFAULT_EMAIL_TEMPLATE, DEFAULT_CONNECTION_TEMPLATE
from core.analytics.metrics import get_email_metrics, get_company_metrics, get_outreach_metrics
from core.storage.database import update_status_by_id, edit_row, edit_lead_row

app = Flask(__name__, template_folder='templates', static_folder='static')


if sys.platform == 'win32':
    PYTHON_BIN = os.path.join(os.getcwd(), ".venv", "Scripts", "python.exe")
    if not os.path.exists(PYTHON_BIN):
        PYTHON_BIN = "python"
else:
    PYTHON_BIN = os.path.join(os.getcwd(), ".venv", "bin", "python")
    if not os.path.exists(PYTHON_BIN):
        PYTHON_BIN = "python3"

# Active task tracking
active_tasks = {}
task_lock = threading.Lock()

class SubprocessRunner:
    def __init__(self, task_id, commands, username):
        self.task_id = task_id
        self.username = username          # The user profile this pipeline is pinned to
        self.commands = commands # List of (script_name, list_of_args)
        self.logs = []
        self.status = "queued"
        self.waiting_for_input = False
        self.process = None
        self.current_step = 0
        self.thread = None
        self.start_time = datetime.now().timestamp()
        
        # Determine unique, timestamped filename based on pipeline task ID
        parts = task_id.split("::")
        pipeline_name = "_".join(parts[1:]) if len(parts) > 1 else "pipeline"
        pipeline_name = pipeline_name.replace("::", "_").replace(" ", "_")
        timestamp = datetime.fromtimestamp(self.start_time).strftime("%Y%m%d_%H%M%S")
        self.log_filename = f"{pipeline_name}_{timestamp}.log"

    def start(self):
        self.status = "running"
        self.thread = threading.Thread(target=self._run_loop)
        self.thread.daemon = True
        self.thread.start()

    def log(self, text):
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        log_line = f"[{timestamp}] {text}"
        self.logs.append(log_line)
        self.write_to_log_file(log_line)
        
        # Detailed user-friendly console logging
        parts = self.task_id.split("::")
        pipeline_name = parts[1] if len(parts) > 1 else "pipeline"
        print(f"[Connectify] [USER: {self.username}] [PIPELINE: {pipeline_name}] {text}", flush=True)

    def write_to_log_file(self, line):
        try:
            # Construct active user logs directory path
            user_logs_dir = os.path.join(BASE_DIR, "users", self.username, "logs")
            os.makedirs(user_logs_dir, exist_ok=True)
            log_file_path = os.path.join(user_logs_dir, self.log_filename)
            with open(log_file_path, "a", encoding="utf-8") as f:
                f.write(line + "\n")
        except Exception as e:
            # Simply print write failure to system standard error
            sys.stderr.write(f"Error writing to pipeline log file: {e}\n")

    def _run_loop(self):
        # Automatically clean up logs older than 30 days for this user
        try:
            from core.logging.config import cleanup_old_logs
            user_logs_dir = os.path.join(BASE_DIR, "users", self.username, "logs")
            cleanup_old_logs(user_logs_dir)
        except Exception as e:
            self.log(f"Warning: failed to clean up old logs: {e}")

        for idx, (script, args) in enumerate(self.commands):
            self.current_step = idx + 1
            if script == "run_linkedin_connect.py":
                pass
            elif script == "run_recruiter_outreach.py":
                try:
                    from config.user_profiles import load_all_configs
                    from core.storage.database import load_all_referrals
                    # Read config for the pinned user (not the currently selected UI user)
                    all_cfg = load_all_configs()
                    user_conf = all_cfg.get("users", {}).get(self.username, {})
                    recruiter_conf = user_conf.get("recruiter_outreach", {})
                    target_count = int(recruiter_conf.get("target_count") or 2)
                    
                    referrals = load_all_referrals()
                    today_str = datetime.now().strftime("%Y-%m-%d")
                    
                    # Count recruiter connection requests and recruiter messages sent today
                    rec_msg_sent = sum(
                        1 for r in referrals
                        if str(r.get('Referral_Status') or '').strip().lower() == 'sent'
                        and str(r.get('Sent_Time') or '').strip().startswith(today_str)
                        and str(r.get('Referral_Source') or '').strip() == 'Existing Recruiter'
                    )
                    rec_conn_sent = sum(
                        1 for r in referrals
                        if str(r.get('Referral_Status') or '').strip().lower() == 'sent'
                        and str(r.get('Sent_Time') or '').strip().startswith(today_str)
                        and str(r.get('Referral_Source') or '').strip() == 'Sent Recruiter Connection'
                    )
                    total_rec_sent_today = rec_msg_sent + rec_conn_sent
                    
                    if total_rec_sent_today >= target_count:
                        self.log(f"Target count of {target_count} reached via recruiter outreach today ({total_rec_sent_today} sent). Skipping recruiter connection requests pipeline.")
                        break
                except Exception as e:
                    self.log(f"Warning: error checking recruiter target limits in runner: {e}")
            script_path = os.path.join(os.getcwd(), script)
            cmd = [PYTHON_BIN, "-u", script_path] + args
            
            # Read fresh env variables from .env to forward to subprocess
            env_copy = os.environ.copy()
            if os.path.exists(".env"):
                with open(".env", "r") as f:
                    for line in f:
                        line = line.strip()
                        if line and not line.startswith("#") and "=" in line:
                            k, v = line.split("=", 1)
                            env_copy[k.strip()] = v.strip()

            # Enable isolated and parallel execution of Chrome instances.
            # CONNECTIFY_USER pins this subprocess to the exact user profile — all calls to
            # get_active_user(), get_selected_user_config(), get_data_dir() etc. inside the
            # subprocess will use this user's isolated directories and config, regardless of
            # which user is currently selected in the UI.
            env_copy["CONNECTIFY_USER"] = self.username
            env_copy["CONNECTIFY_PARALLEL"] = "true"
            env_copy["CONNECTIFY_SUBPROCESS_RUNNER"] = "true"
            
            # Derive per-user, per-pipeline Chrome profile directories
            user_dir_for_runner = os.path.join(BASE_DIR, "users", self.username)
            parts = self.task_id.split("::")
            pipeline_type = parts[1] if len(parts) > 1 else ""
            if pipeline_type == "scraper_pipeline":
                if script == "run_email_scraper.py":
                    env_copy["CHROME_PROFILE_DIR"] = os.path.join(user_dir_for_runner, "chrome-profile-email-scraper")
                elif script == "run_email_sender.py":
                    env_copy["CHROME_PROFILE_DIR"] = os.path.join(user_dir_for_runner, "chrome-profile-email-sender")
                else:
                    env_copy["CHROME_PROFILE_DIR"] = os.path.join(user_dir_for_runner, "chrome-profile-scraper")
            elif pipeline_type == "referral_pipeline":
                env_copy["CHROME_PROFILE_DIR"] = os.path.join(user_dir_for_runner, "chrome-profile-referral")
            elif pipeline_type == "recruiter_pipeline":
                env_copy["CHROME_PROFILE_DIR"] = os.path.join(user_dir_for_runner, "chrome-profile-recruiter")

            # Determine active storage type
            storage_type = "Local (Excel files)"
            sheets_url = None
            try:
                from core.storage.engine import GoogleSheetsStorageProvider
                provider = GoogleSheetsStorageProvider()
                sheets_conf = provider.get_sheets_config(self.username)
                if sheets_conf:
                    storage_type = "Google Sheets"
                    sheets_url = sheets_conf[0]
            except Exception:
                pass

            self.log(f"--- Launching Step {self.current_step}/{len(self.commands)}: {script} ---")
            self.log(f"    - Active Profile   : {self.username}")
            self.log(f"    - Database Storage : {storage_type}" + (f" (URL: {sheets_url})" if sheets_url else ""))
            self.log(f"    - Chrome Profile   : {env_copy.get('CHROME_PROFILE_DIR', 'Default')}")
            self.log(f"    - Script Command   : {script} {' '.join(args) if args else ''}")
            
            try:
                self.process = subprocess.Popen(
                    cmd,
                    stdin=subprocess.PIPE,
                    stdout=subprocess.PIPE,
                    stderr=subprocess.STDOUT,
                    text=True,
                    bufsize=1,
                    preexec_fn=None if sys.platform == 'win32' else os.setsid,
                    env=env_copy
                )
            except Exception as e:
                self.log(f"Failed to start script {script}: {e}")
                self.status = "failed"
                return

            # Read stdout line by line
            while True:
                line = self.process.stdout.readline()
                if not line:
                    break
                
                clean_line = line.rstrip('\r\n')
                self.logs.append(clean_line)
                self.write_to_log_file(clean_line)
                
                # Print to terminal console
                parts = self.task_id.split("::")
                pipeline_name = parts[1] if len(parts) > 1 else "pipeline"
                print(f"[Connectify] [USER: {self.username}] [PIPELINE: {pipeline_name}] {clean_line}", flush=True)
                
                # Dynamic step tracking based on stdout lines for combined outreach script
                compare_line = clean_line.strip()
                if "Executing Phase 1: Post email scraping" in compare_line:
                    self.current_step = 1
                elif "Executing Phase 2: Email sending" in compare_line:
                    self.current_step = 2
                
                # Check if review_for_referral or job search is waiting for option/confirmation selection
                if "Enter choice:" in compare_line or "Options:" in compare_line or "press ENTER to continue" in compare_line or "Send [S] / Skip [K] / Quit [Q]" in compare_line:
                    time.sleep(0.2)
                    self.waiting_for_input = True

            self.process.stdout.close()
            return_code = self.process.wait()
            self.waiting_for_input = False
            
            if return_code != 0:
                self.log(f"Script {script} exited with non-zero status code: {return_code}")
                if self.status not in ("killed", "stopped"):
                    # Exit code 2 = user-initiated Quit (Quality Gate). Show as
                    # 'stopped' rather than 'failed' so the dashboard badge is correct.
                    if return_code == 2:
                        self.status = "stopped"
                    else:
                        self.status = "failed"
                return
            
            self.log(f"Step {self.current_step} completed successfully.")

        if self.status not in ("killed", "stopped"):
            self.status = "success"
            # Print execution summary & adjust status if no candidates were found/processed (only for referral and recruiter pipelines)
            if "referral_pipeline" in self.task_id or "recruiter_pipeline" in self.task_id:
                try:
                    from core.storage.database import load_all_referrals
                    referrals = load_all_referrals()
                    today_str = datetime.now().strftime("%Y-%m-%d")
                    
                    # Count sent today
                    emp_messages_sent = sum(
                        1 for r in referrals
                        if str(r.get('Referral_Status') or '').strip().lower() == 'sent'
                        and str(r.get('Sent_Time') or '').strip().startswith(today_str)
                        and str(r.get('Referral_Source') or '').strip() == 'Existing Employee'
                    )
                    emp_connections_sent = sum(
                        1 for r in referrals
                        if str(r.get('Referral_Status') or '').strip().lower() == 'sent'
                        and str(r.get('Sent_Time') or '').strip().startswith(today_str)
                        and str(r.get('Referral_Source') or '').strip() == 'Sent Employee Connection'
                    )
                    rec_messages_sent = sum(
                        1 for r in referrals
                        if str(r.get('Referral_Status') or '').strip().lower() == 'sent'
                        and str(r.get('Sent_Time') or '').strip().startswith(today_str)
                        and str(r.get('Referral_Source') or '').strip() == 'Existing Recruiter'
                    )
                    rec_connections_sent = sum(
                        1 for r in referrals
                        if str(r.get('Referral_Status') or '').strip().lower() == 'sent'
                        and str(r.get('Sent_Time') or '').strip().startswith(today_str)
                        and str(r.get('Referral_Source') or '').strip() == 'Sent Recruiter Connection'
                    )
                    
                    total_sent = emp_messages_sent + emp_connections_sent + rec_messages_sent + rec_connections_sent
                    
                    self.log("=" * 60)
                    self.log("--- Pipeline Execution Summary ---")
                    self.log(f"Existing Employee Messages Sent   : {emp_messages_sent}")
                    self.log(f"Employee Connection Requests Sent : {emp_connections_sent}")
                    self.log(f"Existing Recruiter Messages Sent   : {rec_messages_sent}")
                    self.log(f"Recruiter Connection Requests Sent : {rec_connections_sent}")
                    self.log(f"Total Outreach Actions Sent Today : {total_sent}")
                    
                    # Load selected config targets for the pinned user
                    from config.user_profiles import load_all_configs
                    all_cfg = load_all_configs()
                    user_conf = all_cfg.get("users", {}).get(self.username, {})
                    connect_conf = user_conf.get("linkedin_connect", {})
                    max_connections = int(connect_conf.get("max_connections_per_run") or 5)
                    
                    if total_sent > 0:
                        self.log("Status: Completed Successfully")
                    else:
                        self.log("Status: Completed – No Candidates Available")
                    self.log("=" * 60)
                except Exception as e:
                    self.log(f"Warning: error compiling execution summary: {e}")

    def send_input(self, text):
        if self.process and self.process.stdin:
            try:
                self.process.stdin.write(text + "\n")
                self.process.stdin.flush()
                self.log(f"Piped input: {text}")
                self.waiting_for_input = False
                return True
            except Exception as e:
                self.log(f"Failed to write to stdin: {e}")
        return False

    def kill(self):
        self.status = "stopped"
        self.waiting_for_input = False
        if self.process:
            try:
                if sys.platform == 'win32':
                    import subprocess as sp
                    sp.run(["taskkill", "/F", "/T", "/PID", str(self.process.pid)], capture_output=True, check=False)
                else:
                    os.killpg(os.getpgid(self.process.pid), signal.SIGTERM)
                self.log("Process terminated by user.")
            except Exception as e:
                self.log(f"Error terminating process: {e}")


def get_excel_data(file_path):
    if not os.path.exists(file_path):
        return []
    try:
        df = pd.read_excel(file_path)
        df = df.fillna("")
        return df.to_dict(orient="records")
    except Exception as e:
        print(f"Error reading {file_path}: {e}")
        return []


@app.route('/')
def home():
    return render_template('index.html')

def index():
    return render_template('index.html')

@app.route('/dashboard')
def dashboard():
    return render_template('dashboard.html')

@app.after_request
def add_header(response):
    if request.path.startswith('/api/'):
        response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, post-check=0, pre-check=0, max-age=0'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '-1'
    return response

@app.route('/api/stats')
def get_stats():
    from core.storage.engine import read_database_rows
    # Always read from disk so the dashboard reflects pipeline updates immediately
    job_tracker = read_database_rows("emails", bypass_cache=True)
    job_leads = read_database_rows("jobs", bypass_cache=True)

    total_emails = len(job_tracker)
    emails_sent = sum(1 for r in job_tracker if str(r.get('Status')).strip().lower() == 'sent')

    total_leads = len(job_leads)

    referral_requests_sent = sum(1 for r in job_leads if str(r.get('Status')).strip().lower() in ('ask for referral', 'asked for referral', 'done', 'referred'))
    done_referrals = sum(1 for r in job_leads if str(r.get('Status')).strip().lower() == 'done')

    return jsonify({
        "total_emails_scraped": total_emails,
        "emails_sent": emails_sent,
        "total_jobs_scraped": total_leads,
        "referral_requests_sent": referral_requests_sent,
        "done_jobs_count": done_referrals
    })


@app.route('/api/email_stats')
def email_stats():
    return jsonify(get_email_metrics())

@app.route('/api/company_stats')
def company_stats():
    return jsonify(get_company_metrics())

@app.route('/api/outreach_stats')
def outreach_stats():
    return jsonify(get_outreach_metrics())

@app.route('/api/data/job_tracker')
def job_tracker_data():
    from core.storage.engine import read_database_rows
    rows = read_database_rows("emails", bypass_cache=True)
    rows.sort(key=lambda x: str(x.get('Timestamp') or ''), reverse=True)
    return jsonify(rows)


@app.route('/api/cache/invalidate', methods=['GET', 'POST'])
def invalidate_cache():
    """Called by background pipeline processes after writing data so the Flask
    in-memory cache is cleared and the next UI request reads fresh data from disk."""
    from core.storage.engine import _invalidate_cached_rows
    username = get_selected_user_name()
    _invalidate_cached_rows(username)  # clears all table caches for the active user
    return jsonify({"status": "ok", "message": f"Cache invalidated for '{username}'"})

@app.route('/api/data/job_leads')
def job_leads_data():
    from core.storage.database import load_job_leads_with_referral_counts
    leads = load_job_leads_with_referral_counts()
    leads.sort(key=lambda x: str(x.get('CreatedDateTime') or ''), reverse=True)
    return jsonify(leads)


@app.route('/api/run/scraper', methods=['POST'])
def start_scraper():
    with task_lock:
        body = request.get_json(silent=True) or {}
        # Use username from request body, or fall back to the currently selected user
        username = body.get("username") or get_selected_user_name()
        phase = body.get("phase", "full")
        task_id = f"{username}::scraper_pipeline::{phase}"
        
        # Check if this specific phase is already running or queued, or if the full pipeline is running
        full_task_id = f"{username}::scraper_pipeline::full"
        if task_id in active_tasks and active_tasks[task_id].status in ("running", "queued"):
            return jsonify({"status": "error", "message": f"Pipeline phase '{phase}' is already running or queued."}), 400
        if phase != "full" and full_task_id in active_tasks and active_tasks[full_task_id].status in ("running", "queued"):
            return jsonify({"status": "error", "message": "The full scraper pipeline is already running."}), 400
        if phase == "full":
            for p in ("phase1", "phase2"):
                p_tid = f"{username}::scraper_pipeline::{p}"
                if p_tid in active_tasks and active_tasks[p_tid].status in ("running", "queued"):
                    return jsonify({"status": "error", "message": f"Cannot run full pipeline while phase '{p}' is running."}), 400
        
        # Route each phase to its dedicated runner script for clean separation of concerns.
        # The combined run_email_outreach.py is used only for the full pipeline run.
        if phase == "phase1":
            commands = [("run_email_scraper.py", [])]
        elif phase == "phase2":
            commands = [("run_email_sender.py", [])]
        else:
            commands = [("run_email_outreach.py", [])]
        
        runner = SubprocessRunner(task_id, commands, username)
        active_tasks[task_id] = runner
        runner.start()
        
        return jsonify({"status": "success", "task_id": task_id})


@app.route('/api/run/referral', methods=['POST'])
def start_referral():
    with task_lock:
        body = request.get_json() or {}
        username = body.get("username") or get_selected_user_name()
        step = body.get("step")
        step_suffix = f"step{step}" if step is not None else "full"
        task_id = f"{username}::referral_pipeline::{step_suffix}"
        
        full_task_id = f"{username}::referral_pipeline::full"
        if task_id in active_tasks and active_tasks[task_id].status in ("running", "queued"):
            return jsonify({"status": "error", "message": "This referral step/pipeline is already running or queued."}), 400
        if step_suffix != "full" and full_task_id in active_tasks and active_tasks[full_task_id].status in ("running", "queued"):
            return jsonify({"status": "error", "message": "The full referral pipeline is already running."}), 400
        if step_suffix == "full":
            for idx in range(1, 7):
                s_tid = f"{username}::referral_pipeline::step{idx}"
                if s_tid in active_tasks and active_tasks[s_tid].status in ("running", "queued"):
                    return jsonify({"status": "error", "message": f"Cannot run full pipeline while step {idx} is running."}), 400
        
        all_commands = [
            ("run_job_search.py", []),
            ("run_referral_review.py", []),
            ("run_referral_outreach_discover.py", []),
            ("run_referral_outreach_send.py", []),
            ("run_url_shortener.py", []),
            ("run_linkedin_connect.py", [])
        ]
        
        if step is not None:
            try:
                step_idx = int(step) - 1
                if step_idx < 0 or step_idx >= len(all_commands):
                    return jsonify({"status": "error", "message": f"Invalid step: {step}"}), 400
                commands = [all_commands[step_idx]]
            except ValueError:
                return jsonify({"status": "error", "message": f"Step must be an integer: {step}"}), 400
        else:
            commands = all_commands
        
        runner = SubprocessRunner(task_id, commands, username)
        active_tasks[task_id] = runner
        runner.start()
        
        return jsonify({"status": "success", "task_id": task_id})


@app.route('/api/run/recruiter', methods=['POST'])
def start_recruiter():
    with task_lock:
        body = request.get_json() or {}
        username = body.get("username") or get_selected_user_name()
        step = body.get("step")
        step_suffix = f"step{step}" if step is not None else "full"
        task_id = f"{username}::recruiter_pipeline::{step_suffix}"
        
        full_task_id = f"{username}::recruiter_pipeline::full"
        if task_id in active_tasks and active_tasks[task_id].status in ("running", "queued"):
            return jsonify({"status": "error", "message": "This recruiter step/pipeline is already running or queued."}), 400
        if step_suffix != "full" and full_task_id in active_tasks and active_tasks[full_task_id].status in ("running", "queued"):
            return jsonify({"status": "error", "message": "The full recruiter pipeline is already running."}), 400
        if step_suffix == "full":
            for idx in range(1, 4):
                s_tid = f"{username}::recruiter_pipeline::step{idx}"
                if s_tid in active_tasks and active_tasks[s_tid].status in ("running", "queued"):
                    return jsonify({"status": "error", "message": f"Cannot run full pipeline while step {idx} is running."}), 400
        
        all_commands = [
            ("run_recruiter_outreach_discover.py", []),
            ("run_recruiter_outreach_send.py", []),
            ("run_recruiter_outreach.py", [])
        ]
        
        if step is not None:
            try:
                step_idx = int(step) - 1
                if step_idx < 0 or step_idx >= len(all_commands):
                    return jsonify({"status": "error", "message": f"Invalid step: {step}"}), 400
                commands = [all_commands[step_idx]]
            except ValueError:
                return jsonify({"status": "error", "message": f"Step must be an integer: {step}"}), 400
        else:
            commands = all_commands
        
        runner = SubprocessRunner(task_id, commands, username)
        active_tasks[task_id] = runner
        runner.start()
        
        return jsonify({"status": "success", "task_id": task_id})


@app.route('/api/tasks')
def get_all_tasks():
    with task_lock:
        result = {}
        for tid, runner in active_tasks.items():
            total_steps = len(runner.commands)
            if "scraper_pipeline" in tid and "full" in tid:
                total_steps = 2
            result[tid] = {
                "status": runner.status,
                "username": runner.username,
                "waiting_for_input": runner.waiting_for_input,
                "current_step": runner.current_step,
                "total_steps": total_steps,
                "start_time": runner.start_time
            }
        return jsonify(result)


@app.route('/api/task/<task_id>/logs')
def get_task_logs(task_id):
    with task_lock:
        if task_id not in active_tasks:
            return jsonify({"status": "error", "message": "Task not found"}), 404
        
        runner = active_tasks[task_id]
        current_step_name = None
        args = []
        if runner.current_step > 0 and runner.current_step <= len(runner.commands):
            cmd_tuple = runner.commands[runner.current_step - 1]
            current_step_name = cmd_tuple[0]
            args = cmd_tuple[1]
        
        return jsonify({
            "status": runner.status,
            "waiting_for_input": runner.waiting_for_input,
            "logs": runner.logs,
            "current_step_name": current_step_name,
            "args": args,
            "is_single_step": len(runner.commands) == 1
        })


@app.route('/api/task/<task_id>/input', methods=['POST'])
def send_task_input(task_id):
    with task_lock:
        if task_id not in active_tasks:
            return jsonify({"status": "error", "message": "Task not found"}), 404
        
        body = request.get_json() or {}
        val = body.get("input", "")
        
        runner = active_tasks[task_id]
        success = runner.send_input(val)
        
        if success:
            return jsonify({"status": "success"})
        else:
            return jsonify({"status": "error", "message": "Failed to send input"}), 500


@app.route('/api/task/<task_id>/kill', methods=['POST'])
def kill_task(task_id):
    with task_lock:
        if task_id not in active_tasks:
            return jsonify({"status": "error", "message": "Task not found"}), 404
        
        runner = active_tasks[task_id]
        runner.kill()
        return jsonify({"status": "success"})


@app.route('/api/users', methods=['GET'])
def get_users_list():
    config = load_all_configs()
    users = list(config.get("users", {}).keys())
    selected = config.get("selected_user", "")
    
    # Extract profile names for all users
    user_details = {}
    for username, data in config.get("users", {}).items():
        profile = data.get("profile", {})
        user_details[username] = {
            "first_name": profile.get("first_name", ""),
            "last_name": profile.get("last_name", "")
        }
        
    return jsonify({
        "users": users,
        "selected_user": selected,
        "user_details": user_details
    })


@app.route('/api/users/select', methods=['POST'])
def select_active_user():
    body = request.get_json() or {}
    user = body.get("user")
    
    # Validate user exists by checking the filesystem (fast, no Sheets calls)
    users_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "users")
    if not user or not os.path.isdir(os.path.join(users_dir, user)):
        return jsonify({"status": "error", "message": "User not found"}), 404
    
    # Write ONLY the active_user.json file — fast, no full config re-save
    active_user_file = os.path.join(users_dir, "active_user.json")
    try:
        with open(active_user_file, "w") as f:
            json.dump({"selected_user": user}, f, indent=2)
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500
    
    # Invalidate the in-memory config cache for the newly selected user
    try:
        from core.storage.engine import _invalidate_cached_config
        _invalidate_cached_config(user)
    except Exception:
        pass
        
    # Auto-sync sheets to local in background (non-blocking) so the switch response returns instantly
    def _bg_sync(username):
        try:
            sync_google_sheets_to_local_if_empty(username)
        except Exception as se:
            print(f"[User Switch] Warning: Sheets auto-sync failed for user '{username}': {se}")

    t = threading.Thread(target=_bg_sync, args=(user,), daemon=True)
    t.start()

    return jsonify({"status": "success", "selected_user": user})


@app.route('/api/users/create', methods=['POST'])
def create_user_profile():
    body = request.get_json() or {}
    username = body.get("username", "").strip()
    if not username:
        return jsonify({"status": "error", "message": "Username is required"}), 400
    
    config = load_all_configs()
    if username in config.get("users", {}):
        return jsonify({"status": "error", "message": "User already exists"}), 400
        
    config["users"][username] = {
        "profile": {
            "first_name": username,
            "last_name": "",
            "email": "",
            "phone": "",
            "resume_name": "",
            "resume_url": "",
            "current_location": "",
            "preferred_locations": "",
            "experience": "",
            "linkedin_url": "",
            "current_ctc": "",
            "expected_ctc": ""
        },
        "email_scraper": {
            "email_template": DEFAULT_EMAIL_TEMPLATE,
            "email_subject": "",
            "search_keywords": [],
            "title_keywords": [],
            "keywords": [],
            "excluded_keywords": [],
            "sender_email": "",
            "interval": "60",
            "review_mode": True,
            "max_emails_per_run": "5",
            "filter_experience_enabled": False,
            "filter_experience_ranges": [],
            "filter_location_enabled": False,
            "filter_locations": [],
        },
        "linkedin_connect": {
            "message_template": DEFAULT_CONNECTION_TEMPLATE,
            "search_keywords": [],
            "title_keywords": [],
            "keywords": [],
            "excluded_keywords": [],
            "interval": "60",
            "search_pages": 2,
            "review_mode": True,
            "max_connections_per_company": "5",
            "max_connections_per_run": "5"
        },
        "recruiter_outreach": {
            "message_template": "Hi {first_name}, let's connect! I saw you handle Talent Acquisition at {company}. I am interested in opportunities there. My resume: {resume}",
            "interval": "120",
            "daily_limit": "5",
            "target_count": "2",
            "review_mode": True
        },
        "referral_outreach": {
            "message_template": "Hi {RECEIVER_NAME},\n\nI hope you're doing well! I saw we're connected on LinkedIn and noticed you work at {COMPANY}.\n\nI'm interested in a role there and would love your guidance or a referral if possible.\n\nJob: {JOB_URL}\nResume: {RESUME}\n\nThank you!",
            "interval": "60",
            "max_referrals_per_run": "5",
            "review_mode": True
        }
    }
    # Do NOT auto-switch to the new user — preserve the current active user session
    save_all_configs(config)
    return jsonify({"status": "success", "username": username, "selected_user": config.get("selected_user", username)})



@app.route('/api/users/delete', methods=['POST'])
def delete_user_profile():
    body = request.get_json() or {}
    username = body.get("username", "").strip()
    if not username:
        return jsonify({"status": "error", "message": "Username is required"}), 400
        
    active_user = get_selected_user_name()
    users_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "users")
    user_folder = os.path.join(users_dir, username)
    
    if not os.path.exists(user_folder) or not os.path.isdir(user_folder):
        return jsonify({"status": "error", "message": "Profile folder not found"}), 404
        
    all_users = [d for d in os.listdir(users_dir) if os.path.isdir(os.path.join(users_dir, d)) and d != "default"]
    
    if len(all_users) <= 1:
        return jsonify({"status": "error", "message": "Cannot delete the only profile. Please create another profile first."}), 400
        
    new_active = active_user
    if username == active_user:
        other_users = [u for u in all_users if u != username]
        new_active = other_users[0]
        active_user_file = os.path.join(users_dir, "active_user.json")
        try:
            with open(active_user_file, "w") as f:
                json.dump({"selected_user": new_active}, f, indent=2)
        except Exception:
            pass
            
    import shutil
    try:
        shutil.rmtree(user_folder, ignore_errors=True)
    except Exception as e:
        return jsonify({"status": "error", "message": f"Failed to delete directory: {str(e)}"}), 500
        
    try:
        from core.storage.engine import _invalidate_cached_config
        _invalidate_cached_config(username)
    except Exception:
        pass
        
    return jsonify({"status": "success", "switched_to": new_active if username == active_user else None})



@app.route('/api/config/sheets/test', methods=['POST'])
def test_sheets_connection():
    body = request.get_json() or {}
    url = body.get("google_sheet_url", "").strip()
    creds = body.get("google_credentials_json", "").strip()
    
    if not url:
        return jsonify({"status": "error", "message": "Google Sheet URL is required"}), 400
    if not creds:
        return jsonify({"status": "error", "message": "Google Credentials JSON content is required"}), 400
        
    try:
        from core.storage.sheets import ensure_worksheets_exist
        ensure_worksheets_exist(url, creds)
        return jsonify({"status": "success"})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 400


@app.route('/api/config/storage/switch', methods=['POST'])
def switch_storage_type():
    """Streams migration progress via Server-Sent Events (SSE).
    
    When a user changes database_type in Settings, this endpoint:
      - Detects the switch direction (local→sheets or sheets→local)
      - Runs the appropriate migration inline with live log streaming
      - Saves the new config after successful migration
    
    The client polls this endpoint using EventSource / fetch+ReadableStream.
    Each SSE event is a JSON line: {"type": "log"|"done"|"error", "message": "..."}
    """
    body = request.get_json() or {}
    new_type = body.get("database_type", "").strip()       # "local" or "google_sheets"
    global_settings_payload = body.get("global_settings") or {}

    if new_type not in ("local", "google_sheets"):
        return jsonify({"status": "error", "message": "Invalid database_type value"}), 400

    def generate():
        import traceback, time as _time

        def emit(event_type, message):
            """Yield a Server-Sent Events line."""
            import json as _json
            yield f"data: {_json.dumps({'type': event_type, 'message': message})}\n\n"

        try:
            from config.user_profiles import (
                load_all_configs, save_all_configs, get_selected_user_name, get_global_settings
            )
            from config.settings import get_job_tracker_file, get_job_leads_file, get_referrals_file
            from config.constants import GOOGLE_SHEET_WORKSHEETS

            username = get_selected_user_name()
            current_type = get_global_settings().get("database_type", "local")

            yield from emit("log", f"Active profile: {username}")
            yield from emit("log", f"Current storage: {current_type}  →  New storage: {new_type}")

            if current_type == new_type:
                if new_type == "local":
                    yield from emit("log", "Storage type unchanged (Local). Saving config...")
                    # Still persist the rest of the settings
                    config = load_all_configs()
                    config["global_settings"] = {**config.get("global_settings", {}), **global_settings_payload, "database_type": new_type}
                    save_all_configs(config)
                    yield from emit("done", "Configuration saved. No migration needed.")
                    return
                else:
                    yield from emit("log", "Storage type unchanged (Google Sheets). Initiating synchronization run...")

            # ---------------------------------------------------------------
            # LOCAL  →  GOOGLE SHEETS
            # ---------------------------------------------------------------
            if new_type == "google_sheets":
                yield from emit("log", "Direction: Local Excel → Google Sheets")

                sheet_url = global_settings_payload.get("google_sheet_url", "").strip()
                creds_content = global_settings_payload.get("google_credentials_json", "").strip()

                if not sheet_url or not creds_content:
                    yield from emit("error", "Google Sheet URL and Credentials JSON are required to switch to Google Sheets storage.")
                    return

                yield from emit("log", "Step 1: Authenticating with Google APIs...")
                try:
                    import gspread, json as _json2
                    from google.oauth2.service_account import Credentials
                    scopes = [
                        "https://spreadsheets.google.com/feeds",
                        "https://www.googleapis.com/auth/drive"
                    ]
                    creds_dict = _json2.loads(creds_content)
                    credentials = Credentials.from_service_account_info(creds_dict, scopes=scopes)
                    client = gspread.authorize(credentials)
                    sh = client.open_by_url(sheet_url)
                    yield from emit("log", f"Authentication OK. Opened sheet: '{sh.title}'")
                except Exception as e:
                    yield from emit("error", f"Authentication failed: {e}")
                    return

                # Bootstrap settings locally so engine can access sheets
                yield from emit("log", "Bootstrapping Google Sheets credentials locally...")
                config = load_all_configs()
                if "global_settings" not in config:
                    config["global_settings"] = {}
                config["global_settings"].update(global_settings_payload)
                save_all_configs(config)

                # Check if User Profile sheet exists and has data (indicating pre-existing cloud settings)
                has_existing_cloud_data = False
                profile_ws_name = GOOGLE_SHEET_WORKSHEETS["profile"]["name"]
                try:
                    pws = sh.worksheet(profile_ws_name)
                    if pws.row_values(2):
                        has_existing_cloud_data = True
                except Exception:
                    pass

                local_paths = {
                    "Job Leads": get_job_leads_file(),
                    "Scraped Emails": get_job_tracker_file(),
                    "Referrals & Connections": get_referrals_file()
                }
                id_cols = {
                    "Job Leads": "JobID",
                    "Scraped Emails": "ID",
                    "Referrals & Connections": "ReferralID"
                }

                if has_existing_cloud_data:
                    yield from emit("log", "Google Sheet contains pre-existing profile settings. Performing PULL migration...")
                    
                    # 1. Pull settings from Google Sheets
                    try:
                        from core.storage.engine import GoogleSheetsStorageProvider
                        cloud_config = GoogleSheetsStorageProvider().get_config(username, bypass_cache=True)
                        if cloud_config:
                            # Update local config with cloud settings
                            config = load_all_configs()
                            if "users" not in config:
                                config["users"] = {}
                            
                            # Merge cloud settings
                            user_config = cloud_config.get("users", {}).get(username, cloud_config)
                            config["users"][username] = user_config
                            # Retain sheet settings
                            config["global_settings"].update(cloud_config.get("global_settings", {}))
                            save_all_configs(config)
                            yield from emit("log", "  Successfully downloaded user profile settings from Google Sheets.")
                    except Exception as e:
                        yield from emit("log", f"  Warning: Could not pull cloud settings: {e}")

                    # 2. Pull database rows from Google Sheets (bidirectional sync)
                    import openpyxl as _xl
                    yield from emit("log", "Step 2: Syncing database rows bidirectionally...")
                    for key, info in GOOGLE_SHEET_WORKSHEETS.items():
                        if key not in ("jobs", "emails", "referrals"):
                            continue
                        ws_name = info["name"]
                        headers = info["headers"]
                        local_file = local_paths[ws_name]
                        id_col = id_cols[ws_name]

                        yield from emit("log", f"  Processing '{ws_name}'...")
                        _time.sleep(1.0)

                        try:
                            ws = sh.worksheet(ws_name)
                            cloud_rows = ws.get_all_records()
                        except Exception as e:
                            yield from emit("log", f"  Warning: Could not read cloud sheet '{ws_name}': {e}")
                            continue

                        local_rows = []
                        local_ids = set()
                        if os.path.exists(local_file):
                            try:
                                wb = _xl.load_workbook(local_file)
                                lws = wb.active
                                col_map = {cell.value: idx for idx, cell in enumerate(lws[1], start=1)}
                                id_idx = col_map.get(id_col)
                                for row in range(2, lws.max_row + 1):
                                    row_id = str(lws.cell(row=row, column=id_idx).value or "").rstrip(".0") if id_idx else ""
                                    if row_id:
                                        local_ids.add(row_id)
                                        row_data = {}
                                        for h, ci in col_map.items():
                                            row_data[h] = lws.cell(row=row, column=ci).value or ""
                                        local_rows.append(row_data)
                            except Exception as e:
                                yield from emit("log", f"  Warning: Could not read local file for '{ws_name}': {e}")

                        # Pull missing rows from Cloud to Local
                        new_local_count = 0
                        try:
                            if not os.path.exists(local_file):
                                wb = _xl.Workbook()
                                lws = wb.active
                                lws.title = ws_name.replace(" & ", " ")
                                lws.append(headers)
                            else:
                                wb = _xl.load_workbook(local_file)
                                lws = wb.active

                            for r in cloud_rows:
                                r_id = str(r.get(id_col, "") or "").rstrip(".0")
                                if r_id and r_id not in local_ids:
                                    lws.append([str(r.get(h, "") or "") for h in headers])
                                    local_ids.add(r_id)
                                    new_local_count += 1
                            if new_local_count > 0:
                                wb.save(local_file)
                                yield from emit("log", f"    Downloaded {new_local_count} rows from Google Sheets to local.")
                        except Exception as e:
                            yield from emit("log", f"    Warning: Failed to save pulled rows locally: {e}")

                        # Push missing rows from Local to Cloud
                        new_cloud_rows = []
                        for r in local_rows:
                            r_id = str(r.get(id_col, "") or "").rstrip(".0")
                            cloud_ids = {str(cr.get(id_col, "") or "").rstrip(".0") for cr in cloud_rows}
                            if r_id and r_id not in cloud_ids:
                                new_cloud_rows.append(r)

                        if new_cloud_rows:
                            try:
                                for r in new_cloud_rows:
                                    row_vals = [str(r.get(h, "") or "") for h in headers]
                                    ws.append_row(row_vals)
                                    _time.sleep(0.3)
                                yield from emit("log", f"    Uploaded {len(new_cloud_rows)} local rows to Google Sheets.")
                            except Exception as e:
                                yield from emit("log", f"    Warning: Upload error for '{ws_name}': {e}")

                        try:
                            from core.storage.sheets import _cache_invalidate
                            _cache_invalidate(ws_name)
                        except Exception:
                            pass

                else:
                    yield from emit("log", "Google Sheet is empty. Performing PUSH migration (local -> cloud)...")
                    import openpyxl as _xl
                    yield from emit("log", "Step 2: Migrating local Excel data to Google Sheets...")

                    try:
                        from core.storage.sheets import ensure_worksheets_exist
                        ensure_worksheets_exist(sheet_url, creds_content)
                    except Exception as e:
                        yield from emit("log", f"  Warning: Sheets initialization failed: {e}")

                    for key, info in GOOGLE_SHEET_WORKSHEETS.items():
                        if key not in ("jobs", "emails", "referrals"):
                            continue
                        ws_name = info["name"]
                        headers = info["headers"]
                        local_file = local_paths[ws_name]
                        id_col = id_cols[ws_name]

                        yield from emit("log", f"  Processing '{ws_name}'...")
                        _time.sleep(1.5)

                        try:
                            ws = sh.worksheet(ws_name)
                            cloud_rows = ws.get_all_records()
                        except Exception as e:
                            yield from emit("log", f"  Warning: Could not read worksheet '{ws_name}': {e}")
                            cloud_rows = []

                        cloud_ids = {str(r.get(id_col, "") or "").rstrip(".0") for r in cloud_rows}

                        if not os.path.exists(local_file):
                            yield from emit("log", f"  No local file found for '{ws_name}'. Skipping.")
                            continue

                        try:
                            wb = _xl.load_workbook(local_file)
                            lws = wb.active
                            col_map = {cell.value: idx for idx, cell in enumerate(lws[1], start=1)}
                            id_idx = col_map.get(id_col)
                            new_rows = []
                            for row in range(2, lws.max_row + 1):
                                row_id = str(lws.cell(row=row, column=id_idx).value or "").rstrip(".0") if id_idx else ""
                                if row_id and row_id not in cloud_ids:
                                    row_data = {}
                                    for h, ci in col_map.items():
                                        row_data[h] = lws.cell(row=row, column=ci).value or ""
                                    new_rows.append(row_data)
                        except Exception as e:
                            yield from emit("log", f"  Warning: Could not read local file for '{ws_name}': {e}")
                            continue

                        yield from emit("log", f"  Local rows: {lws.max_row - 1}  |  Cloud rows: {len(cloud_rows)}  |  New to upload: {len(new_rows)}")

                        if new_rows:
                            try:
                                for r in new_rows:
                                    row_vals = [str(r.get(h, "") or "") for h in headers]
                                    ws.append_row(row_vals)
                                    _time.sleep(0.3)
                                yield from emit("log", f"  Uploaded {len(new_rows)} new rows to '{ws_name}'.")
                            except Exception as e:
                                yield from emit("log", f"  Warning: Upload error for '{ws_name}': {e}")
                        else:
                            yield from emit("log", f"  '{ws_name}' already up to date.")

                        try:
                            from core.storage.sheets import _cache_invalidate
                            _cache_invalidate(ws_name)
                        except Exception:
                            pass

                    # Push configuration to Google Sheets
                    yield from emit("log", "Step 3: Uploading local configuration settings to Google Sheets...")
                    try:
                        from core.storage.engine import GoogleSheetsStorageProvider
                        config_to_save = load_all_configs()
                        user_config_dict = config_to_save.get("users", {}).get(username, config_to_save)
                        GoogleSheetsStorageProvider().save_config(username, user_config_dict)
                        yield from emit("log", "  Successfully uploaded local configuration to Google Sheets.")
                    except Exception as e:
                        yield from emit("log", f"  Warning: Could not save config to Google Sheets: {e}")

            # ---------------------------------------------------------------
            # GOOGLE SHEETS  →  LOCAL
            # ---------------------------------------------------------------
            else:  # new_type == "local"
                yield from emit("log", "Direction: Google Sheets → Local Excel")

                current_settings = get_global_settings()
                sheet_url = current_settings.get("google_sheet_url", "").strip()
                creds_content = current_settings.get("google_credentials_json", "").strip()

                if not sheet_url or not creds_content:
                    yield from emit("log", "No Google Sheets credentials found. Switching to local without migration.")
                else:
                    yield from emit("log", "Step 1: Authenticating with Google APIs...")
                    try:
                        import gspread, json as _json3
                        from google.oauth2.service_account import Credentials
                        scopes = [
                            "https://spreadsheets.google.com/feeds",
                            "https://www.googleapis.com/auth/drive"
                        ]
                        creds_dict = _json3.loads(creds_content)
                        credentials = Credentials.from_service_account_info(creds_dict, scopes=scopes)
                        client = gspread.authorize(credentials)
                        sh = client.open_by_url(sheet_url)
                        yield from emit("log", f"Authentication OK. Opened sheet: '{sh.title}'")
                    except Exception as e:
                        yield from emit("error", f"Authentication failed: {e}")
                        return

                    local_paths = {
                        "Job Leads": get_job_leads_file(),
                        "Scraped Emails": get_job_tracker_file(),
                        "Referrals & Connections": get_referrals_file()
                    }
                    id_cols = {
                        "Job Leads": "JobID",
                        "Scraped Emails": "ID",
                        "Referrals & Connections": "ReferralID"
                    }

                    import openpyxl as _xl2

                    yield from emit("log", "Step 2: Pulling Google Sheets data into local Excel files...")
                    for key, info in GOOGLE_SHEET_WORKSHEETS.items():
                        if key not in ("jobs", "emails", "referrals"):
                            continue
                        ws_name = info["name"]
                        headers = info["headers"]
                        local_file = local_paths[ws_name]
                        id_col = id_cols[ws_name]

                        yield from emit("log", f"  Processing '{ws_name}'...")
                        _time.sleep(1.5)

                        try:
                            ws = sh.worksheet(ws_name)
                            cloud_rows = ws.get_all_records()
                        except Exception as e:
                            yield from emit("log", f"  Warning: Could not read '{ws_name}': {e}")
                            continue

                        if not cloud_rows:
                            yield from emit("log", f"  No data in '{ws_name}'. Skipping.")
                            continue

                        # Load existing local IDs
                        existing_ids = set()
                        if os.path.exists(local_file):
                            try:
                                wb = _xl2.load_workbook(local_file)
                                lws = wb.active
                                col_map = {cell.value: idx for idx, cell in enumerate(lws[1], start=1)}
                                id_idx = col_map.get(id_col)
                                if id_idx:
                                    for row in range(2, lws.max_row + 1):
                                        val = lws.cell(row=row, column=id_idx).value
                                        if val is not None:
                                            existing_ids.add(str(val).rstrip(".0"))
                            except Exception:
                                pass

                        new_rows = [
                            r for r in cloud_rows
                            if str(r.get(id_col, "") or "").rstrip(".0") not in existing_ids
                            and str(r.get(id_col, "") or "").strip()
                        ]

                        yield from emit("log", f"  Cloud rows: {len(cloud_rows)}  |  Local existing: {len(existing_ids)}  |  New to write: {len(new_rows)}")

                        if not new_rows:
                            yield from emit("log", f"  Local already up to date.")
                            continue

                        try:
                            if os.path.exists(local_file):
                                wb = _xl2.load_workbook(local_file)
                                lws = wb.active
                            else:
                                wb = _xl2.Workbook()
                                lws = wb.active
                                lws.title = ws_name.replace(" & ", " ")
                                lws.append(headers)

                            for r in new_rows:
                                lws.append([str(r.get(h, "") or "") for h in headers])
                            wb.save(local_file)
                            yield from emit("log", f"  Written {len(new_rows)} rows to local file.")
                        except Exception as e:
                            yield from emit("log", f"  Warning: Could not write to local file for '{ws_name}': {e}")

            # ---------------------------------------------------------------
            # Save updated config after successful migration
            # ---------------------------------------------------------------
            yield from emit("log", "Step 3: Saving new storage configuration...")
            config = load_all_configs()
            existing_gs = config.get("global_settings", {})
            merged_gs = {**existing_gs, **global_settings_payload, "database_type": new_type}
            config["global_settings"] = merged_gs
            save_all_configs(config)
            yield from emit("done", f"Migration complete! Active storage is now: {new_type.replace('_', ' ').title()}")

        except GeneratorExit:
            return
        except Exception as e:
            yield from emit("error", f"Unexpected error during migration: {traceback.format_exc()}")

    return Response(generate(), mimetype="text/event-stream",
                    headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"})




@app.route('/api/system/update/check', methods=['GET'])
def check_system_updates():
    try:
        import subprocess
        repo_dir = os.path.dirname(os.path.abspath(__file__))
        
        # 1. Run git fetch origin
        fetch_res = subprocess.run(["git", "fetch", "origin"], cwd=repo_dir, capture_output=True, text=True, timeout=15)
        if fetch_res.returncode != 0:
            return jsonify({
                "status": "error",
                "message": f"Git fetch failed: {fetch_res.stderr.strip()}"
            }), 500
            
        # 2. Get local commit hash and description
        local_hash = subprocess.run(["git", "rev-parse", "--short", "HEAD"], cwd=repo_dir, capture_output=True, text=True).stdout.strip()
        local_desc = subprocess.run(["git", "log", "-1", "--format=%s (%cd)", "--date=short", "HEAD"], cwd=repo_dir, capture_output=True, text=True).stdout.strip()
        
        # 3. Get remote tracking commit hash
        remote_hash = subprocess.run(["git", "rev-parse", "--short", "origin/main"], cwd=repo_dir, capture_output=True, text=True).stdout.strip()
        
        # 4. Check commits behind
        behind_log = subprocess.run(["git", "log", "HEAD..origin/main", "--oneline"], cwd=repo_dir, capture_output=True, text=True).stdout.strip()
        
        commits_behind = []
        if behind_log:
            commits_behind = [c.strip() for c in behind_log.split("\n") if c.strip()]
            
        updates_available = len(commits_behind) > 0
        
        return jsonify({
            "status": "success",
            "updates_available": updates_available,
            "current_commit": local_hash,
            "current_desc": local_desc,
            "latest_commit": remote_hash,
            "behind_by_commits": len(commits_behind),
            "commits_list": commits_behind
        })
    except Exception as e:
        return jsonify({
            "status": "error",
            "message": f"Unexpected error checking updates: {str(e)}"
        }), 500


@app.route('/api/system/update/pull', methods=['POST'])
def pull_system_updates():
    try:
        import subprocess
        repo_dir = os.path.dirname(os.path.abspath(__file__))
        
        # 1. Stash any local modifications to ensure clean merge/pull
        subprocess.run(["git", "stash"], cwd=repo_dir, capture_output=True, text=True, timeout=15)
        
        # 2. Run git pull origin main
        pull_res = subprocess.run(["git", "pull", "origin", "main"], cwd=repo_dir, capture_output=True, text=True, timeout=30)
        
        # 3. Pop stashed changes back
        subprocess.run(["git", "stash", "pop"], cwd=repo_dir, capture_output=True, text=True, timeout=15)
        
        if pull_res.returncode != 0:
            return jsonify({
                "status": "error",
                "message": f"Git pull failed: {pull_res.stderr.strip()}",
                "log": pull_res.stdout.strip() + "\n" + pull_res.stderr.strip()
            }), 500
            
        return jsonify({
            "status": "success",
            "log": pull_res.stdout.strip()
        })
    except Exception as e:
        return jsonify({
            "status": "error",
            "message": f"Unexpected error pulling updates: {str(e)}"
        }), 500




@app.route('/api/users/config', methods=['GET'])
def get_user_configuration():
    # Only load the active user's config — no need to fetch all users (avoids N Google Sheets calls)
    from core.storage.engine import get_user_config
    username = get_selected_user_name()
    user_data = get_user_config(username, bypass_cache=False)  # use cache; 15s TTL is fine here
    global_settings = user_data.get("global_settings", {})
    return jsonify({
        "username": username,
        "config": user_data,
        "global_settings": global_settings
    })


@app.route('/api/users/config', methods=['POST'])
def save_user_configuration():
    body = request.get_json() or {}
    from core.storage.engine import get_user_config, save_user_config, _invalidate_cached_config
    username = get_selected_user_name()

    # Load only this user's existing config (uses cache — fast)
    existing_user = get_user_config(username, bypass_cache=False)
    if not existing_user:
        return jsonify({"status": "error", "message": f"User {username} not found"}), 404

    # Helper: deep merge incoming section over existing section so no fields are lost
    def merge_section(existing, incoming):
        if not incoming:
            return existing
        merged = dict(existing)
        merged.update({k: v for k, v in incoming.items() if v is not None and v != ''})
        return merged

    # For profile and global_settings, always use the full incoming value (fully managed by UI)
    user_profile = body.get("profile") or {}
    global_settings_incoming = body.get("global_settings") or {}

    # For pipeline sections, merge incoming over existing so stale/missing fields are preserved
    email_scraper_incoming = body.get("email_scraper") or {}
    linkedin_connect_incoming = body.get("linkedin_connect") or {}
    recruiter_outreach_incoming = body.get("recruiter_outreach") or {}
    referral_outreach_incoming = body.get("referral_outreach") or {}

    existing_scraper = existing_user.get("email_scraper") or {}
    existing_connect = existing_user.get("linkedin_connect") or {}
    existing_recruiter = existing_user.get("recruiter_outreach") or {}
    existing_referral = existing_user.get("referral_outreach") or {}

    # Apply scraper defaults for missing fields
    scraper_defaults = {
        "interval": "60", "review_mode": True, "max_emails_per_run": "5",
        "search_keywords": [], "title_keywords": [], "keywords": [],
        "excluded_keywords": [], "email_subject": "", "email_template": "", "sender_email": ""
    }
    existing_scraper = {**scraper_defaults, **existing_scraper}

    # Apply connect defaults for missing fields
    connect_defaults = {
        "interval": "60", "search_pages": 2, "review_mode": True,
        "max_connections_per_company": "5", "max_connections_per_run": "5",
        "search_keywords": [], "title_keywords": [], "keywords": [],
        "excluded_keywords": [], "message_template": ""
    }
    existing_connect = {**connect_defaults, **existing_connect}

    # Apply recruiter defaults for missing fields
    recruiter_defaults = {
        "interval": "120", "target_count": "2", "review_mode": True,
        "message_template": "", "direct_message_template": ""
    }
    existing_recruiter = {**recruiter_defaults, **existing_recruiter}

    # Merge: incoming takes precedence, existing fills gaps
    email_scraper = {**existing_scraper, **{k: v for k, v in email_scraper_incoming.items() if v is not None}}
    linkedin_connect = {**existing_connect, **{k: v for k, v in linkedin_connect_incoming.items() if v is not None}}
    recruiter_outreach = {**existing_recruiter, **{k: v for k, v in recruiter_outreach_incoming.items() if v is not None}}
    referral_outreach = {**existing_referral, **{k: v for k, v in referral_outreach_incoming.items() if v is not None}}

    # Preserve existing global_settings (credentials, etc.) and overlay incoming changes
    existing_global = existing_user.get("global_settings", {})
    merged_global = {**existing_global, **{k: v for k, v in global_settings_incoming.items() if v is not None and v != ''}}

    updated_config = dict(existing_user)
    updated_config["profile"] = user_profile
    updated_config["email_scraper"] = email_scraper
    updated_config["linkedin_connect"] = linkedin_connect
    updated_config["recruiter_outreach"] = recruiter_outreach
    updated_config["referral_outreach"] = referral_outreach
    updated_config["global_settings"] = merged_global

    # 1. Save locally immediately (fast — just writes config.json)
    from core.storage.engine import LocalStorageProvider
    LocalStorageProvider().save_config(username, updated_config)
    _invalidate_cached_config(username)
    from core.storage.engine import _set_cached_config
    _set_cached_config(username, updated_config)

    # 2. Sync to Google Sheets in background so the response returns instantly
    def _bg_sheets_sync(uname, cfg):
        try:
            db_type = cfg.get("global_settings", {}).get("database_type", "local")
            if db_type == "google_sheets":
                from core.storage.engine import GoogleSheetsStorageProvider
                GoogleSheetsStorageProvider().save_config(uname, cfg)
        except Exception as e:
            print(f"[Config Save] Google Sheets sync failed (background): {e}")

    t = threading.Thread(target=_bg_sheets_sync, args=(username, updated_config), daemon=True)
    t.start()

    return jsonify({"status": "success"})


@app.route('/api/users/config/keywords', methods=['POST'])
def save_user_keywords():
    body = request.get_json() or {}
    kws_type = body.get("type")
    keywords = body.get("keywords", [])

    valid_types = ("scraper", "connect", "scraper-excluded", "connect-excluded")
    if kws_type not in valid_types:
        return jsonify({"status": "error", "message": "Invalid keywords type"}), 400

    config = load_all_configs()
    username = get_selected_user_name()
    if username not in config.get("users", {}):
        return jsonify({"status": "error", "message": f"User {username} not found"}), 404

    if kws_type in ("scraper", "scraper-excluded"):
        target_section = "email_scraper"
    else:
        target_section = "linkedin_connect"

    field = "keywords" if kws_type in ("scraper", "connect") else "excluded_keywords"
    config["users"][username][target_section][field] = [k.strip() for k in keywords if str(k).strip()]

    # Save only the active user directly (fast local write + background Sheets sync)
    from core.storage.engine import get_user_config, LocalStorageProvider, _invalidate_cached_config, _set_cached_config
    updated_config = get_user_config(username, bypass_cache=False)
    updated_config.setdefault(target_section, {})[field] = [k.strip() for k in keywords if str(k).strip()]

    LocalStorageProvider().save_config(username, updated_config)
    _invalidate_cached_config(username)
    _set_cached_config(username, updated_config)

    def _bg_kw_sync(uname, cfg):
        try:
            db_type = cfg.get("global_settings", {}).get("database_type", "local")
            if db_type == "google_sheets":
                from core.storage.engine import GoogleSheetsStorageProvider
                GoogleSheetsStorageProvider().save_config(uname, cfg)
        except Exception as e:
            print(f"[Keywords Save] Google Sheets sync failed (background): {e}")

    threading.Thread(target=_bg_kw_sync, args=(username, updated_config), daemon=True).start()
    return jsonify({"status": "success"})

@app.route('/api/users/resume/upload', methods=['POST'])
def upload_user_resume():
    if 'resume' not in request.files:
        return jsonify({"status": "error", "message": "No file uploaded"}), 400
    file = request.files['resume']
    if file.filename == '':
        return jsonify({"status": "error", "message": "No file selected"}), 400
        
    if file:
        config = load_all_configs()
        username = get_selected_user_name()
        resumes_dir = get_resumes_dir()
        os.makedirs(resumes_dir, exist_ok=True)
        
        # Resume Replacement Logic: remove previous active resume if it exists
        if username in config.get("users", {}):
            old_resume_name = config["users"][username].get("profile", {}).get("resume_name", "")
            if old_resume_name:
                old_resume_path = os.path.join(resumes_dir, old_resume_name)
                if os.path.exists(old_resume_path):
                    try:
                        os.remove(old_resume_path)
                    except Exception as e:
                        app.logger.warning(f"Failed to delete old resume file {old_resume_path}: {e}")

        filename = secure_filename(file.filename)
        new_filename = filename
        save_path = os.path.join(resumes_dir, new_filename)
        file.save(save_path)
        
        if username in config.get("users", {}):
            config["users"][username]["profile"]["resume_name"] = new_filename
            save_all_configs(config)
            
        return jsonify({"status": "success", "resume_name": new_filename})


@app.route('/api/users/resume/download/<username>', methods=['GET'])
def download_user_resume(username):
    config = load_all_configs()
    user_data = config.get("users", {}).get(username, {})
    resume_name = user_data.get("profile", {}).get("resume_name", "")
    
    resumes_dir = get_resumes_dir()
    user_resume_path = os.path.join(resumes_dir, resume_name) if resume_name else ""

    if resume_name and os.path.exists(user_resume_path):
        return send_from_directory(resumes_dir, resume_name, as_attachment=False)

    # Fall back to the default resume path
    default_path = get_resume_file_path(user_data.get("profile", {}))
    if os.path.exists(default_path):
        return send_file(default_path, as_attachment=False)

    return "Resume not found", 404


@app.route('/api/config', methods=['GET'])
def get_config():
    user_conf = get_selected_user_config()
    global_conf = get_global_settings()
    profile = user_conf.get("profile", {})
    scraper = user_conf.get("email_scraper", {})
    
    response_data = {
        "LINKEDIN_EMAIL": global_conf.get("linkedin_email", ""),
        "LINKEDIN_PASSWORD": global_conf.get("linkedin_password", ""),
        "SEARCH_KEYWORDS": "|".join(scraper.get("search_keywords") or scraper.get("keywords") or []),
        "SEARCH_LOCATION": global_conf.get("search_location", ""),
        "SEARCH_TIME_RANGE": global_conf.get("search_time_range", "r604800"),
        "DRY_RUN": global_conf.get("dry_run", "0"),
        "MAX_RUN_DURATION_SECONDS": global_conf.get("max_run_duration_seconds", "600"),
        "RESUME_LINK": profile.get("resume_url", ""),
        "OUTREACH_MODE": "both",
        "MAX_APPLY": global_conf.get("max_apply", "5"),
        "SMTP_EMAIL": global_conf.get("smtp_email", ""),
        "SMTP_PASSWORD": global_conf.get("smtp_password", ""),
        "SISTER_NAME": (profile.get("first_name", "") + " " + profile.get("last_name", "")).strip() or profile.get("full_name", ""),
        "SISTER_EMAIL": profile.get("email", ""),
        "PHONE_NUMBER": profile.get("phone", "")
    }
    return jsonify(response_data)


@app.route('/api/config', methods=['POST'])
def save_config():
    body = request.get_json() or {}
    config = load_all_configs()
    username = get_selected_user_name()
    if username in config.get("users", {}):
        if "SISTER_NAME" in body:
            full_name = (body["SISTER_NAME"] or "").strip()
            parts = full_name.split(maxsplit=1)
            first_name = parts[0] if parts else ""
            last_name = parts[1] if len(parts) > 1 else ""
            config["users"][username]["profile"]["first_name"] = first_name
            config["users"][username]["profile"]["last_name"] = last_name
        if "SISTER_EMAIL" in body:
            config["users"][username]["profile"]["email"] = body["SISTER_EMAIL"]
        if "PHONE_NUMBER" in body:
            config["users"][username]["profile"]["phone"] = body["PHONE_NUMBER"]
        if "RESUME_LINK" in body:
            config["users"][username]["profile"]["resume_url"] = body["RESUME_LINK"]
        if "SEARCH_KEYWORDS" in body:
            keywords_list = [k.strip() for k in body["SEARCH_KEYWORDS"].split("|") if k.strip()]
            config["users"][username]["email_scraper"]["search_keywords"] = keywords_list
            config["users"][username]["email_scraper"]["title_keywords"] = keywords_list
            config["users"][username]["email_scraper"]["keywords"] = keywords_list
        
        # global settings
        if "LINKEDIN_EMAIL" in body:
            config["global_settings"]["linkedin_email"] = body["LINKEDIN_EMAIL"]
        if "LINKEDIN_PASSWORD" in body:
            config["global_settings"]["linkedin_password"] = body["LINKEDIN_PASSWORD"]
        if "SEARCH_LOCATION" in body:
            config["global_settings"]["search_location"] = body["SEARCH_LOCATION"]
        if "SEARCH_TIME_RANGE" in body:
            config["global_settings"]["search_time_range"] = body["SEARCH_TIME_RANGE"]
        if "DRY_RUN" in body:
            config["global_settings"]["dry_run"] = body["DRY_RUN"]
        if "MAX_APPLY" in body:
            config["global_settings"]["max_apply"] = body["MAX_APPLY"]
        if "SMTP_EMAIL" in body:
            config["global_settings"]["smtp_email"] = body["SMTP_EMAIL"]
        if "SMTP_PASSWORD" in body:
            config["global_settings"]["smtp_password"] = body["SMTP_PASSWORD"]
            
        save_all_configs(config)
    return jsonify({"status": "success"})


@app.route('/api/data/update_status', methods=['POST'])
def update_row_status():
    body = request.get_json() or {}
    db_type = body.get("db_type")
    row_id = body.get("id")
    status = body.get("status")
    
    if not db_type or row_id is None or not status:
        return jsonify({"status": "error", "message": "Missing required fields"}), 400
        
    try:
        row_id = int(row_id)
    except ValueError:
        pass
        
    if db_type == "scraper":
        from core.storage.engine import read_database_rows, write_database_rows
        rows = read_database_rows("emails")
        updated = False
        for r in rows:
            if str(r.get('ID')).rstrip(".0") == str(row_id).rstrip(".0"):
                r['Status'] = status
                r['Timestamp'] = datetime.utcnow().isoformat()
                updated = True
                break
        if updated:
            write_database_rows("emails", rows)
            return jsonify({"status": "success"})
        return jsonify({"status": "error", "message": "ID not found"}), 404
        
    elif db_type == "referral":
        success = update_status_by_id(row_id, status)
        if success:
            return jsonify({"status": "success"})
        return jsonify({"status": "error", "message": "JobID not found"}), 404
        
    else:
        return jsonify({"status": "error", "message": "Invalid db_type"}), 400


@app.route('/api/data/delete_row', methods=['POST'])
def delete_table_row():
    body = request.get_json() or {}
    db_type = body.get("db_type")
    row_id = body.get("id")
    
    if not db_type or row_id is None:
        return jsonify({"status": "error", "message": "Missing required fields"}), 400
        
    table_key = "emails" if db_type == "scraper" else "jobs"
    id_field = "ID" if db_type == "scraper" else "JobID"
    
    from core.storage.engine import read_database_rows, write_database_rows
    try:
        rows = read_database_rows(table_key)
        filtered_rows = [r for r in rows if str(r.get(id_field)).rstrip(".0") != str(row_id).rstrip(".0")]
        if len(filtered_rows) < len(rows):
            write_database_rows(table_key, filtered_rows)
            return jsonify({"status": "success"})
        return jsonify({"status": "error", "message": f"{id_field} not found"}), 404
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route('/api/data/edit_row', methods=['POST'])
def edit_table_row():
    body = request.get_json() or {}
    db_type = body.get("db_type")
    row_id = body.get("id")
    
    if not db_type or row_id is None:
        return jsonify({"status": "error", "message": "Missing required fields"}), 400
        
    try:
        row_id = int(row_id)
    except ValueError:
        return jsonify({"status": "error", "message": "Invalid ID"}), 400
        
    if db_type == "scraper":
        email = body.get("email")
        status = body.get("status")
        keyword = body.get("keyword")
        post_url = body.get("post_url")
        company_name = body.get("company_name")
        experience = body.get("experience")
        location = body.get("location")
        if not email or not status:
            return jsonify({"status": "error", "message": "Missing required fields"}), 400
        success = edit_row(row_id, email, status, keyword,
                           post_url=post_url, company_name=company_name,
                           experience=experience, location=location,
                           path=get_job_tracker_file())
        if success:
            return jsonify({"status": "success"})
        return jsonify({"status": "error", "message": "ID not found"}), 404
    elif db_type == "referral":
        company = body.get("company")
        url = body.get("url")
        shorten = body.get("shorten")
        keyword = body.get("keyword")
        position = body.get("position")
        status = body.get("status")
        if not company or not url or not status:
            return jsonify({"status": "error", "message": "Missing company, url, or status"}), 400
        success = edit_lead_row(row_id, company, url, shorten, keyword, position, status, get_job_leads_file())
        if success:
            return jsonify({"status": "success"})
        return jsonify({"status": "error", "message": "JobID not found"}), 404
    else:
        return jsonify({"status": "error", "message": "Invalid db_type"}), 400


@app.route('/api/data/referrals')
def referrals_data():
    from core.storage.database import get_sheets_config, init_referrals_store, load_all_referrals
    # Only initialise local file when not using Google Sheets
    if not get_sheets_config():
        init_referrals_store()
    # Do NOT pass path= here — load_all_referrals() detects Sheets mode automatically
    rows = load_all_referrals()
    rows.sort(key=lambda x: str(x.get('Sent_Time') or ''), reverse=True)
    return jsonify(rows)


@app.route('/api/data/update_referral_status', methods=['POST'])
def update_referral_status():
    body = request.get_json() or {}
    referral_id = body.get("id")
    status = body.get("status") 
    
    if referral_id is None or not status:
        return jsonify({"status": "error", "message": "Missing required fields"}), 400
        
    try:
        referral_id = int(referral_id)
    except ValueError:
        pass
        
    from core.storage.database import edit_referral_contact_row, sync_job_lead_referral_statuses
    success = edit_referral_contact_row(referral_id, {"Referral_Status": status}, get_referrals_file())
    if success:
        sync_job_lead_referral_statuses()
        return jsonify({"status": "success"})
    return jsonify({"status": "error", "message": "ReferralID not found"}), 404


@app.route('/api/data/edit_referral_row', methods=['POST'])
def edit_referral_row():
    body = request.get_json() or {}
    referral_id = body.get("id")
    if referral_id is None:
        return jsonify({"status": "error", "message": "Missing ReferralID"}), 400
        
    try:
        referral_id = int(referral_id)
    except ValueError:
        return jsonify({"status": "error", "message": "Invalid ReferralID"}), 400
        
    name = body.get("name")
    email = body.get("email")
    profile_url = body.get("profile_url")
    source = body.get("source")
    status = body.get("status")
    company = body.get("company")
    job_url = body.get("job_url") or body.get("company_url")
    notes = body.get("notes")
    verification = body.get("employment_verification_status") or "Verified"
    
    if not name or not profile_url or not status or not company:
        return jsonify({"status": "error", "message": "Missing required fields"}), 400
        
    update_data = {
        "CompanyName": company,
        "Job_URL": job_url or "",
        "Referral_Person_Name": name,
        "Referral_Person_Email": email or "",
        "Referral_Person_Profile_URL": profile_url,
        "Referral_Source": source or "Existing Connection",
        "Referral_Status": status,
        "Employment_Verification_Status": verification,
        "Error_Reason": notes or ""
    }
    
    from core.storage.database import edit_referral_contact_row, sync_job_lead_referral_statuses
    success = edit_referral_contact_row(referral_id, update_data, get_referrals_file())
    if success:
        sync_job_lead_referral_statuses()
        return jsonify({"status": "success"})
    return jsonify({"status": "error", "message": "ReferralID not found"}), 404


@app.route('/api/data/delete_referral_row', methods=['POST'])
def delete_referral_row():
    body = request.get_json() or {}
    referral_id = body.get("id")
    if referral_id is None:
        return jsonify({"status": "error", "message": "Missing ReferralID"}), 400
        
    try:
        referral_id = int(referral_id)
    except ValueError:
        pass
        
    from core.storage.database import get_sheets_config
    sheets_conf = get_sheets_config()
    if sheets_conf:
        url, creds = sheets_conf
        from core.storage.sheets import read_rows, write_rows
        try:
            rows = read_rows(url, creds, "Referrals & Connections")
        except Exception as e:
            return jsonify({"status": "error", "message": str(e)}), 400
        filtered_rows = [r for r in rows if str(r.get("ReferralID")).strip() != str(referral_id).strip()]
        if len(filtered_rows) < len(rows):
            write_rows(url, creds, "Referrals & Connections", filtered_rows)
            return jsonify({"status": "success"})
        return jsonify({"status": "error", "message": "ReferralID not found in Google Sheets"}), 404

    path = get_referrals_file()
    if not os.path.exists(path):
        return jsonify({"status": "error", "message": "File not found"}), 404
        
    try:
        wb = openpyxl.load_workbook(path)
        if "Referrals" not in wb.sheetnames:
            return jsonify({"status": "error", "message": "Referrals sheet not found"}), 404
        ws = wb["Referrals"]
        col_indices = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}
        id_col = col_indices.get('ReferralID')
        if not id_col:
            return jsonify({"status": "error", "message": "ReferralID column not found"}), 500
            
        deleted = False
        for row in range(2, ws.max_row + 1):
            val = ws.cell(row=row, column=id_col).value
            if val is not None and str(val).strip() == str(referral_id).strip():
                ws.delete_rows(row)
                deleted = True
                break
                
        if deleted:
            ws._tables.clear()
            from config.constants import REFERRAL_HEADERS
            ref_range = f"A1:{chr(64 + len(REFERRAL_HEADERS))}{ws.max_row}"
            from openpyxl.worksheet.table import Table, TableStyleInfo
            tab = Table(displayName="ReferralsTable", ref=ref_range)
            style = TableStyleInfo(
                name="TableStyleLight9",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            tab.tableStyleInfo = style
            ws.add_table(tab)
            wb.save(path)
            _trigger_mac_excel_reload(path)
            from core.storage.database import sync_job_lead_referral_statuses
            sync_job_lead_referral_statuses()
            return jsonify({"status": "success"})
            
        return jsonify({"status": "error", "message": "ReferralID not found"}), 404
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500



@app.route('/api/dev-reload')
def dev_reload():
    def event_stream():
        while True:
            time.sleep(10)
            yield "data: ping\n\n"
    return Response(event_stream(), mimetype="text/event-stream")


def sanitize_excel_files():
    # Loop over all users and sanitize their LinkedIn_Job_Tracker.xlsx file to match the latest schema
    config_dir = "users"
    if os.path.exists(config_dir):
        for user in os.listdir(config_dir):
            user_path = os.path.join(config_dir, user)
            if os.path.isdir(user_path):
                excel_path = os.path.join(user_path, "data", "LinkedIn_Job_Tracker.xlsx")
                if os.path.exists(excel_path):
                    try:
                        from core.storage.database import trim_job_leads_excel_to_schema
                        trim_job_leads_excel_to_schema(excel_path)
                        print(f"Sanitized: aligned {excel_path} to latest schema")
                    except Exception as e:
                        print(f"Error sanitizing {excel_path}: {e}")


def start_git_autoupdater(app):
    # Only run once at startup (in the Werkzeug reloader child process or if debug is off)
    if os.environ.get("WERZEUG_RUN_MAIN") == "true" or not app.debug:
        repo_dir = os.path.dirname(os.path.abspath(__file__))
        print("[Auto-Updater] Checking for git updates at startup...")
        try:
            # Determine the current branch name dynamically
            branch_res = subprocess.run(["git", "rev-parse", "--abbrev-ref", "HEAD"], cwd=repo_dir, capture_output=True, text=True, timeout=10)
            current_branch = branch_res.stdout.strip()
            if not current_branch:
                current_branch = "main"

            # 1. Fetch from remote
            fetch_res = subprocess.run(["git", "fetch", "origin"], cwd=repo_dir, capture_output=True, text=True, timeout=30)
            if fetch_res.returncode == 0:
                # 2. Check if local is behind remote branch
                remote_ref = f"origin/{current_branch}"
                behind_log = subprocess.run(["git", "log", f"HEAD..{remote_ref}", "--oneline"], cwd=repo_dir, capture_output=True, text=True, timeout=10).stdout.strip()
                if behind_log:
                    print(f"[Auto-Updater] Remote changes detected on branch '{current_branch}'. Pulling updates...")
                    # 3. Pull latest changes
                    # Stash local changes to avoid conflicts with system-generated files
                    subprocess.run(["git", "stash"], cwd=repo_dir, capture_output=True, text=True, timeout=30)
                    pull_res = subprocess.run(["git", "pull", "origin", current_branch], cwd=repo_dir, capture_output=True, text=True, timeout=60)
                    subprocess.run(["git", "stash", "pop"], cwd=repo_dir, capture_output=True, text=True, timeout=30)
                    
                    if pull_res.returncode == 0:
                        print(f"[Auto-Updater] Code successfully updated from remote branch '{current_branch}'.")
                    else:
                        print(f"[Auto-Updater] Git pull failed: {pull_res.stderr.strip()}", file=sys.stderr)
                else:
                    print(f"[Auto-Updater] Code is already up to date on branch '{current_branch}'.")
            else:
                print(f"[Auto-Updater] Git fetch failed: {fetch_res.stderr.strip()}", file=sys.stderr)
        except Exception as e:
            print(f"[Auto-Updater] Error checking updates: {e}", file=sys.stderr)



def sync_google_sheets_to_local_if_empty(username):
    """If database_type is google_sheets and local files/configs are empty or out of sync,
    synchronize and download all configuration and table data from Google Sheets.
    """
    try:
        from core.storage.engine import get_user_config, GoogleSheetsStorageProvider, LocalStorageProvider
        # Read the current local bootstrap config
        config = LocalStorageProvider().get_config(username)
        db_type = config.get("global_settings", {}).get("database_type", "local")
        if db_type != "google_sheets":
            return
            
        sheet_url = config.get("global_settings", {}).get("google_sheet_url", "").strip()
        creds_content = config.get("global_settings", {}).get("google_credentials_json", "").strip()
        if not sheet_url or not creds_content:
            return
            
        print(f"[Storage Sync] Pulling latest configuration settings from Google Sheets for user '{username}'...")
        # 1. Pull settings/profile config from Google Sheets
        try:
            cloud_config = GoogleSheetsStorageProvider().get_config(username, bypass_cache=True)
            if cloud_config:
                # Merge or write cloud_config locally
                # To prevent overwriting the local google_sheet_url/credentials bootstrap if they are missing on sheet,
                # we preserve them
                if "global_settings" not in cloud_config:
                    cloud_config["global_settings"] = {}
                cloud_config["global_settings"]["database_type"] = "google_sheets"
                cloud_config["global_settings"]["google_sheet_url"] = sheet_url
                cloud_config["global_settings"]["google_credentials_json"] = creds_content
                
                LocalStorageProvider().save_config(username, cloud_config)
                # Invalidate in-memory caches
                try:
                    from core.storage.engine import _invalidate_cached_config
                    _invalidate_cached_config(username)
                except Exception:
                    pass
                print(f"[Storage Sync]   Successfully updated local config.json with profile details and settings from Google Sheets.")
        except Exception as e:
            print(f"[Storage Sync] Warning: Could not pull configuration from Google Sheets: {e}")
            
        # 2. Pull database rows from Google Sheets
        from config.settings import get_job_tracker_file, get_job_leads_file, get_referrals_file
        orig_user = os.environ.get("CONNECTIFY_USER")
        os.environ["CONNECTIFY_USER"] = username
        try:
            local_files = [get_job_tracker_file(), get_job_leads_file(), get_referrals_file()]
        finally:
            if orig_user:
                os.environ["CONNECTIFY_USER"] = orig_user
            elif "CONNECTIFY_USER" in os.environ:
                del os.environ["CONNECTIFY_USER"]
                
        needs_sync = False
        for f in local_files:
            if not os.path.exists(f) or os.path.getsize(f) < 6000:
                needs_sync = True
                break
                
        if needs_sync:
            print(f"[Storage Sync] Local cache files for user '{username}' are empty or missing. Initializing Google Sheets pull...")
            import gspread, openpyxl as _xl
            from google.oauth2.service_account import Credentials
            from config.constants import GOOGLE_SHEET_WORKSHEETS
            
            scopes = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
            import json as _json
            creds_dict = _json.loads(creds_content)
            credentials = Credentials.from_service_account_info(creds_dict, scopes=scopes)
            client = gspread.authorize(credentials)
            sh = client.open_by_url(sheet_url)
            
            # Temporarily set active user env for paths resolution
            orig_user = os.environ.get("CONNECTIFY_USER")
            os.environ["CONNECTIFY_USER"] = username
            try:
                local_paths = {
                    "Job Leads": get_job_leads_file(),
                    "Scraped Emails": get_job_tracker_file(),
                    "Referrals & Connections": get_referrals_file()
                }
            finally:
                if orig_user:
                    os.environ["CONNECTIFY_USER"] = orig_user
                elif "CONNECTIFY_USER" in os.environ:
                    del os.environ["CONNECTIFY_USER"]
                    
            id_cols = {
                "Job Leads": "JobID",
                "Scraped Emails": "ID",
                "Referrals & Connections": "ReferralID"
            }
            
            for key, info in GOOGLE_SHEET_WORKSHEETS.items():
                if key not in ("jobs", "emails", "referrals"):
                    continue
                ws_name = info["name"]
                headers = info["headers"]
                local_file = local_paths[ws_name]
                
                try:
                    ws = sh.worksheet(ws_name)
                    cloud_rows = ws.get_all_records()
                except Exception as e:
                    print(f"[Storage Sync] Warning: Could not read cloud worksheet '{ws_name}': {e}")
                    continue
                    
                # Rewrite local file with cloud rows to fully sync
                try:
                    wb = _xl.Workbook()
                    lws = wb.active
                    lws.title = ws_name.replace(" & ", " ")
                    lws.append(headers)
                    for r in cloud_rows:
                        lws.append([str(r.get(h, "") or "") for h in headers])
                    wb.save(local_file)
                    print(f"[Storage Sync]   Synced {len(cloud_rows)} rows for '{ws_name}' locally.")
                except Exception as e:
                    print(f"[Storage Sync]   Error writing local file for '{ws_name}': {e}")
    except Exception as e:
        print(f"[Storage Sync] Warning: Automatic Google Sheets sync failed: {e}")


if __name__ == '__main__':
    # Ensure standard directories exist
    os.makedirs("templates", exist_ok=True)
    os.makedirs("static/css", exist_ok=True)
    os.makedirs("static/js", exist_ok=True)
    
    # Sanitize existing databases
    sanitize_excel_files()
    
    # Run startup database deduplication
    try:
        from config.user_profiles import load_all_configs
        from core.storage.database import deduplicate_all_tables
        all_cfg = load_all_configs()
        users = list(all_cfg.get("users", {}).keys())
        if not users:
            users = ["default"]
        for u in users:
            deduplicate_all_tables(u)
            try:
                sync_google_sheets_to_local_if_empty(u)
            except Exception as se:
                print(f"[Startup Database] Warning: Sheets auto-sync failed for user '{u}': {se}")
        print("[Startup Database] Safely scanned, auto-synced, and deduplicated all user tables.")
    except Exception as e:
        print(f"[Startup Database] Warning: Deduplication run failed: {e}")
    
    # Start the Git auto-updater thread
    start_git_autoupdater(app)
    
    # Watch templates and static files to trigger reloader
    extra_files = []
    for extra_dir in ["templates", "static"]:
        if os.path.exists(extra_dir):
            for root, dirs, files in os.walk(extra_dir):
                for file in files:
                    extra_files.append(os.path.join(root, file))
    
    print("Connectify Automation Hub starting at http://127.0.0.1:5001")
    app.run(host='127.0.0.1', port=5001, debug=True, extra_files=extra_files)

