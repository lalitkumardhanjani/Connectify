import os
import time
import openpyxl
from config.settings import JOB_LEADS_FILE
from core.integrations.url_shortener import shorten_url
from core.logging.config import logger

def load_job_data_excel(filename):
    if not os.path.exists(filename):
        logger.error(f"Excel file '{filename}' not found.")
        return [], None, None, []
    try:
        wb = openpyxl.load_workbook(filename)
        ws = wb.active
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            rows.append(dict(zip(headers, row)))
        return rows, wb, ws, headers
    except Exception as e:
        logger.exception(f"Error loading '{filename}': {e}")
        return [], None, None, []

def save_job_data_excel(wb, ws, headers, rows, filename):
    # 1. ALWAYS write locally first
    if ws is not None and wb is not None:
        try:
            ws.delete_rows(2, ws.max_row)
            for row_dict in rows:
                ws.append([row_dict.get(h, "") for h in headers])
            wb.save(filename)
            logger.info(f"Successfully saved updated job data to '{filename}'.")
        except Exception as e:
            logger.error(f"Error saving '{filename}': {e}")

    # 2. Update Google Sheets backup if configured
    from core.storage.database import get_sheets_config
    sheets_conf = get_sheets_config()
    if sheets_conf:
        url, creds = sheets_conf
        from core.storage.sheets import write_rows
        try:
            write_rows(url, creds, "Job Leads", rows)
            logger.info("Successfully saved updated job data to Google Sheets backup.")
        except Exception as e:
            logger.error(f"Error saving job leads to Google Sheets: {e}")

def run_url_shortener():
    """Runs the URL shortener workflow to process long Company URLs using TinyURL."""
    rows, wb, ws, headers = load_job_data_excel(JOB_LEADS_FILE)
    if not rows:
        logger.warning("No jobs to process for URL shortening.")
        return

    logger.info(f"Loaded {len(rows)} jobs from '{JOB_LEADS_FILE}'.")

    def col_index(name):
        for h in headers:
            if isinstance(h, str) and h.lower() == name.lower():
                return h
        return None

    company_url_col = col_index('CompanyURL')
    shorten_url_col = col_index('ShortenURL')
    status_col = col_index('Status')
    shorturl_col = col_index('ShortUrlCreated')

    updated_jobs_count = 0
    for i, row in enumerate(rows):
        original_url = row.get(company_url_col)
        
        if shorturl_col and row.get(shorturl_col) == 'Yes':
            logger.info(f"Job {i+1}: Already processed (ShortUrlCreated='Yes'), skipping.")
            continue
            
        if not original_url or "shorturl.at" in original_url or original_url.startswith("https://tinyurl.com/"):
            logger.info(f"Job {i+1}: Skipping (missing URL or already shortened): {original_url}")
            continue
            
        logger.info(f"Job {i+1}: Shortening URL: {original_url}")
        shortened = shorten_url(original_url)
        if shortened:
            if shorten_url_col:
                row[shorten_url_col] = shortened
            if status_col and not row.get(status_col):
                row[status_col] = 'NEW'
            if shorturl_col:
                row[shorturl_col] = 'Yes'
            updated_jobs_count += 1
            logger.info(f"  -> Shortened to: {shortened}")
        else:
            logger.info("  -> Failed to shorten URL. Keeping empty.")
        time.sleep(2)

    if updated_jobs_count > 0:
        save_job_data_excel(wb, ws, headers, rows, JOB_LEADS_FILE)
        logger.info(f"\nSuccessfully updated {updated_jobs_count} URLs in '{JOB_LEADS_FILE}'.")
        # Reload sheet on macOS
        from core.storage.database import _trigger_mac_excel_reload
        _trigger_mac_excel_reload(JOB_LEADS_FILE)
    else:
        logger.info("\nNo new URLs were shortened or updated.")

if __name__ == "__main__":
    run_url_shortener()
