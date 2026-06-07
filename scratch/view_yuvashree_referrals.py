import sys
import os
import openpyxl

sys.path.append(os.getcwd())
from config.settings import get_job_leads_file

path = get_job_leads_file()
print(f"Path: {path}")

wb = openpyxl.load_workbook(path)
if "Referrals" not in wb.sheetnames:
    print("Referrals sheet not found!")
else:
    ws = wb["Referrals"]
    headers = [cell.value for cell in ws[1]]
    print(f"Headers: {headers}")
    
    col_indices = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}
    name_idx = col_indices.get("Referral_Person_Name")
    status_idx = col_indices.get("Referral_Status")
    source_idx = col_indices.get("Referral_Source")
    
    print("\nAll Records:")
    for r in range(2, ws.max_row + 1):
        row_dict = {headers[c-1]: ws.cell(row=r, column=c).value for c in range(1, len(headers) + 1)}
        print(f"Row {r}: {row_dict}")

