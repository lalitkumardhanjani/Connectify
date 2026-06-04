import os
import openpyxl
from selenium.webdriver.chrome.options import Options
from selenium import webdriver

EXCEL_FILE = "LinkedIn_Job_Tracker.xlsx"

def load_job_data_excel(filename):
    """Load job data from the Excel tracker and return rows, workbook, sheet, headers."""
    if not os.path.exists(filename):
        print(f"Excel file '{filename}' not found!")
        return [], None, None, []
    try:
        wb = openpyxl.load_workbook(filename)
        ws = wb.active
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            rows.append(dict(zip(headers, row)))
        return rows, wb, ws, headers
    except Exception as e:
        print(f"Error loading Excel file: {e}")
        return [], None, None, []

def save_job_data_excel(wb, ws, headers, rows):
    """Write rows back to the Excel worksheet and save the workbook."""
    ws.delete_rows(2, ws.max_row)
    for row_dict in rows:
        ws.append([row_dict.get(h, "") for h in headers])
    wb.save(EXCEL_FILE)
    print(f"Excel file '{EXCEL_FILE}' saved.")

def main():
    rows, wb, ws, headers = load_job_data_excel(EXCEL_FILE)
    if not rows:
        print("No jobs found in Excel!")
        return

    # Resolve column names (case‐insensitive)
    def col(name):
        for h in headers:
            if isinstance(h, str) and h.lower() == name.lower():
                return h
        return None

    url_col = col('CompanyURL')
    company_col = col('CompanyName')
    status_col = col('Status')
    total = len(rows)
    print(f"Loaded {total} jobs from Excel.")

    # Only process rows whose Status column is exactly "NEW"
    # All other rows are left untouched (preserved in processed_rows).
    processed_rows = []
    jobs_to_process = []
    for row in rows:
        status = row.get(status_col)
        if status == "NEW":
            jobs_to_process.append(row)
        else:
            processed_rows.append(row)
    total = len(jobs_to_process)
    if total == 0:
        print("No jobs left to review.")
        return

    # Chrome setup matching other scripts (using linkedin_job_driver.py)
    from linkedin_job_driver import get_driver
    try:
        driver = get_driver()
    except Exception as e:
        print(f"Error starting Chrome: {e}")
        return


    updated_rows = []
    try:
        for index, row in enumerate(jobs_to_process, start=1):
            url = row.get(url_col, "")
            company = row.get(company_col, "Unknown")

            if not url:
                print(f"No URL found for {company}. Skipping...")
                updated_rows.append(row)
                continue

            print("\n" + "=" * 60)
            print(f"Job {index}/{total}")
            print(f"Company : {company}")
            print("=" * 60)
            driver.get(url)

            while True:
                print("\nOptions:")
                print("1 -> Ask for referral (keep row)")
                print("2 -> Not Interested (set status)")
                print("q -> Quit")
                choice = input("\nEnter choice: ").strip().lower()
                if choice in ['1', '2', 'q']:
                    break
                print("Invalid choice. Please enter 1, 2, or q.")

            if choice == 'q':
                print("Stopping script...")
                # Append remaining unprocessed rows unchanged
                remaining = jobs_to_process[index-1:]
                updated_rows.extend(remaining)
                break
            elif choice == '1':
                print("Marked for referral. Keeping row.")
                if status_col:
                    row[status_col] = "Ask for referral"
                updated_rows.append(row)
            elif choice == '2':
                print("Not interested. Updating status.")
                if status_col:
                    row[status_col] = "Not Interested"
                updated_rows.append(row)

            # After handling the choice, save current state (processed + updated + remaining)
            remaining = jobs_to_process[index:]
            all_rows = processed_rows + updated_rows + remaining
            save_job_data_excel(wb, ws, headers, all_rows)
            print("Saved progress to Excel.")
    except Exception as e:
        print(f"\nFatal error: {e}")
    finally:
        print("\nClosing browser...")
        driver.quit()
        print("Script finished.")

if __name__ == "__main__":
    main()
