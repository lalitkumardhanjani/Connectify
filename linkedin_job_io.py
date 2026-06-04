import os
import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import datetime
from linkedin_job_config import TRACKER_FILE
from linkedin_job_helpers import is_valid_external_url, normalize_external_url

seen_external_urls = set()

HEADERS = [
    'JobID', 'CompanyName', 'CompanyURL', 'ShortenURL', 'SearchKeyword', 'Status', 
    'ReferralAsked', 'ShortUrlCreated', 'ReferralPerson', 'Remarks', 'CreatedDateTime'
]

def load_saved_jobs():
    jobs = []
    normalized_urls = set()

    if os.path.exists(TRACKER_FILE):
        try:
            wb = openpyxl.load_workbook(TRACKER_FILE)
            ws = wb.active
            # Find the header indices
            col_indices = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}
            
            company_url_col = col_indices.get('CompanyURL')
            company_col = col_indices.get('CompanyName')
            job_id_col = col_indices.get('JobID')
            keyword_col = col_indices.get('SearchKeyword')

            if company_url_col:
                for row in ws.iter_rows(min_row=2, values_only=True):
                    # Skip empty rows
                    if not any(cell is not None for cell in row):
                        continue
                    company_url = row[company_url_col - 1]
                    if company_url:
                        company_url_str = str(company_url).strip()
                        normalized_urls.add(normalize_external_url(company_url_str))
                        
                        jobs.append({
                            "url": company_url_str,
                            "company": str(row[company_col - 1]).strip() if company_col and row[company_col - 1] else "",
                            "job_id": int(row[job_id_col - 1]) if job_id_col and row[job_id_col - 1] else None,
                            "keyword": str(row[keyword_col - 1]).strip() if keyword_col and row[keyword_col - 1] else ""
                        })
        except Exception as e:
            print(f"Error loading tracker Excel file: {str(e)}")

    return jobs, normalized_urls

def update_excel_table(ws):
    # Clear any existing tables
    ws._tables.clear()
    
    # Determine the reference range
    max_row = max(ws.max_row, 1)
    ref_range = f"A1:K{max_row}"
    
    # Create Table
    tab = Table(displayName="JobTrackerTable", ref=ref_range)
    
    # Apply styling
    style = TableStyleInfo(
        name="TableStyleLight9", 
        showFirstColumn=False,
        showLastColumn=False, 
        showRowStripes=True, 
        showColumnStripes=False
    )
    tab.tableStyleInfo = style
    ws.add_table(tab)

def save_job(data):
    url = (data.get("url") or "").strip()
    company = (data.get("company") or "").strip()
    search_keyword = (data.get("search_keyword") or "").strip()

    if not url:
        return False

    if not is_valid_external_url(url):
        return False

    normalized_url = normalize_external_url(url)
    jobs, existing_urls = load_saved_jobs()

    # Exact raw URL duplicate check AND normalized URL duplicate check
    if normalized_url in existing_urls:
        print(f"Duplicate skipped (normalized): {url}")
        return False

    for existing_job in jobs:
        if existing_job.get("url") == url:
            print(f"Duplicate skipped (exact raw URL): {url}")
            return False

    try:
        if not os.path.exists(TRACKER_FILE):
            clean_json_outputs()

        wb = openpyxl.load_workbook(TRACKER_FILE)
        ws = wb.active

        # Auto-increment JobID
        max_id = 0
        col_indices = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}
        job_id_col = col_indices.get('JobID', 1)

        for row in range(2, ws.max_row + 1):
            val = ws.cell(row=row, column=job_id_col).value
            if val is not None:
                try:
                    val_int = int(val)
                    if val_int > max_id:
                        max_id = val_int
                except ValueError:
                    pass

        new_job_id = max_id + 1
        created_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # Map to headers structure
        row_data = []
        for header in HEADERS:
            if header == 'JobID':
                row_data.append(new_job_id)
            elif header == 'CompanyName':
                row_data.append(company)
            elif header == 'CompanyURL':
                row_data.append(url)
            elif header == 'ShortenURL':
                row_data.append('')
            elif header == 'SearchKeyword':
                row_data.append(search_keyword)
            elif header == 'Status':
                row_data.append('NEW')
            elif header == 'ReferralAsked':
                row_data.append('No')
            elif header == 'ShortUrlCreated':
                row_data.append('No')
            elif header == 'ReferralPerson':
                row_data.append('')
            elif header == 'Remarks':
                row_data.append('')
            elif header == 'CreatedDateTime':
                row_data.append(created_time)

        # Append new row
        ws.append(row_data)

        # Sort all rows by JobID in ascending order
        rows_to_sort = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(cell is not None for cell in row):
                rows_to_sort.append(list(row))

        job_id_idx = HEADERS.index('JobID')
        rows_to_sort.sort(key=lambda r: int(r[job_id_idx]) if r[job_id_idx] is not None else 0)

        # Clear existing sheet rows from row 2 onwards
        while ws.max_row > 1:
            ws.delete_rows(2)

        # Re-append sorted rows
        for r in rows_to_sort:
            ws.append(r)

        # Refresh formal Excel Table
        update_excel_table(ws)

        wb.save(TRACKER_FILE)

        # On macOS, if the user has Excel/Numbers open, reload the file so they see updates instantly
        try:
            import subprocess
            abs_path = os.path.abspath(TRACKER_FILE)
            
            # AppleScript to reopen the file in Microsoft Excel or Numbers if they are running and have the file open
            reload_script = f'''
            tell application "System Events"
                set excelRunning to (name of processes) contains "Microsoft Excel"
                set numbersRunning to (name of processes) contains "Numbers"
            end tell
            
            if excelRunning then
                tell application "Microsoft Excel"
                    try
                        if exists workbook "{TRACKER_FILE}" then
                            close workbook "{TRACKER_FILE}" saving no
                            open POSIX file "{abs_path}"
                        end if
                    end try
                end tell
            end if
            
            if numbersRunning then
                tell application "Numbers"
                    try
                        set docName to "{TRACKER_FILE}"
                        set docExists to false
                        repeat with doc in documents
                            if name of doc is docName then
                                set docExists to true
                                exit repeat
                            end if
                        end repeat
                        if docExists then
                            close document docName saving no
                            open POSIX file "{abs_path}"
                        end if
                    end try
                end tell
            end if
            '''
            subprocess.run(["osascript", "-e", reload_script], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        except Exception:
            pass

        seen_external_urls.add(normalized_url)
        print(f"Saved & Sorted in Excel Tracker (JobID {new_job_id}): {url}")
        return True

    except Exception as e:
        print(f"Error saving job to Excel tracker: {str(e)}")
        return False

def clean_json_outputs():
    # Only initialize if the file doesn't exist. Never wipe out data on startup.
    if not os.path.exists(TRACKER_FILE):
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Jobs"
            ws.append(HEADERS)
            update_excel_table(ws)
            wb.save(TRACKER_FILE)
            print(f"Initialized new Excel tracker: {TRACKER_FILE}")
        except Exception as e:
            print(f"Unable to initialize Excel tracker: {str(e)}")
    else:
        # Just ensure headers and Table are present
        try:
            wb = openpyxl.load_workbook(TRACKER_FILE)
            ws = wb.active
            if ws.max_row == 0 or ws.cell(row=1, column=1).value != 'JobID':
                ws.insert_rows(1)
                for col_idx, header in enumerate(HEADERS, start=1):
                    ws.cell(row=1, column=col_idx, value=header)
            update_excel_table(ws)
            wb.save(TRACKER_FILE)
        except Exception:
            pass
