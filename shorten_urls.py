import json
import requests
import os
from datetime import datetime
import time # Ensure time is imported for sleep

EXCEL_FILE = "LinkedIn_Job_Tracker.xlsx"
# Change to TinyURL API endpoint
SHORTENER_SERVICE_URL = "https://tinyurl.com/api-create.php"
def load_job_data_excel(filename):
    """Load job data from Excel file and return list of row dictionaries."""
    import openpyxl
    if not os.path.exists(filename):
        print(f"Error: Excel file '{filename}' not found.")
        return []
    try:
        wb = openpyxl.load_workbook(filename)
        ws = wb.active
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            rows.append(dict(zip(headers, row)))
        return rows, wb, ws, headers
    except Exception as e:
        print(f"Error loading '{filename}': {e}")
        return [], None, None, []

def save_job_data_excel(wb, ws, headers, rows):
    """Write updated rows back to the Excel worksheet and save workbook."""
    try:
        # Clear existing data rows
        ws.delete_rows(2, ws.max_row)
        for row_dict in rows:
            ws.append([row_dict.get(h, "") for h in headers])
        wb.save(EXCEL_FILE)
        print(f"Successfully saved updated job data to '{EXCEL_FILE}'.")
    except Exception as e:
        print(f"Error saving '{EXCEL_FILE}': {e}")

def shorten_url(long_url):
    """Shortens a given URL using TinyURL API."""
    try:
        # TinyURL API is a simple GET request that returns the shortened URL as plain text
        response = requests.get(f"{SHORTENER_SERVICE_URL}?url={long_url}", timeout=10)
        response.raise_for_status() # Raise an exception for HTTP errors
        
        shortened = response.text.strip()

        # TinyURL returns the long URL itself or an error message if it fails to shorten
        if shortened.startswith("Error") or shortened == long_url:
            print(f"TinyURL API returned an error or original URL for '{long_url}': {shortened}")
            return None
        return shortened

    except requests.exceptions.RequestException as e:
        print(f"Error shortening URL '{long_url}' with TinyURL: {e}")
        return None
    except Exception as e:
        print(f"An unexpected error occurred while shortening URL '{long_url}': {e}")
        return None

def main():
    rows, wb, ws, headers = load_job_data_excel(EXCEL_FILE)
    if not rows:
        print("No jobs to process. Exiting.")
        return

    print(f"Loaded {len(rows)} jobs from '{EXCEL_FILE}'.")

    # Determine column names (case‑insensitive)
    def col_index(name):
        for h in headers:
            if isinstance(h, str) and h.lower() == name.lower():
                return h
        return None

    company_url_col = col_index('CompanyURL')
    shorten_url_col = col_index('ShortenURL')
    status_col = col_index('Status')
    shorturl_col = col_index('ShortUrlCreated')

    updated_jobs_count = 0
    for i, row in enumerate(rows):
        original_url = row.get(company_url_col)
        # Skip if already processed (ShortUrlCreated == 'Yes')
        if shorturl_col and row.get(shorturl_col) == 'Yes':
            print(f"Job {i+1}: Already processed (ShortUrlCreated='Yes'), skipping.")
            continue
        if not original_url or "shorturl.at" in original_url or original_url.startswith("https://tinyurl.com/"):
            print(f"Job {i+1}: Skipping (missing URL or already shortened): {original_url}")
            continue
        print(f"Job {i+1}: Shortening URL: {original_url}")
        shortened = shorten_url(original_url)
        if shortened:
            if shorten_url_col:
                row[shorten_url_col] = shortened
            # Preserve existing status (e.g. 'Ask for referral' or 'Not Interested')
            if status_col and not row.get(status_col):
                row[status_col] = 'NEW'
            if shorturl_col:
                row[shorturl_col] = 'Yes'
            updated_jobs_count += 1
            print(f"  -> Shortened to: {shortened}")
        else:
            print(f"  -> Failed to shorten URL. Keeping empty.")
        time.sleep(2)

    if updated_jobs_count > 0:
        save_job_data_excel(wb, ws, headers, rows)
        print(f"\nSuccessfully updated {updated_jobs_count} URLs in '{EXCEL_FILE}'.")
    else:
        print("\nNo new URLs were shortened or updated.")

if __name__ == "__main__":
    main()