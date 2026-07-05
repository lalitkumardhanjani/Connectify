#!/usr/bin/env python
"""
migrate_to_local.py — Reverse migration: Google Sheets → Local Excel files.

Reads all three worksheets (Job Leads, Scraped Emails, Referrals & Connections)
from the configured Google Sheet and writes them back into the user's local Excel
workbooks. Safe to run multiple times (idempotent by primary key).
"""
import os
import sys
import json
import pandas as pd

sys.path.append(os.path.dirname(os.path.abspath(__file__)))


def main():
    try:
        sys.stdout.reconfigure(encoding='utf-8')
    except Exception:
        pass

    print("=================================================================")
    print("   Connectify  Google Sheets → Local Excel  Migration Utility   ")
    print("=================================================================")

    from config.user_profiles import get_selected_user_config, get_selected_user_name
    from config.constants import GOOGLE_SHEET_WORKSHEETS
    from config.settings import get_job_tracker_file, get_job_leads_file, get_referrals_file

    username = get_selected_user_name()
    if not username:
        print("❌ No active user profile found. Launch the app first.")
        sys.exit(1)

    print(f"Active Profile : {username}")
    user_conf = get_selected_user_config()
    global_settings = user_conf.get("global_settings", {})
    sheet_url = global_settings.get("google_sheet_url")
    creds_content = global_settings.get("google_credentials_json")

    if not sheet_url or not creds_content:
        print("❌ Google Sheets is not configured for this profile.")
        sys.exit(1)

    # --- Authenticate ---
    print("\nStep 1: Authenticating with Google APIs...")
    try:
        import gspread
        from google.oauth2.service_account import Credentials
        scopes = [
            "https://spreadsheets.google.com/feeds",
            "https://www.googleapis.com/auth/drive"
        ]
        creds_dict = json.loads(creds_content)
        credentials = Credentials.from_service_account_info(creds_dict, scopes=scopes)
        client = gspread.authorize(credentials)
        print("✅ Authentication successful.")
    except Exception as e:
        print(f"❌ Authentication Failed: {e}")
        sys.exit(1)

    # --- Open Spreadsheet ---
    print("\nStep 2: Connecting to Google Sheet...")
    try:
        sh = client.open_by_url(sheet_url)
        print(f"✅ Opened: '{sh.title}'")
    except Exception as e:
        print(f"❌ Failed to open Google Sheet: {e}")
        sys.exit(1)

    local_paths = {
        "Job Leads": get_job_leads_file(),
        "Scraped Emails": get_job_tracker_file(),
        "Referrals & Connections": get_referrals_file()
    }

    id_cols = {
        "Job Leads": "JobID",
        "Scraped Emails": "ID",
        "Referrals & Connections": "ReferralID"
    }

    import time, openpyxl
    from openpyxl.worksheet.table import Table, TableStyleInfo

    print("\nStep 3: Pulling data from Google Sheets into local Excel files...")

    for key, info in GOOGLE_SHEET_WORKSHEETS.items():
        if key == "config":
            continue
        ws_name = info["name"]
        headers = info["headers"]
        local_file = local_paths[ws_name]
        id_col = id_cols[ws_name]

        print(f"\n📁 Processing '{ws_name}'...")
        time.sleep(1.5)  # throttle to avoid 429

        try:
            ws = sh.worksheet(ws_name)
            cloud_rows = ws.get_all_records()
        except Exception as e:
            print(f"  ⚠️ Could not read worksheet '{ws_name}': {e}")
            continue

        if not cloud_rows:
            print(f"  ℹ️ No data in Google Sheets for '{ws_name}'. Skipping.")
            continue

        # Load existing local IDs to deduplicate
        existing_ids = set()
        if os.path.exists(local_file):
            try:
                wb = openpyxl.load_workbook(local_file)
                lws = wb.active
                col_map = {cell.value: idx for idx, cell in enumerate(lws[1], start=1)}
                id_idx = col_map.get(id_col)
                if id_idx:
                    for row in range(2, lws.max_row + 1):
                        val = lws.cell(row=row, column=id_idx).value
                        if val is not None:
                            existing_ids.add(str(val).rstrip(".0"))
            except Exception:
                pass

        # Filter to only rows not already in local file
        new_rows = []
        for r in cloud_rows:
            row_id = str(r.get(id_col, "") or "").rstrip(".0")
            if row_id and row_id not in existing_ids:
                new_rows.append(r)

        print(f"  ☁️  Google Sheet rows : {len(cloud_rows)}")
        print(f"  💾 Local existing IDs: {len(existing_ids)}")
        print(f"  ➕ New rows to write : {len(new_rows)}")

        if not new_rows:
            print(f"  ✅ Local Excel is already up to date — no new rows needed.")
            continue

        # Create or open workbook
        if os.path.exists(local_file):
            wb = openpyxl.load_workbook(local_file)
            lws = wb.active
        else:
            wb = openpyxl.Workbook()
            lws = wb.active
            lws.title = ws_name.replace(" & ", " ")
            lws.append(headers)

        for r in new_rows:
            lws.append([str(r.get(h, "") or "") for h in headers])

        wb.save(local_file)
        print(f"  ✅ Written {len(new_rows)} rows to '{os.path.basename(local_file)}'.")

    # --- Update config to local ---
    print("\nStep 4: Downloading configuration settings from Google Sheets and switching active storage to Local Excel...")
    try:
        from core.storage.engine import GoogleSheetsStorageProvider, LocalStorageProvider
        sheets_provider = GoogleSheetsStorageProvider()
        full_config = sheets_provider.get_config(username)
        
        if "global_settings" not in full_config:
            full_config["global_settings"] = {}
        full_config["global_settings"]["database_type"] = "local"
        
        local_provider = LocalStorageProvider()
        local_provider.save_config(username, full_config)
        print("✅ Config updated. Configuration settings downloaded from Google Sheets and Active Storage is now LOCAL EXCEL.")
    except Exception as e:
        print(f"❌ Error downloading and updating configuration: {e}")

    print("\n🎉 Reverse migration completed successfully!")


if __name__ == "__main__":
    main()
