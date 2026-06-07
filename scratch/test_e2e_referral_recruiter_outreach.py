import os
import sys
import unittest
import openpyxl
from datetime import datetime
from unittest.mock import MagicMock, patch

# Ensure workspace is in PYTHONPATH
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

from config.constants import JOB_LEADS_HEADERS, REFERRAL_HEADERS
from config.user_profiles import get_selected_user_name, load_all_configs, save_all_configs

TEST_LEADS_FILE = os.path.abspath(os.path.join(os.path.dirname(__file__), "test_job_leads.xlsx"))

def get_mock_job_leads_file():
    return TEST_LEADS_FILE

class TestE2EReferralRecruiterOutreach(unittest.TestCase):
    
    def setUp(self):
        # Clean up any leftover test files
        if os.path.exists(TEST_LEADS_FILE):
            os.remove(TEST_LEADS_FILE)
            
        # Back up active user config
        self.username = get_selected_user_name()
        self.orig_config = load_all_configs()
        
    def tearDown(self):
        # Restore active user config
        if self.orig_config:
            save_all_configs(self.orig_config)
        # Clean up test files
        if os.path.exists(TEST_LEADS_FILE):
            os.remove(TEST_LEADS_FILE)

    @patch('core.storage.database.get_job_leads_file', get_mock_job_leads_file)
    @patch('config.settings.get_job_leads_file', get_mock_job_leads_file)
    @patch('config.settings.JOB_LEADS_FILE', TEST_LEADS_FILE)
    @patch('pipelines.linkedin_outreach.services.connector.get_driver')
    @patch('pipelines.linkedin_outreach.services.connector.login_to_linkedin', return_value=True)
    @patch('pipelines.linkedin_outreach.services.referral_outreach.get_driver')
    @patch('pipelines.linkedin_outreach.services.recruiter_connector.get_driver')
    def test_end_to_end_referral_and_recruiter_outreach(self, mock_recruiter_driver, mock_referral_driver, mock_login, mock_connector_driver):
        # 1. Initialize E2E excel worksheets
        from core.storage.database import init_job_leads_store, init_referrals_store
        init_job_leads_store()
        init_referrals_store()
        
        # 2. Add E2E Job Leads
        wb = openpyxl.load_workbook(TEST_LEADS_FILE)
        ws_jobs = wb["Jobs"]
        
        # Row 2: Interested job (for Referral Employee discovery)
        ws_jobs.append([101, "DBA", "Google", "https://google.com/jobs", "", "DBA", "Interested", "Yes", "2026-06-07T10:00:00"])
        # Row 3: Asked for Referral job (for Recruiter outreach)
        ws_jobs.append([102, "DBA", "Microsoft", "https://microsoft.com/jobs", "", "DBA", "Asked for Referral", "Yes", "2026-06-07T10:00:00"])
        wb.save(TEST_LEADS_FILE)
        
        # Setup E2E settings configuration
        config = load_all_configs()
        user_conf = config.get("users", {}).get(self.username, {})
        user_conf["linkedin_connect"] = {
            "interval": "15",
            "review_mode": True,
            "max_connections_per_run": "2",
            "keywords": ["DBA"],
            "message_template": "Hello {PERSON_NAME}, let's connect."
        }
        user_conf["recruiter_outreach"] = {
            "interval": "15",
            "target_count": "2",
            "review_mode": True,
            "message_template": "Hi {PERSON_NAME}, I want to connect at {company}."
        }
        user_conf["profile"] = {
            "resume_url": "https://shorturl.at/mock-resume"
        }
        save_all_configs(config)
        
        # 3. referral discovery (Pipeline 4) E2E Test
        # Mock Referral employee search & scraping
        mock_driver = MagicMock()
        mock_referral_driver.return_value = mock_driver
        
        # Mock employee URL finder and connections list
        with patch('pipelines.linkedin_outreach.services.referral_outreach.find_company_employees_search_url', return_value="https://linkedin.com/search/employees/Google"), \
             patch('pipelines.linkedin_outreach.services.referral_outreach.scrape_connections_from_search') as mock_scrape:
             
             # Setup Mock Connections
             mock_scrape.side_effect = [
                 # Google connections
                 [
                     {"name": "Alice Employee", "designation": "Staff Engineer at Google", "profile_url": "https://linkedin.com/in/alice-google"},
                     {"name": "Bob Employee", "designation": "DBA at Google", "profile_url": "https://linkedin.com/in/bob-google"}
                 ]
             ]
             
             from pipelines.linkedin_outreach.services.referral_outreach import run_phase_one_discovery
             print("\n[E2E] Running Referral Discovery (Pipeline 4)...")
             run_phase_one_discovery()
             
        # Verify Pending referral rows created
        wb = openpyxl.load_workbook(TEST_LEADS_FILE)
        ws_referrals = wb["Referrals"]
        
        referral_rows = []
        for r in range(2, ws_referrals.max_row + 1):
            referral_rows.append([ws_referrals.cell(row=r, column=c).value for c in range(1, len(REFERRAL_HEADERS) + 1)])
            
        print("[E2E] Discovered employee contacts in Referrals sheet:", referral_rows)
        self.assertEqual(len(referral_rows), 2, "Discovery should collect exactly 2 connections (Google target)")
        self.assertEqual(referral_rows[0][3], "Alice Employee")
        self.assertEqual(referral_rows[0][7], "Existing Employee")
        self.assertEqual(referral_rows[0][8], "Pending")
        
        # 4. Referral Messaging (Pipeline 5) E2E Test
        # Setup Quality Gate prompts:
        # First contact (Alice Employee) -> Skip ('k')
        # Second contact (Bob Employee) -> Send ('s')
        mock_input = MagicMock(side_effect=['k', 's'])
        
        # Mock selenium UI actions
        with patch('pipelines.linkedin_outreach.services.referral_outreach.open_messaging_from_profile', return_value=True), \
             patch('pipelines.linkedin_outreach.services.referral_outreach.insert_message_draft', return_value=True), \
             patch('pipelines.linkedin_outreach.services.referral_outreach.upload_resume_attachment', return_value=True), \
             patch('pipelines.linkedin_outreach.services.referral_outreach.click_send_message', return_value=True), \
             patch('pipelines.linkedin_outreach.services.referral_outreach.verify_delivery', return_value=True), \
             patch('builtins.input', mock_input):
             
             from pipelines.linkedin_outreach.services.referral_outreach import run_phase_two_messaging
             print("\n[E2E] Running Referral Messaging (Pipeline 5) under Quality Gate...")
             run_phase_two_messaging()
             
        # Verify status transitions E2E in spreadsheet
        wb = openpyxl.load_workbook(TEST_LEADS_FILE)
        ws_referrals = wb["Referrals"]
        
        final_referrals = {}
        for r in range(2, ws_referrals.max_row + 1):
            name = ws_referrals.cell(row=r, column=4).value
            status = ws_referrals.cell(row=r, column=9).value
            final_referrals[name] = status
            
        print("[E2E] Referral statuses after phase 2:", final_referrals)
        self.assertEqual(final_referrals["Alice Employee"], "Skipped")
        self.assertEqual(final_referrals["Bob Employee"], "Sent")
        
        # 5. Recruiter Outreach Complete Workflow E2E Test
        # Set up recruiter connector mocks
        mock_rec_driver = MagicMock()
        mock_recruiter_driver.return_value = mock_rec_driver
        
        # Mock connection request quality gate options:
        # John Recruiter -> Send ('s'), Jane Recruiter -> Send ('s')
        mock_rec_input = MagicMock(side_effect=['s', 's'])
        
        # Mock selenium action calls and connections discovery
        with patch('pipelines.linkedin_outreach.services.recruiter_connector.find_people_with_connect_button') as mock_find_rec, \
             patch('pipelines.linkedin_outreach.services.recruiter_connector.send_connection_request', return_value=True), \
             patch('builtins.input', mock_rec_input):
              
             # Return 3 mock recruiter cards to test limit capping (target is 2)
             mock_find_rec.return_value = [
                 {"name": "John Recruiter", "role": "TA Lead at Microsoft", "profile_url": "https://linkedin.com/in/john-recruiter", "button": MagicMock()},
                 {"name": "Jane Recruiter", "role": "HR Manager at Microsoft", "profile_url": "https://linkedin.com/in/jane-recruiter", "button": MagicMock()},
                 {"name": "Jack Recruiter", "role": "Recruiter at Microsoft", "profile_url": "https://linkedin.com/in/jack-recruiter", "button": MagicMock()}
             ]
              
             from pipelines.linkedin_outreach.services.recruiter_connector import run_recruiter_connector
             print("\n[E2E] Running Recruiter Outreach Pipeline workflow...")
             run_recruiter_connector()
              
        # Verify Recruiter rows added to Referrals worksheet
        wb = openpyxl.load_workbook(TEST_LEADS_FILE)
        ws_referrals = wb["Referrals"]
        
        final_contacts = []
        for r in range(2, ws_referrals.max_row + 1):
            row_vals = [ws_referrals.cell(row=r, column=c).value for c in (4, 8, 9)]
            final_contacts.append(row_vals)
            
        print("[E2E] Final directory contacts in Referrals sheet:", final_contacts)
        john_contact = [c for c in final_contacts if c[0] == "John Recruiter"]
        jane_contact = [c for c in final_contacts if c[0] == "Jane Recruiter"]
        jack_contact = [c for c in final_contacts if c[0] == "Jack Recruiter"]
        
        self.assertEqual(len(john_contact), 1, "Recruiter John Recruiter should be registered")
        self.assertEqual(john_contact[0][1], "Sent Recruiter Connection")
        self.assertEqual(john_contact[0][2], "Sent")
        
        self.assertEqual(len(jane_contact), 1, "Recruiter Jane Recruiter should be registered")
        self.assertEqual(jane_contact[0][1], "Sent Recruiter Connection")
        self.assertEqual(jane_contact[0][2], "Sent")
        
        self.assertEqual(len(jack_contact), 0, "Recruiter Jack Recruiter should NOT be registered (capped by limit)")
        
        # Verify Job Leads status is updated to Done
        ws_jobs = wb["Jobs"]
        job_statuses = {ws_jobs.cell(row=r, column=3).value: ws_jobs.cell(row=r, column=7).value for r in range(2, ws_jobs.max_row + 1)}
        print("[E2E] Job leads status mapping after Recruiter run:", job_statuses)
        self.assertEqual(job_statuses["Microsoft"], "Done", "Microsoft job status should be updated to Done")
        
        print("\n[E2E] REFERRAL AND RECRUITER OUTREACH E2E TEST WORKFLOW PASSED SUCCESSFULLY!")

if __name__ == "__main__":
    unittest.main()
