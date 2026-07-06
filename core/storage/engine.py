import os
import json
import time
import threading
from datetime import datetime
from config.settings import BASE_DIR
from config.constants import GOOGLE_SHEET_WORKSHEETS
from core.logging.config import logger

# ---------------------------------------------------------------------------
# Thread-safe environment override for CONNECTIFY_USER
# ---------------------------------------------------------------------------
_thread_local_env = threading.local()
_orig_setitem = os._Environ.__setitem__
_orig_getitem = os._Environ.__getitem__
_orig_delitem = os._Environ.__delitem__
_orig_getenv = os.getenv

def _thread_safe_setitem(self, key, value):
    if key == "CONNECTIFY_USER":
        _thread_local_env.connectify_user = value
    _orig_setitem(self, key, value)

def _thread_safe_getitem(self, key):
    if key == "CONNECTIFY_USER" and getattr(_thread_local_env, "connectify_user", None) is not None:
        return _thread_local_env.connectify_user
    return _orig_getitem(self, key)

def _thread_safe_delitem(self, key):
    if key == "CONNECTIFY_USER" and hasattr(_thread_local_env, "connectify_user"):
        delattr(_thread_local_env, "connectify_user")
    _orig_delitem(self, key)

def _thread_safe_getenv(key, default=None):
    if key == "CONNECTIFY_USER" and getattr(_thread_local_env, "connectify_user", None) is not None:
        return _thread_local_env.connectify_user
    return _orig_getenv(key, default)

os._Environ.__setitem__ = _thread_safe_setitem
os._Environ.__getitem__ = _thread_safe_getitem
os._Environ.__delitem__ = _thread_safe_delitem
os.getenv = _thread_safe_getenv


# ---------------------------------------------------------------------------
# Caching Layer (In-memory, Thread-Safe, TTL-based)
# ---------------------------------------------------------------------------
_cache_lock = threading.Lock()
_row_cache = {}          # { (username, table_key): (fetched_at_monotonic, data) }
_config_cache = {}       # { username: (fetched_at_monotonic, config_dict) }
CACHE_TTL_SECONDS = 30   # Cache lifetime for Google Sheets reads


def _get_cached_rows(username: str, table_key: str):
    with _cache_lock:
        entry = _row_cache.get((username, table_key))
        if entry:
            ts, data = entry
            if time.monotonic() - ts < CACHE_TTL_SECONDS:
                return data
    return None


def _set_cached_rows(username: str, table_key: str, data: list):
    with _cache_lock:
        _row_cache[(username, table_key)] = (time.monotonic(), data)


def _invalidate_cached_rows(username: str, table_key: str = None):
    with _cache_lock:
        if table_key:
            _row_cache.pop((username, table_key), None)
        else:
            keys_to_remove = [k for k in _row_cache.keys() if k[0] == username]
            for k in keys_to_remove:
                _row_cache.pop(k, None)


def _get_cached_config(username: str):
    with _cache_lock:
        entry = _config_cache.get(username)
        if entry:
            ts, config = entry
            if time.monotonic() - ts < CACHE_TTL_SECONDS:
                return config
    return None


def _set_cached_config(username: str, config: dict):
    with _cache_lock:
        _config_cache[username] = (time.monotonic(), config)


def _invalidate_cached_config(username: str):
    with _cache_lock:
        _config_cache.pop(username, None)


def _clear_all_caches():
    with _cache_lock:
        _row_cache.clear()
        _config_cache.clear()



# ---------------------------------------------------------------------------
# Dictionary Flattening/Unflattening Utilities
# ---------------------------------------------------------------------------
def flatten_dict(d, prefix=""):
    """Flattens a nested dictionary into a flat dictionary of dot-notated paths."""
    items = []
    for k, v in d.items():
        new_key = f"{prefix}.{k}" if prefix else k
        if isinstance(v, dict):
            items.extend(flatten_dict(v, new_key).items())
        else:
            items.append((new_key, v))
    return dict(items)


def unflatten_dict(d):
    """Restructures a dot-notated flat dictionary back into its nested format."""
    result = {}
    for k, v in d.items():
        parts = k.split(".")
        current = result
        for p in parts[:-1]:
            current = current.setdefault(p, {})
        
        # Auto-deserialize lists, nested JSON-like structures, booleans and numbers
        val = v
        if isinstance(v, str):
            v_stripped = v.strip()
            # Boolean coercion (handles Google Sheets True/False strings)
            if v_stripped.lower() == "true":
                val = True
            elif v_stripped.lower() == "false":
                val = False
            # JSON list/dict coercion
            elif (v_stripped.startswith("[") and v_stripped.endswith("]")) or (v_stripped.startswith("{") and v_stripped.endswith("}")):
                try:
                    val = json.loads(v_stripped)
                except Exception:
                    pass
            # Numeric coercion for pure integer/float strings (only for known numeric keys)
            elif parts[-1] in ("interval", "max_emails_per_run", "max_connections_per_company",
                                "max_connections_per_run", "target_count", "smtp_port",
                                "max_apply", "max_run_duration_seconds") and v_stripped.lstrip("-").replace(".", "", 1).isdigit():
                try:
                    val = int(v_stripped) if "." not in v_stripped else float(v_stripped)
                except Exception:
                    pass
        current[parts[-1]] = val
    return result


# ---------------------------------------------------------------------------
# Setting Metadata & Groups mapping for visual Google Sheets layout
# ---------------------------------------------------------------------------
SETTING_METADATA = {
    # Profile
    "profile.first_name": ("User Profile", "First Name"),
    "profile.last_name": ("User Profile", "Last Name"),
    "profile.email": ("User Profile", "Email Address"),
    "profile.phone": ("User Profile", "Phone Number"),
    "profile.linkedin_url": ("User Profile", "LinkedIn Profile URL"),
    "profile.resume_name": ("User Profile", "Resume Filename"),
    "profile.resume_url": ("User Profile", "Resume Short URL"),
    "profile.experience": ("User Profile", "Years of Experience"),
    "profile.current_location": ("User Profile", "Current Location"),
    "profile.preferred_locations": ("User Profile", "Preferred Locations"),
    "profile.current_ctc": ("User Profile", "Current CTC (LPA)"),
    "profile.expected_ctc": ("User Profile", "Expected CTC (LPA)"),
    "profile.notice_period": ("User Profile", "Notice Period"),
    "profile.last_working_day": ("User Profile", "Last Working Day"),

    # Email Scraper
    "email_scraper.interval": ("Email Scraper", "Search Execution Frequency (seconds)"),
    "email_scraper.max_emails_per_run": ("Email Scraper", "Max Emails per Run"),
    "email_scraper.review_mode": ("Email Scraper", "Review Mode Enabled (True/False)"),
    "email_scraper.sender_email": ("Email Scraper", "Custom Sender Email"),
    "email_scraper.search_keywords": ("Email Scraper", "Search Keywords"),
    "email_scraper.title_keywords": ("Email Scraper", "Title Keywords"),
    "email_scraper.keywords": ("Email Scraper", "General Keywords"),
    "email_scraper.excluded_keywords": ("Email Scraper", "Excluded Keywords"),
    "email_scraper.email_subject": ("Email Scraper", "Outreach Email Subject"),
    "email_scraper.email_template": ("Email Scraper", "Outreach Email Template"),
    "email_scraper.filter_experience_enabled": ("Email Scraper", "Filter Experience Enabled (True/False)"),
    "email_scraper.filter_experience_ranges": ("Email Scraper", "Acceptable Experience Ranges"),
    "email_scraper.filter_location_enabled": ("Email Scraper", "Filter Location Enabled (True/False)"),
    "email_scraper.filter_locations": ("Email Scraper", "Preferred Locations List"),

    # LinkedIn Connect
    "linkedin_connect.interval": ("LinkedIn Connections", "LinkedIn Connections Run Interval (minutes)"),
    "linkedin_connect.search_pages": ("LinkedIn Connections", "Pages to Search per Keyword"),
    "linkedin_connect.review_mode": ("LinkedIn Connections", "Review Mode Enabled (True/False)"),
    "linkedin_connect.max_connections_per_company": ("LinkedIn Connections", "Max Connections per Company"),
    "linkedin_connect.max_connections_per_run": ("LinkedIn Connections", "Max Connections per Run"),
    "linkedin_connect.search_keywords": ("LinkedIn Connections", "Search Keywords"),
    "linkedin_connect.title_keywords": ("LinkedIn Connections", "Title Keywords"),
    "linkedin_connect.keywords": ("LinkedIn Connections", "General Keywords"),
    "linkedin_connect.excluded_keywords": ("LinkedIn Connections", "Excluded Keywords"),
    "linkedin_connect.message_template": ("LinkedIn Connections", "Connection Note Template"),

    # Recruiter Outreach
    "recruiter_outreach.interval": ("Recruiter Outreach", "Recruiter Outreach Run Interval (minutes)"),
    "recruiter_outreach.target_count": ("Recruiter Outreach", "Target Count"),
    "recruiter_outreach.review_mode": ("Recruiter Outreach", "Review Mode Enabled (True/False)"),
    "recruiter_outreach.message_template": ("Recruiter Outreach", "Recruiter Message Template"),
    "recruiter_outreach.direct_message_template": ("Recruiter Outreach", "Direct Message Template"),

    # Referral Outreach
    "referral_outreach.message_template": ("Referral Outreach", "Referral Request Template"),

    # Global Settings
    "global_settings.database_type": ("Global Settings", "Database Storage Type"),
    "global_settings.google_sheet_url": ("Global Settings", "Google Sheet URL"),
    "global_settings.google_credentials_json": ("Global Settings", "Google Service Account Credentials (JSON)"),
    "global_settings.linkedin_email": ("Global Settings", "LinkedIn Login Email"),
    "global_settings.linkedin_password": ("Global Settings", "LinkedIn Login Password"),
    "global_settings.smtp_email": ("Global Settings", "SMTP Server Email"),
    "global_settings.smtp_password": ("Global Settings", "SMTP Server Password"),
    "global_settings.smtp_server": ("Global Settings", "SMTP Server Host"),
    "global_settings.smtp_port": ("Global Settings", "SMTP Server Port"),
    "global_settings.search_location": ("Global Settings", "LinkedIn Search Location Target"),
    "global_settings.search_time_range": ("Global Settings", "LinkedIn Search Time Code"),
    "global_settings.dry_run": ("Global Settings", "Dry Run Mode (1 for Yes, 0 for No)"),
    "global_settings.max_apply": ("Global Settings", "Max Auto-Applications per Job"),
    "global_settings.max_run_duration_seconds": ("Global Settings", "Max Pipeline Run Time (seconds)"),
}

def get_setting_meta(key: str):
    """Returns a tuple of (Category, Setting Name) for a given technical key."""
    if key in SETTING_METADATA:
        return SETTING_METADATA[key]
    
    # Dynamic fallback for any future added settings
    parts = key.split(".", 1)
    if len(parts) == 2:
        cat_raw, name_raw = parts
        cat = cat_raw.replace("_", " ").title()
        name = name_raw.replace("_", " ").title()
    else:
        cat = "Other Settings"
        name = key.replace("_", " ").title()
    return cat, name


# ---------------------------------------------------------------------------
# Base Storage Provider Interface
# ---------------------------------------------------------------------------
class BaseStorageProvider:
    def get_config(self, username: str, bypass_cache: bool = False) -> dict:
        raise NotImplementedError()

    def save_config(self, username: str, config: dict):
        raise NotImplementedError()

    def read_rows(self, username: str, table_key: str, bypass_cache: bool = False) -> list:
        raise NotImplementedError()

    def write_rows(self, username: str, table_key: str, data: list):
        raise NotImplementedError()

    def append_row(self, username: str, table_key: str, row: dict):
        raise NotImplementedError()


# ---------------------------------------------------------------------------
# Local Storage Provider (Default)
# ---------------------------------------------------------------------------
class LocalStorageProvider(BaseStorageProvider):
    def get_config_path(self, username: str) -> str:
        return os.path.join(BASE_DIR, "users", username, "config.json")

    def get_config(self, username: str, bypass_cache: bool = False) -> dict:
        path = self.get_config_path(username)
        if os.path.exists(path):
            try:
                with open(path, "r", encoding="utf-8") as f:
                    return json.load(f)
            except Exception as e:
                logger.error(f"Error reading local config for {username}: {e}")
        return {}

    def save_config(self, username: str, config: dict):
        path = self.get_config_path(username)
        os.makedirs(os.path.dirname(path), exist_ok=True)
        try:
            with open(path, "w", encoding="utf-8") as f:
                json.dump(config, f, indent=2)
        except Exception as e:
            logger.error(f"Error saving local config for {username}: {e}")


    def get_excel_path(self, username: str, table_key: str) -> str:
        filename_map = {
            "jobs": "LinkedIn_Job_Tracker.xlsx",
            "emails": "job_tracker.xlsx",
            "referrals": "referrals.xlsx"
        }
        if table_key not in filename_map:
            raise ValueError(f"Unknown table key: {table_key}")
        return os.path.join(BASE_DIR, "users", username, "data", filename_map[table_key])

    def read_rows(self, username: str, table_key: str, bypass_cache: bool = False) -> list:
        path = self.get_excel_path(username, table_key)
        if not os.path.exists(path):
            # Auto-initialize with empty headers if file is missing
            self.write_rows(username, table_key, [])
            return []

        try:
            import openpyxl
            wb = openpyxl.load_workbook(path, data_only=True)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            if not headers:
                return []
            
            rows = []
            for r in range(2, ws.max_row + 1):
                row_dict = {}
                for col_idx, h in enumerate(headers, start=1):
                    val = ws.cell(row=r, column=col_idx).value
                    row_dict[h] = val if val is not None else ""
                rows.append(row_dict)
            return rows
        except Exception as e:
            logger.error(f"Error reading local Excel database file '{path}': {e}")
            return []

    def write_rows(self, username: str, table_key: str, data: list):
        path = self.get_excel_path(username, table_key)
        os.makedirs(os.path.dirname(path), exist_ok=True)
        
        headers = GOOGLE_SHEET_WORKSHEETS[table_key]["headers"]
        sheet_title = GOOGLE_SHEET_WORKSHEETS[table_key]["name"].replace(" & ", " ")
        
        try:
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = sheet_title
            ws.append(headers)
            
            for row in data:
                row_vals = []
                for h in headers:
                    val = row.get(h)
                    if val is not None and val != "":
                        val_str = str(val).strip()
                        if val_str.isdigit():
                            val = int(val_str)
                        else:
                            try:
                                # Only check float representation if it has a decimal point
                                if "." in val_str:
                                    val = float(val_str)
                                    if val.is_integer():
                                        val = int(val)
                            except ValueError:
                                pass
                    else:
                        val = ""
                    row_vals.append(val)
                ws.append(row_vals)
            
            wb.save(path)
        except Exception as e:
            logger.error(f"Error writing to local Excel database file '{path}': {e}")

    def append_row(self, username: str, table_key: str, row: dict):
        path = self.get_excel_path(username, table_key)
        if not os.path.exists(path):
            self.write_rows(username, table_key, [row])
            return

        headers = GOOGLE_SHEET_WORKSHEETS[table_key]["headers"]
        try:
            import openpyxl
            wb = openpyxl.load_workbook(path)
            ws = wb.active
            row_vals = []
            for h in headers:
                val = row.get(h)
                if val is not None and val != "":
                    val_str = str(val).strip()
                    if val_str.isdigit():
                        val = int(val_str)
                    else:
                        try:
                            if "." in val_str:
                                val = float(val_str)
                                if val.is_integer():
                                    val = int(val)
                        except ValueError:
                            pass
                else:
                    val = ""
                row_vals.append(val)
            ws.append(row_vals)
            wb.save(path)
        except Exception as e:
            logger.error(f"Error appending row to local Excel database file '{path}': {e}")




# ---------------------------------------------------------------------------
# Google Sheets Storage Provider (Centralized Cloud Backend)
# ---------------------------------------------------------------------------
class GoogleSheetsStorageProvider(BaseStorageProvider):
    def get_sheets_config(self, username: str):
        """Extracts Sheets connection credentials from the local bootstrap configuration."""
        path = os.path.join(BASE_DIR, "users", username, "config.json")
        if os.path.exists(path):
            try:
                with open(path, "r", encoding="utf-8") as f:
                    cfg = json.load(f)
                    gs = cfg.get("global_settings", {})
                    url = gs.get("google_sheet_url")
                    creds = gs.get("google_credentials_json")
                    if url and creds:
                        return url, creds
            except Exception:
                pass
        return None

    def get_config(self, username: str, bypass_cache: bool = False) -> dict:
        # 1. Check in-memory cache
        if not bypass_cache:
            cached = _get_cached_config(username)
            if cached is not None:
                return cached

        # 2. Retrieve Sheets Credentials
        sheets_conf = self.get_sheets_config(username)
        if not sheets_conf:
            logger.warning(f"Google Sheets not configured for user {username}. Falling back to Local Storage configs.")
            return LocalStorageProvider().get_config(username, bypass_cache=bypass_cache)

        url, creds_content = sheets_conf
        
        profile_ws = GOOGLE_SHEET_WORKSHEETS["profile"]["name"]
        templates_ws = GOOGLE_SHEET_WORKSHEETS["templates"]["name"]
        keywords_ws = GOOGLE_SHEET_WORKSHEETS["keywords"]["name"]
        settings_ws = GOOGLE_SHEET_WORKSHEETS["settings"]["name"]

        # Cache invalidation for all worksheets if cache is bypassed
        if bypass_cache:
            try:
                from core.storage.sheets import _cache_invalidate
                for ws_name in (profile_ws, templates_ws, keywords_ws, settings_ws):
                    _cache_invalidate(ws_name)
            except Exception:
                pass

        try:
            from core.storage.sheets import read_rows, ensure_worksheets_exist
            # Ensure all new sheets exist and visual formatting is applied
            ensure_worksheets_exist(url, creds_content)

            profile_rows = read_rows(url, creds_content, profile_ws)
            if not profile_rows:
                logger.warning(f"Google Sheets User Profile is empty for user {username}. Falling back to Local Storage configs and auto-uploading to sync.")
                local_config = LocalStorageProvider().get_config(username, bypass_cache=bypass_cache)
                if local_config:
                    try:
                        self.save_config(username, local_config)
                    except Exception as upload_err:
                        logger.error(f"Failed to auto-populate empty Google Sheet settings: {upload_err}")
                _set_cached_config(username, local_config)
                return local_config

            templates_rows = read_rows(url, creds_content, templates_ws)
            keywords_rows = read_rows(url, creds_content, keywords_ws)
            settings_rows = read_rows(url, creds_content, settings_ws)

            # Reconstruct flat config dict from modular sheets
            flat_dict = {}

            # A. Profile
            profile_row = profile_rows[0] if profile_rows else {}
            profile_field_mapping = {
                "First Name": "profile.first_name", "Last Name": "profile.last_name", "Email Address": "profile.email",
                "Phone Number": "profile.phone", "LinkedIn URL": "profile.linkedin_url", "Resume Filename": "profile.resume_name",
                "Resume Short URL": "profile.resume_url", "Years of Experience": "profile.experience",
                "Current Location": "profile.current_location", "Preferred Locations": "profile.preferred_locations",
                "Current CTC": "profile.current_ctc", "Expected CTC": "profile.expected_ctc",
                "Notice Period": "profile.notice_period", "Last Working Day": "profile.last_working_day"
            }
            for header, flat_key in profile_field_mapping.items():
                val = profile_row.get(header)
                if val is not None:
                    flat_dict[flat_key] = val

            # B. Message Templates
            for r in templates_rows:
                key = r.get("Key")
                subj = r.get("Subject")
                body = r.get("Body")
                if key:
                    flat_dict[key] = body if body is not None else ""
                    if key == "email_scraper.email_template" and subj:
                        flat_dict["email_scraper.email_subject"] = subj

            # C. Keyword Lists
            keyword_mapping = {
                "Scraper Search Keywords": "email_scraper.search_keywords",
                "Scraper Title Keywords": "email_scraper.title_keywords",
                "Scraper Excluded Keywords": "email_scraper.excluded_keywords",
                "Connect Search Keywords": "linkedin_connect.search_keywords",
                "Connect Title Keywords": "linkedin_connect.title_keywords",
                "Connect Excluded Keywords": "linkedin_connect.excluded_keywords"
            }
            temp_kw_lists = {flat_key: [] for flat_key in keyword_mapping.values()}
            for row in keywords_rows:
                for header, flat_key in keyword_mapping.items():
                    val = row.get(header)
                    if val is not None and str(val).strip():
                        temp_kw_lists[flat_key].append(str(val).strip())
            for flat_key, lst in temp_kw_lists.items():
                flat_dict[flat_key] = lst

            # D. Application Settings
            for r in settings_rows:
                key = r.get("Key")
                val = r.get("Value")
                if key:
                    flat_dict[key] = val if val is not None else ""

            # Unflatten dictionary to nested config format
            config_dict = unflatten_dict(flat_dict)

            # Merge with local bootstrap configuration (deep merge to preserve defaults not in Sheets)
            local_config = LocalStorageProvider().get_config(username, bypass_cache=bypass_cache)
            
            def deep_merge(target, source):
                for k, v in source.items():
                    if k not in target:
                        target[k] = v
                    elif isinstance(v, dict) and isinstance(target[k], dict):
                        deep_merge(target[k], v)
                        
            if local_config:
                deep_merge(config_dict, local_config)

            _set_cached_config(username, config_dict)
            return config_dict
        except Exception as e:
            logger.error(f"Error loading config from Google Sheet for user {username}: {e}. Falling back to Local.")
            fallback_conf = LocalStorageProvider().get_config(username, bypass_cache=bypass_cache)
            _set_cached_config(username, fallback_conf)
            return fallback_conf

    def save_config(self, username: str, config: dict):
        # 1. Save local bootstrap settings first (handles credentials storage)
        LocalStorageProvider().save_config(username, config)

        # 2. Check Sheets Credentials
        sheets_conf = self.get_sheets_config(username)
        if not sheets_conf:
            return

        url, creds_content = sheets_conf
        flat_dict = flatten_dict(config)

        # A. Profile data
        profile_headers = GOOGLE_SHEET_WORKSHEETS["profile"]["headers"]
        profile_field_mapping = {
            "First Name": "profile.first_name", "Last Name": "profile.last_name", "Email Address": "profile.email",
            "Phone Number": "profile.phone", "LinkedIn URL": "profile.linkedin_url", "Resume Filename": "profile.resume_name",
            "Resume Short URL": "profile.resume_url", "Years of Experience": "profile.experience",
            "Current Location": "profile.current_location", "Preferred Locations": "profile.preferred_locations",
            "Current CTC": "profile.current_ctc", "Expected CTC": "profile.expected_ctc",
            "Notice Period": "profile.notice_period", "Last Working Day": "profile.last_working_day"
        }
        profile_row = {}
        for header, flat_key in profile_field_mapping.items():
            profile_row[header] = str(flat_dict.get(flat_key, ""))
        profile_data_dicts = [profile_row]

        # B. Templates data
        templates_data_dicts = [
            {"Template Name": "Outreach Email", "Subject": str(flat_dict.get("email_scraper.email_subject", "")), "Body": str(flat_dict.get("email_scraper.email_template", "")), "Key": "email_scraper.email_template"},
            {"Template Name": "LinkedIn Connection Note", "Subject": "", "Body": str(flat_dict.get("linkedin_connect.message_template", "")), "Key": "linkedin_connect.message_template"},
            {"Template Name": "Recruiter Message", "Subject": "", "Body": str(flat_dict.get("recruiter_outreach.message_template", "")), "Key": "recruiter_outreach.message_template"},
            {"Template Name": "Recruiter Direct Message", "Subject": "", "Body": str(flat_dict.get("recruiter_outreach.direct_message_template", "")), "Key": "recruiter_outreach.direct_message_template"},
            {"Template Name": "Referral Request", "Subject": "", "Body": str(flat_dict.get("referral_outreach.message_template", "")), "Key": "referral_outreach.message_template"}
        ]

        # C. Keyword Lists data
        keyword_mapping = {
            "Scraper Search Keywords": "email_scraper.search_keywords",
            "Scraper Title Keywords": "email_scraper.title_keywords",
            "Scraper Excluded Keywords": "email_scraper.excluded_keywords",
            "Connect Search Keywords": "linkedin_connect.search_keywords",
            "Connect Title Keywords": "linkedin_connect.title_keywords",
            "Connect Excluded Keywords": "linkedin_connect.excluded_keywords"
        }
        kw_lists = []
        for header in [
            "Scraper Search Keywords", "Scraper Title Keywords", "Scraper Excluded Keywords",
            "Connect Search Keywords", "Connect Title Keywords", "Connect Excluded Keywords"
        ]:
            flat_key = keyword_mapping[header]
            val = flat_dict.get(flat_key, [])
            if isinstance(val, str):
                try:
                    lst = json.loads(val)
                except Exception:
                    lst = [x.strip() for x in val.split(",") if x.strip()]
            elif isinstance(val, list):
                lst = val
            else:
                lst = []
            kw_lists.append(lst)

        max_len = max(len(lst) for lst in kw_lists) if kw_lists else 0
        keywords_data_dicts = []
        for i in range(max_len):
            row = {}
            cols = [
                "Scraper Search Keywords", "Scraper Title Keywords", "Scraper Excluded Keywords",
                "Connect Search Keywords", "Connect Title Keywords", "Connect Excluded Keywords"
            ]
            for idx, col in enumerate(cols):
                row[col] = kw_lists[idx][i] if i < len(kw_lists[idx]) else ""
            keywords_data_dicts.append(row)

        # D. Settings data
        settings_data_dicts = []
        for k, v in flat_dict.items():
            if (k.startswith("profile.") or 
                k.endswith("_template") or 
                k.endswith("_keywords") or 
                k == "email_scraper.email_subject" or
                k == "recruiter_outreach.direct_message_template" or
                k == "global_settings.google_credentials_json"):
                continue

            val_str = json.dumps(v) if isinstance(v, (list, dict)) else str(v)
            cat, name = get_setting_meta(k)
            settings_data_dicts.append({
                "Category": cat,
                "Setting Name": name,
                "Value": val_str,
                "Key": k
            })
        settings_data_dicts.sort(key=lambda x: (x["Category"], x["Setting Name"]))

        try:
            from core.storage.sheets import write_rows, ensure_worksheets_exist
            ensure_worksheets_exist(url, creds_content)

            write_rows(url, creds_content, GOOGLE_SHEET_WORKSHEETS["profile"]["name"], profile_data_dicts)
            write_rows(url, creds_content, GOOGLE_SHEET_WORKSHEETS["templates"]["name"], templates_data_dicts)
            write_rows(url, creds_content, GOOGLE_SHEET_WORKSHEETS["keywords"]["name"], keywords_data_dicts)
            write_rows(url, creds_content, GOOGLE_SHEET_WORKSHEETS["settings"]["name"], settings_data_dicts)

            _invalidate_cached_config(username)
            _set_cached_config(username, config)
        except Exception as e:
            logger.error(f"Failed to save configuration to Google Sheet: {e}")

    def read_rows(self, username: str, table_key: str, bypass_cache: bool = False) -> list:
        # Check cache first
        if not bypass_cache:
            cached = _get_cached_rows(username, table_key)
            if cached is not None:
                return cached

        sheets_conf = self.get_sheets_config(username)
        if not sheets_conf:
            logger.warning(f"Google Sheets not configured. Reading from local database instead.")
            return LocalStorageProvider().read_rows(username, table_key, bypass_cache=bypass_cache)

        url, creds_content = sheets_conf
        ws_name = GOOGLE_SHEET_WORKSHEETS[table_key]["name"]

        try:
            from core.storage.sheets import read_rows
            
            # If bypassing cache, we must invalidate gspread's internally held cache in sheets.py
            if bypass_cache:
                try:
                    from core.storage.sheets import _cache_invalidate
                    _cache_invalidate(ws_name)
                except Exception:
                    pass

            data = read_rows(url, creds_content, ws_name)
            
            # Standardize ID types (convert decimal keys/ID to int strings for consistency with Local Excel)
            id_col = "JobID" if table_key == "jobs" else ("ID" if table_key == "emails" else "ReferralID")
            for row in data:
                if id_col in row and row[id_col] != "":
                    try:
                        row[id_col] = int(float(row[id_col]))
                    except (ValueError, TypeError):
                        pass

            # --- AUTOMATIC SELF-HEALING SYNC ---
            try:
                local_data = LocalStorageProvider().read_rows(username, table_key, bypass_cache=True)
                if local_data:
                    from core.utils.url_utils import normalize_external_url
                    sheets_ids = {str(r.get(id_col)).strip().rstrip(".0") for r in data if r.get(id_col)}
                    
                    # For jobs, build a lookup set of normalized URLs to prevent duplicates by URL
                    sheets_urls = set()
                    if table_key == "jobs":
                        for r in data:
                            c_url = normalize_external_url(r.get("CompanyURL") or "")
                            lc_url = normalize_external_url(r.get("LinkedIn_Company_URL") or "")
                            if c_url:
                                sheets_urls.add(c_url)
                            if lc_url and c_url:
                                sheets_urls.add((lc_url, c_url))
                                
                    missing_rows = []
                    for lr in local_data:
                        lid = str(lr.get(id_col)).strip().rstrip(".0")
                        
                        # Primary key ID check
                        if lid and lid in sheets_ids:
                            continue
                            
                        # URL matching checks for jobs
                        if table_key == "jobs":
                            c_url = normalize_external_url(lr.get("CompanyURL") or "")
                            lc_url = normalize_external_url(lr.get("LinkedIn_Company_URL") or "")
                            if c_url and c_url in sheets_urls:
                                continue
                            if lc_url and c_url and (lc_url, c_url) in sheets_urls:
                                continue
                                
                        # Age check: skip syncing rows created less than 120 seconds ago to avoid Sheets delay issues
                        created_str = lr.get("CreatedDateTime") or lr.get("Timestamp")
                        is_recent = False
                        if created_str:
                            for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S.%f", "%Y-%m-%dT%H:%M:%S"):
                                try:
                                    created_dt = datetime.strptime(str(created_str).split(".")[0], fmt)
                                    age_seconds = (datetime.now() - created_dt).total_seconds()
                                    if age_seconds < 120:
                                        is_recent = True
                                        break
                                except ValueError:
                                    continue
                                    
                        if is_recent:
                            continue
                            
                        missing_rows.append(lr)
                    
                    if missing_rows:
                        logger.info(f"Self-healing sync: Found {len(missing_rows)} missing rows in Google Sheets for '{table_key}'. Syncing now...")
                        from core.storage.sheets import append_row
                        for mr in missing_rows:
                            try:
                                append_row(url, creds_content, ws_name, mr)
                                data.append(mr.copy())
                            except Exception as se:
                                logger.warning(f"Self-healing sync failed to append row {mr.get(id_col)}: {se}")
            except Exception as le:
                logger.warning(f"Self-healing sync: Error reading local database for comparison: {le}")

            if not bypass_cache:
                _set_cached_rows(username, table_key, data)
            return data
        except Exception as e:
            logger.error(f"Error reading Google Sheets worksheet '{ws_name}': {e}. Falling back to Local.")
            return LocalStorageProvider().read_rows(username, table_key, bypass_cache=bypass_cache)

    def write_rows(self, username: str, table_key: str, data: list):
        # 1. ALWAYS write to local Excel database first as a guaranteed offline mirror copy
        try:
            LocalStorageProvider().write_rows(username, table_key, data)
        except Exception as le:
            logger.warning(f"Failed to write local database mirror backup for '{table_key}': {le}")

        sheets_conf = self.get_sheets_config(username)
        if not sheets_conf:
            return

        url, creds_content = sheets_conf
        ws_name = GOOGLE_SHEET_WORKSHEETS[table_key]["name"]

        # 2. Try to write to Google Sheets online copy
        try:
            from core.storage.sheets import write_rows
            write_rows(url, creds_content, ws_name, data)
            
            # Invalidate and reset cache
            _invalidate_cached_rows(username, table_key)
            _set_cached_rows(username, table_key, data)
        except Exception as e:
            logger.error(f"Error writing to Google Sheets worksheet '{ws_name}': {e}. (Saved locally, will sync later).")

    def append_row(self, username: str, table_key: str, row: dict):
        # 1. ALWAYS write to local Excel database first as a guaranteed offline mirror copy
        try:
            LocalStorageProvider().append_row(username, table_key, row)
        except Exception as le:
            logger.warning(f"Failed to append local database mirror backup for '{table_key}': {le}")

        sheets_conf = self.get_sheets_config(username)
        if not sheets_conf:
            return

        url, creds_content = sheets_conf
        ws_name = GOOGLE_SHEET_WORKSHEETS[table_key]["name"]

        # 2. Try to write to Google Sheets online copy
        try:
            from core.storage.sheets import append_row
            append_row(url, creds_content, ws_name, row)
            
            # Update cache in-memory if present, and slide the TTL freshness window
            with _cache_lock:
                entry = _row_cache.get((username, table_key))
                if entry:
                    ts, data = entry
                    data.append(row.copy())
                    _row_cache[(username, table_key)] = (time.monotonic(), data)
        except Exception as e:
            logger.error(f"Error appending to Google Sheets worksheet '{ws_name}': {e}. (Saved locally, will sync later).")


# ---------------------------------------------------------------------------
# Storage Manager (Dynamic Provider Resolver)
# ---------------------------------------------------------------------------
class StorageManager:
    _instance = None
    _lock = threading.Lock()

    def __new__(cls):
        with cls._lock:
            if cls._instance is None:
                cls._instance = super(StorageManager, cls).__new__(cls)
                cls._instance.providers = {
                    "local": LocalStorageProvider(),
                    "google_sheets": GoogleSheetsStorageProvider()
                }
            return cls._instance

    def get_provider(self, username: str) -> BaseStorageProvider:
        # Determine the user's active database type from their local bootstrap config file
        path = os.path.join(BASE_DIR, "users", username, "config.json")
        db_type = "local"
        if os.path.exists(path):
            try:
                with open(path, "r", encoding="utf-8") as f:
                    cfg = json.load(f)
                    db_type = cfg.get("global_settings", {}).get("database_type", "local")
            except Exception:
                pass
        return self.providers.get(db_type, self.providers["local"])


# ---------------------------------------------------------------------------
# Global Access Helpers
# ---------------------------------------------------------------------------
def get_active_username():
    """Resolves active runner username via env or config file."""
    env_user = os.getenv("CONNECTIFY_USER")
    if env_user:
        return env_user
        
    active_user_file = os.path.join(BASE_DIR, "users", "active_user.json")
    if os.path.exists(active_user_file):
        try:
            with open(active_user_file, "r") as f:
                return json.load(f).get("selected_user") or "default"
        except Exception:
            pass
    return "default"


def get_active_storage_provider() -> BaseStorageProvider:
    username = get_active_username()
    return StorageManager().get_provider(username)


def get_user_config(username: str = None, bypass_cache: bool = False) -> dict:
    if not username:
        username = get_active_username()
    if not bypass_cache:
        cached = _get_cached_config(username)
        if cached is not None:
            return cached
    config = StorageManager().get_provider(username).get_config(username, bypass_cache=bypass_cache)
    if not bypass_cache:
        _set_cached_config(username, config)
    return config


def save_user_config(config: dict, username: str = None):
    if not username:
        username = get_active_username()
    StorageManager().get_provider(username).save_config(username, config)
    _invalidate_cached_config(username)
    _set_cached_config(username, config)


def read_database_rows(table_key: str, username: str = None, bypass_cache: bool = False) -> list:
    if not username:
        username = get_active_username()
    if not bypass_cache:
        cached = _get_cached_rows(username, table_key)
        if cached is not None:
            return cached
    data = StorageManager().get_provider(username).read_rows(username, table_key, bypass_cache=bypass_cache)
    if not bypass_cache:
        _set_cached_rows(username, table_key, data)
    return data


def write_database_rows(table_key: str, data: list, username: str = None):
    if not username:
        username = get_active_username()
    StorageManager().get_provider(username).write_rows(username, table_key, data)
    _invalidate_cached_rows(username, table_key)
    _set_cached_rows(username, table_key, data)


def append_database_row(table_key: str, row: dict, username: str = None):
    if not username:
        username = get_active_username()
    StorageManager().get_provider(username).append_row(username, table_key, row)
    
    # Update cache in-memory if present, and slide the TTL freshness window
    with _cache_lock:
        entry = _row_cache.get((username, table_key))
        if entry:
            ts, data = entry
            data.append(row.copy())
            _row_cache[(username, table_key)] = (time.monotonic(), data)

