import os
import sys
import openpyxl
from unittest.mock import MagicMock, patch

# Ensure the root directory is in the path
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

from config.settings import get_job_tracker_file
from config.user_profiles import get_selected_user_name, load_all_configs
from pipelines.email_outreach.pipeline import run_phase_two

class DummyScraper:
    def __init__(self):
        self.driver = MagicMock()

def setup_dummy_job_tracker():
    tracker_file = get_job_tracker_file()
    os.makedirs(os.path.dirname(tracker_file), exist_ok=True)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["ID", "Email", "Status", "Keyword", "Timestamp"])
    ws.append([1, "manager1@test.com", "", "Python DBA", ""])
    ws.append([2, "manager2@test.com", "", "Java DBA", ""])
    ws.append([3, "manager3@test.com", "", "C# DBA", ""])
    wb.save(tracker_file)
    print(f"Created dummy job tracker at {tracker_file}")
    return tracker_file

def test_skipped_status():
    tracker_file = setup_dummy_job_tracker()
    
    # Set scraper review_mode = True in active user's config
    username = get_selected_user_name()
    config = load_all_configs()
    if username in config.get("users", {}):
        config["users"][username]["email_scraper"]["review_mode"] = True
        from config.user_profiles import save_all_configs
        save_all_configs(config)
        print(f"Set review_mode = True for {username}")

    scraper = DummyScraper()
    
    # We will mock send_email_via_gmail to return "skipped" for the first email and "quit" for the second
    mock_send = MagicMock(side_effect=lambda driver, email, review_mode=None: "skipped" if email == "manager1@test.com" else "quit")
    
    print("\n--- Running Phase Two with mocked send_email_via_gmail (Skip, then Quit) ---")
    with patch('pipelines.email_outreach.pipeline.send_email_via_gmail', mock_send):
        run_phase_two(scraper, review_mode=True)
        
    # Verify the results in the Excel file
    wb = openpyxl.load_workbook(tracker_file)
    ws = wb.active
    
    row2_email = ws.cell(row=2, column=2).value
    row2_status = ws.cell(row=2, column=3).value
    row3_email = ws.cell(row=3, column=2).value
    row3_status = ws.cell(row=3, column=3).value
    row4_email = ws.cell(row=4, column=2).value
    row4_status = ws.cell(row=4, column=3).value
    
    print(f"After Run 1:")
    print(f"Row 2 (manager1): Email={row2_email}, Status={row2_status}")
    print(f"Row 3 (manager2): Email={row3_email}, Status={row3_status}")
    print(f"Row 4 (manager3): Email={row4_email}, Status={row4_status}")
    
    # manager1@test.com should have been skipped, so status should be 'skipped'
    assert row2_status == "skipped", f"Expected manager1 status to be 'skipped', got: {row2_status}"
    
    # manager2@test.com was processed but returned "quit", status should remain empty/None
    assert row3_status is None or row3_status == "", f"Expected manager2 status to be empty, got: {row3_status}"
    
    # 2. Now run the pipeline again. Since manager1 is already 'skipped', it should skip it automatically and call send_email_via_gmail directly for manager2!
    mock_send_run2 = MagicMock(side_effect=lambda driver, email, review_mode=None: "quit")
    
    print("\n--- Running Phase Two again (Run 2) ---")
    with patch('pipelines.email_outreach.pipeline.send_email_via_gmail', mock_send_run2):
        run_phase_two(scraper, review_mode=True)
        
    wb = openpyxl.load_workbook(tracker_file)
    ws = wb.active
    row2_status = ws.cell(row=2, column=3).value
    row3_status = ws.cell(row=3, column=3).value
    
    print(f"After Run 2:")
    print(f"Row 2 (manager1): Status={row2_status}")
    print(f"Row 3 (manager2): Status={row3_status}")
    
    # Since manager1 was skipped/ignored in run 2, it should not have called mock_send_run2 for it.
    # The first mock call should be for manager2!
    assert mock_send_run2.call_count == 1, f"Expected send_email_via_gmail to be called once (for manager2), but was called {mock_send_run2.call_count} times."
    
    # Verify the argument of the single call is manager2@test.com
    called_email = mock_send_run2.call_args[0][1]
    assert called_email == "manager2@test.com", f"Expected to process manager2@test.com, but processed: {called_email}"
    
    print("\nALL SKIPPED STATUS UNIT TESTS PASSED SUCCESSFULLY!")

if __name__ == "__main__":
    try:
        test_skipped_status()
    finally:
        # Clean up dummy file
        tracker_file = get_job_tracker_file()
        if os.path.exists(tracker_file):
            os.remove(tracker_file)
            print("Cleaned up dummy job tracker.")
