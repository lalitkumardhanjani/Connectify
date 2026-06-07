import os
import sys
import openpyxl
import json
import urllib.request
import shutil

# Ensure the root directory is in the path
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

from config.settings import get_job_tracker_file
from core.storage.database import init_scraper_store

def test_edit_row():
    tracker_file = get_job_tracker_file()
    backup_file = tracker_file + ".bak"
    
    # Back up current tracker file if it exists
    has_backup = False
    if os.path.exists(tracker_file):
        shutil.copyfile(tracker_file, backup_file)
        has_backup = True
        print(f"Backed up current job tracker to {backup_file}")
    
    try:
        # Create a fresh test tracker file
        if os.path.exists(tracker_file):
            os.remove(tracker_file)
            
        init_scraper_store()
        
        # Add test rows: one 'New'
        wb = openpyxl.load_workbook(tracker_file)
        ws = wb.active
        ws.append([1, "editme@example.com", "New", "Python", "2026-06-07T08:50:00"])
        wb.save(tracker_file)
        print("Set up test database rows.")
        
        # Call the /api/data/edit_row endpoint to change status to "Skipped"
        payload = {
            "db_type": "scraper",
            "id": 1,
            "email": "editme@example.com",
            "status": "Skipped",
            "keyword": "Python"
        }
        
        req = urllib.request.Request(
            "http://127.0.0.1:5001/api/data/edit_row",
            data=json.dumps(payload).encode('utf-8'),
            headers={'Content-Type': 'application/json'}
        )
        
        print("Sending edit request to /api/data/edit_row...")
        with urllib.request.urlopen(req) as response:
            res_data = json.loads(response.read().decode())
            print("Response:", res_data)
            
        assert res_data.get("status") == "success", f"Expected 'success', got: {res_data}"
        
        # Verify the change was persisted in the Excel file
        wb = openpyxl.load_workbook(tracker_file)
        ws = wb.active
        status_value = ws.cell(row=2, column=3).value
        print(f"Persisted status in Excel file: {status_value}")
        
        assert status_value == "Skipped", f"Expected 'Skipped', got: {status_value}"
        print("Edit row API integration test passed successfully!")
        
    finally:
        # Restore backup
        if os.path.exists(tracker_file):
            os.remove(tracker_file)
        if has_backup:
            shutil.copyfile(backup_file, tracker_file)
            os.remove(backup_file)
            print("Restored original job tracker.")
        else:
            print("No original job tracker to restore.")

if __name__ == "__main__":
    test_edit_row()
