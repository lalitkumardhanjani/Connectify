import os
import openpyxl
from config.settings import JOB_TRACKER_FILE
from config.user_profiles import get_selected_user_config
from config.constants import DBA_KEYWORDS_DEFAULT
from core.integrations.selenium_driver import get_driver
from core.storage.database import init_scraper_store, update_status
from core.logging.config import logger
from core.utils.string_utils import parse_preferred_locations

from pipelines.email_outreach.services.scraper import LinkedInScraper, ScraperTargetReached
from pipelines.email_outreach.services.sender import send_email_via_gmail

def run_phase_one(scraper):
    """Executes email scraping from LinkedIn posts based on selected keyword list."""
    init_scraper_store()
    user_conf = get_selected_user_config()
    email_scraper_conf = user_conf.get("email_scraper", {})
    keywords = email_scraper_conf.get("search_keywords") or email_scraper_conf.get("keywords") or DBA_KEYWORDS_DEFAULT

    # Read the user-configured per-keyword timeout (Search Execution Frequency setting).
    # Default to 60s if not set. This is the maximum time spent scrolling each keyword
    # before moving on to the next one.
    try:
        timeout_seconds = int(email_scraper_conf.get("interval") or 60)
    except (ValueError, TypeError):
        timeout_seconds = 60
    logger.info(f"Per-keyword timeout set to {timeout_seconds}s (from Search Execution Frequency setting).")
    
    # Retrieve preferred locations from candidate profile
    profile = user_conf.get("profile", {})
    pref_location = profile.get("preferred_locations", "")
    locations = parse_preferred_locations(pref_location)
    if not locations:
        locations = [""]
    
    try:
        for loc in locations:
            for kw in keywords:
                search_query = f"{kw} job {loc}".strip() if loc else f"{kw} job"
                logger.info(f"=== Processing keyword: '{kw}' at location: '{loc}' (Search query: '{search_query}') ===")
                if scraper.search_for_keyword(search_query):
                    scraper.process_keyword(kw, timeout_seconds=timeout_seconds)
                else:
                    logger.warning(f"Skipping query '{search_query}' due to navigation failure.")
    except ScraperTargetReached as e:
        logger.info(f"Phase 1 scraping completed early: {e}")


def run_phase_two(scraper, review_mode=None):
    """Executes composing and sending emails via Selenium browser to discovered addresses."""
    if not os.path.exists(JOB_TRACKER_FILE):
        logger.warning(f"Excel database '{JOB_TRACKER_FILE}' not found – nothing to send. Please run Phase 1 (Fetch Emails) first to scrape contact leads.")
        return
        
    wb = openpyxl.load_workbook(JOB_TRACKER_FILE)
    ws = wb.active
    col_map = {cell.value: idx+1 for idx, cell in enumerate(ws[1])}
    email_col = col_map.get('Email')
    status_col = col_map.get('Status')
    post_url_col = col_map.get('PostURL')
    
    if not email_col or not status_col:
        logger.error("Excel schema missing required columns.")
        return
        
    pending_rows = []
    for row in range(2, ws.max_row + 1):
        status = (ws.cell(row=row, column=status_col).value or '').strip().lower()
        if status in ('sent', 'skipped'):
            continue
        email = ws.cell(row=row, column=email_col).value
        if not email:
            continue
        post_url = ws.cell(row=row, column=post_url_col).value if post_url_col else ''
        pending_rows.append((row, email, post_url or ''))
        
    user_conf = get_selected_user_config()
    email_scraper_conf = user_conf.get("email_scraper", {})
    max_emails = int(email_scraper_conf.get("max_emails_per_run") or 5)
    emails_sent = 0

    if not pending_rows:
        logger.info(f"No pending or unsent emails found in '{JOB_TRACKER_FILE}'. All scraped email records are already sent or the file is empty.")
        return
        
    logger.info(f"Found {len(pending_rows)} pending emails to process. Target limit: {max_emails} per run.")
    for row, email, post_url in pending_rows:
        if emails_sent >= max_emails:
            logger.info(f"Target limit of {max_emails} emails sent reached for this run. Stopping.")
            break
            
        logger.info(f"Processing email outreach for {email} (row {row})")
        sent = send_email_via_gmail(scraper.driver, email, post_url=post_url, review_mode=review_mode)
        if sent == "skipped":
            logger.info(f"Email to {email} skipped – updating status to 'skipped'.")
            update_status(email, 'skipped')
            continue
        elif sent == "quit":
            logger.info("Quitting email outreach pipeline as requested by user.")
            break
        elif sent:
            update_status(email, 'sent')
            emails_sent += 1
        else:
            logger.info(f"Email to {email} not sent – leaving status unchanged.")


def run_pipeline(phase="full", review_mode=None):
    """Orchestrates the entire Email Outreach pipeline process."""
    logger.info(f"Starting Email Scraper & Outreach Pipeline (Phase: {phase})")
    
    driver = get_driver()
    scraper = LinkedInScraper(driver)
    
    try:
        if not scraper.login():
            logger.error("Login failed – aborting Email Outreach pipeline.")
            return False
            
        if phase in ("full", "phase1"):
            logger.info("Executing Phase 1: Post email scraping...")
            run_phase_one(scraper)
            
        if phase in ("full", "phase2"):
            logger.info("Executing Phase 2: Email sending...")
            run_phase_two(scraper, review_mode=review_mode)
            
        return True
    except ScraperTargetReached as e:
        logger.info(f"Pipeline stopped: {e}")
        return True
    except Exception as e:
        logger.exception(f"Unhandled error during pipeline execution: {e}")
        return False
    finally:
        logger.info("Closing browser session...")
        driver.quit()
        logger.info("Email Scraper & Outreach pipeline complete.")
        
if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description="Run the Email Scraper and Outreach pipeline.")
    parser.add_argument("--phase", choices=["full", "phase1", "phase2"], default="full")
    parser.add_argument("--review", action="store_true", help="Review before sending emails")
    args = parser.parse_args()
    run_pipeline(phase=args.phase, review_mode=args.review)
